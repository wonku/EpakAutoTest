# -*- coding: utf-8 -*-
"""
Generate Excel + XMind from a local Markdown spec (paste Feishu wiki content).

Feishu wiki URL (cannot be fetched without login):
https://tvd6quau8vr.feishu.cn/wiki/QHcZwLiGOiSK4rkBtBacbPBRnSf

Convention:
  # Module title
  - requirement bullet 1
  - requirement bullet 2
"""
from __future__ import annotations

import json
import re
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
SPEC_PATH = ROOT / "specs" / "Feishu_QHcZwLiGOiSK4rkBtBacbPBRnSf.md"
XLSX_PATH = ROOT / "Feishu_QHcZwLiGOiSK4rkBtBacbPBRnSf_TestCases.xlsx"
XMIND_PATH = ROOT / "Feishu_QHcZwLiGOiSK4rkBtBacbPBRnSf_MindMap.xmind"
DOC_URL = "https://tvd6quau8vr.feishu.cn/wiki/QHcZwLiGOiSK4rkBtBacbPBRnSf"


def tid() -> str:
    return str(uuid.uuid4()).replace("-", "")[:16]


def parse_spec(text: str) -> list[dict]:
    """Parse # headers and - bullets into modules."""
    modules: list[dict] = []
    current: dict | None = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("<!--"):
            continue
        if line.startswith("#"):
            title = line.lstrip("#").strip()
            if title.startswith("【") and "示例" in title:
                pass  # keep as normal module
            current = {"title": title, "items": []}
            modules.append(current)
            continue
        if line.startswith(("- ", "* ")) and current:
            current["items"].append(line[2:].strip())
            continue
    return [m for m in modules if m["items"]]


def expand_cases(modules: list[dict]) -> list[tuple]:
    """Rows: module, case_id, title, priority, preconditions, steps, expected, notes."""
    rows = []
    seq = 0
    for mod in modules:
        mod_title = mod["title"]
        for item in mod["items"]:
            seq += 1
            base = f"{mod_title}｜{item}"
            rows.append(
                (
                    mod_title,
                    f"TC-{seq:03d}-A",
                    f"{item}-主流程成功",
                    "P0",
                    "测试环境可用；具备合法账号/测试数据；权限符合 happy path。",
                    "1. 按 PRD 准备前置数据（字典、配置、开关状态）。\n"
                    "2. 按产品设计完成完整操作路径（含必要二次确认）。\n"
                    "3. 校验页面提示、列表/详情字段、关键状态机是否与 PRD 一致。\n"
                    "4. 如涉及接口：核对请求参数、HTTP 状态码、业务错误码与返回体核心字段。\n"
                    "5. 如涉及异步任务：核对任务状态流转与最终一致性（轮询/回调/通知）。",
                    "全流程成功；核心数据落库或外部系统一致；无未处理前端报错；审计/日志符合约定（如有）。",
                    f"需求条目：{item}",
                )
            )
            rows.append(
                (
                    mod_title,
                    f"TC-{seq:03d}-B",
                    f"{item}-异常与边界校验",
                    "P1",
                    "同主流程环境；可构造非法/边界输入。",
                    "1. 逐项构造：缺少必填、格式错误、超长文本、特殊字符、边界数值、非法枚举。\n"
                    "2. 触发提交/保存/导入等动作。\n"
                    "3. 观察前端校验与后端返回（不得静默失败）。\n"
                    "4. 确认数据库与外部系统不产生脏数据或不一致中间态。",
                    "阻断或友好提示；错误文案与 PRD/交互稿一致；无副作用扩散；必要时记录可观测日志。",
                    "对照 PRD「校验规则」表格逐条勾选。",
                )
            )
            rows.append(
                (
                    mod_title,
                    f"TC-{seq:03d}-C",
                    f"{item}-权限、菜单与按钮显隐",
                    "P1",
                    "具备多角色账号（无权限/只读/管理员等）。",
                    "1. 无权限角色访问菜单 URL 或入口。\n"
                    "2. 检查按钮级权限（列表操作、批量、导出、删除等）。\n"
                    "3. 切换有权限角色回归关键路径。\n"
                    "4. 如涉及数据权限：横向越权尝试（改 ID、改租户）。",
                    "最小权限生效：无权限不可见或明确拒绝；越权返回 403/业务拒绝且无数据泄露。",
                    base,
                )
            )
            rows.append(
                (
                    mod_title,
                    f"TC-{seq:03d}-D",
                    f"{item}-并发、重复操作与幂等",
                    "P2",
                    "可双开会话或使用脚本压测（按需）。",
                    "1. 快速连续双击提交、重复扫码回调、重复推送 webhook（如有）。\n"
                    "2. 并发修改同一资源的两条冲突路径。\n"
                    "3. 观察锁、版本号、乐观锁或业务幂等键是否生效。",
                    "不产生重复单据/重复计费/重复通知；最终状态与 PRD 收敛规则一致。",
                    "与研发确认幂等键与防重策略。",
                )
            )
            rows.append(
                (
                    mod_title,
                    f"TC-{seq:03d}-E",
                    f"{item}-兼容、性能与弱网体验",
                    "P2",
                    "主流浏览器或移动端容器；可使用节流工具模拟弱网。",
                    "1. 覆盖 Chrome / Edge / Safari（或企业规定清单）。\n"
                    "2. 关键列表分页、导出、上传在大数据量下的响应时间与超时提示。\n"
                    "3. 弱网下操作：loading、重试、断线恢复是否符合交互规范。",
                    "无明显阻塞脚本错误；超时与错误可理解；核心路径在约定性能阈值内可完成。",
                    "指标以 PRD 非功能章节为准，无则记录基线。",
                )
            )
    return rows


def build_excel(modules: list[dict], cases: list[tuple]) -> None:
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "文档与说明"
    ws_info["A1"] = "飞书 Wiki 地址"
    ws_info["B1"] = DOC_URL
    ws_info["A2"] = "说明"
    ws_info["B2"] = (
        "当前环境无法自动读取飞书正文（需登录）。请将文档全文粘贴到 specs/"
        "Feishu_QHcZwLiGOiSK4rkBtBacbPBRnSf.md 后重新运行本脚本。"
    )
    ws_info["A3"] = "解析结果"
    ws_info["B3"] = f"模块数 {len(modules)}；生成用例行数 {len(cases)}"
    for col in ("A", "B"):
        ws_info.column_dimensions[col].width = 22 if col == "A" else 88
    for r in (1, 2, 3):
        ws_info[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("测试用例")
    headers = [
        "模块",
        "用例编号",
        "用例标题",
        "优先级",
        "前置条件",
        "测试步骤",
        "预期结果",
        "备注/追溯",
    ]
    hf = PatternFill("solid", fgColor="4472C4")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for i, row in enumerate(cases, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col != 6 else 42
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)


def add_topic(attached: list, title: str, children: list[str] | None = None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    attached.append(node)
    if children:
        for ch in children:
            add_topic(node["children"]["attached"], ch)
    return node


def build_xmind(modules: list[dict]) -> None:
    root_children: list = []
    add_topic(
        root_children,
        "使用说明",
        [
            "飞书 Wiki 需登录，自动抓取不可用",
            "将正文粘贴到 specs/Feishu_QHcZwLiGOiSK4rkBtBacbPBRnSf.md",
            "运行 python generate_feishu_wiki_tests.py 重新生成",
            DOC_URL,
        ],
    )
    for m in modules:
        bullets = m["items"]
        sub = [f"需求：{b}" for b in bullets[:20]]
        if len(bullets) > 20:
            sub.append(f"... 另有 {len(bullets) - 20} 条需求（见 Excel）")
        mid = add_topic(root_children, m["title"], [])
        for b in bullets:
            add_topic(
                mid["children"]["attached"],
                b,
                ["主流程 P0", "边界/异常 P1", "权限 P1", "并发幂等 P2", "兼容性能 P2"],
            )

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "Feishu Wiki 测试脑图",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "Feishu：QHcZwLiGOiSK4rkBtBacbPBRnSf（粘贴 MD 后生成）",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_children},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "generate_feishu_wiki_tests", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def main() -> None:
    text = SPEC_PATH.read_text(encoding="utf-8")
    # Strip HTML comments
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    modules = parse_spec(text)
    if not modules:
        raise SystemExit(
            f"No modules parsed from {SPEC_PATH}. "
            "Use '# 模块' headings and '- ' bullet requirements."
        )
    cases = expand_cases(modules)
    build_excel(modules, cases)
    build_xmind(modules)
    print("Wrote:", XLSX_PATH)
    print("Wrote:", XMIND_PATH)


if __name__ == "__main__":
    main()
