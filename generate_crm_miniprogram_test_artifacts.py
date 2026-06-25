# -*- coding: utf-8 -*-
"""CRM 小程序展示询价信息+线索历史信息 — 测试用例 Excel + XMind."""
import json
import shutil
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cases_crm_miniprogram_inquiry import CASES

OUT_DIR = Path(__file__).resolve().parent
TESTCASES_DIR = OUT_DIR / "testcases"
TESTCASES_DIR.mkdir(parents=True, exist_ok=True)
DOC_NAME = "CRM-小程序展示询价信息+线索历史信息.docx"
XLSX_PATH = TESTCASES_DIR / "CRM_MiniProgram_Inquiry_LeadHistory_TestCases.xlsx"
XMIND_PATH = TESTCASES_DIR / "CRM_MiniProgram_Inquiry_LeadHistory_MindMap.xmind"
ROOT_XLSX = OUT_DIR / "CRM_MiniProgram_Inquiry_LeadHistory_TestCases.xlsx"
ROOT_XMIND = OUT_DIR / "CRM_MiniProgram_Inquiry_LeadHistory_MindMap.xmind"

HEADERS = [
    "用例ID",
    "模块",
    "优先级",
    "是否阻塞",
    "首轮必测",
    "场景",
    "前置条件",
    "测试步骤",
    "预期结果",
    "实际结果",
    "备注",
    "用例状态",
]

STATUS_OPTIONS = "PASS,FAIL,BLOCK,N/A"


def tid():
    return str(uuid.uuid4()).replace("-", "")[:16]


def add_topic(parent_attached, title, children_titles=None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    if children_titles:
        for ct in children_titles:
            add_topic(node["children"]["attached"], ct)
    return node


def case_to_row(c):
    module, cid, scene, pri, blocking, first, pre, steps, expect, note = c
    return (cid, module, pri, blocking, first, scene, pre, steps, expect, "", note, "")


def _style_header(ws, headers, fill_color="4472C4"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_cases_sheet(ws, rows):
    _style_header(ws, HEADERS)
    block_fill = PatternFill("solid", fgColor="FFC7CE")
    p0_fill = PatternFill("solid", fgColor="DDEBF7")

    for i, row in enumerate(rows, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row[3] == "是":
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=i, column=col).fill = block_fill
        elif row[2] == "P0":
            for col in range(1, len(HEADERS) + 1):
                fill = ws.cell(row=i, column=col).fill
                if not fill or fill.fgColor.rgb in ("00000000", "FFFFFFFF", None):
                    ws.cell(row=i, column=col).fill = p0_fill

    widths = {
        1: 12,
        2: 24,
        3: 8,
        4: 10,
        5: 10,
        6: 30,
        7: 26,
        8: 46,
        9: 38,
        10: 18,
        11: 20,
        12: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    status_col = get_column_letter(HEADERS.index("用例状态") + 1)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "请选择 PASS、FAIL、BLOCK 或 N/A"
    dv.errorTitle = "用例状态无效"
    dv.prompt = "执行后填写：PASS / FAIL / BLOCK / N/A"
    dv.promptTitle = "用例状态"
    ws.add_data_validation(dv)
    dv.add(f"{status_col}2:{status_col}{max(len(rows) + 1, 500)}")


def build_xmind():
    root_attached = []

    m1 = add_topic(root_attached, "1 客户详情｜Tab与导航")
    add_topic(
        m1["children"]["attached"],
        "Tab入口",
        ["询价记录", "线索历史记录", "空态"],
    )

    m2 = add_topic(root_attached, "2 询价记录｜列表卡片")
    add_topic(
        m2["children"]["attached"],
        "卡片字段",
        ["单号/状态/时间", "询价类型", "人员网格", "是否成单"],
    )
    add_topic(
        m2["children"]["attached"],
        "交互与同步",
        ["查看详情", "PC新增同步", "多类型数据"],
    )

    m3 = add_topic(root_attached, "3 询价详情页")
    add_topic(
        m3["children"]["attached"],
        "区块",
        ["基本信息", "询价要求", "出厂报价卡片", "平台报价卡片"],
    )
    add_topic(
        m3["children"]["attached"],
        "流转记录",
        ["步骤卡片倒序", "退回/转单", "返回导航"],
    )

    m4 = add_topic(root_attached, "4 业务员脱敏")
    add_topic(
        m4["children"]["attached"],
        "与PC一致",
        ["列表脱敏", "详情脱敏", "角色权限"],
    )

    m5 = add_topic(root_attached, "5 数据同步")
    add_topic(
        m5["children"]["attached"],
        "一致性",
        ["状态同步", "PC编辑同步", "只读展示"],
    )

    m6 = add_topic(root_attached, "6 线索历史｜多线索卡片")
    add_topic(
        m6["children"]["attached"],
        "来源线索切换",
        ["横向滑动", "当前查看", "条数统计"],
    )
    add_topic(
        m6["children"]["attached"],
        "切换过滤",
        ["仅本线索轨迹", "单线索场景"],
    )

    m7 = add_topic(root_attached, "7 线索历史｜跟进轨迹")
    add_topic(
        m7["children"]["attached"],
        "时间轴",
        ["转换节点", "动态+活动合并倒序", "转换前后边界"],
    )
    add_topic(
        m7["children"]["attached"],
        "内容",
        ["系统动态", "附件预览", "日期分组"],
    )

    m8 = add_topic(root_attached, "8 权限与异常")
    add_topic(
        m8["children"]["attached"],
        "异常",
        ["无权限", "弱网重试", "长文本"],
    )

    m9 = add_topic(root_attached, "9 非功能")
    add_topic(
        m9["children"]["attached"],
        "兼容性能",
        ["双端适配", "大数据量"],
    )

    m10 = add_topic(root_attached, "10 端到端")
    add_topic(
        m10["children"]["attached"],
        "全链路",
        ["PC询价→小程序", "多线索历史", "脱敏+同步"],
    )

    for c in CASES:
        mod_short = c[0].split("-", 1)[-1] if "-" in c[0] else c[0]
        title = f"{c[1]} {c[2][:28]}"
        pri_tag = f"[{c[3]}]"
        flags = []
        if c[4] == "是":
            flags.append("阻塞")
        if c[5] == "是":
            flags.append("首轮")
        flag_str = "/".join(flags) if flags else ""
        leaf = f"{pri_tag}{'('+flag_str+')' if flag_str else ''} {c[2][:24]}"
        module_num = c[0].split("-")[0]
        parent_map = {
            "1": m1,
            "2": m2,
            "3": m3,
            "4": m4,
            "5": m5,
            "6": m6,
            "7": m7,
            "8": m8,
            "9": m9,
            "10": m10,
        }
        parent = parent_map.get(module_num, m10)
        add_topic(parent["children"]["attached"], leaf)

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "CRM小程序询价线索",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "CRM 小程序展示询价信息+线索历史 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "CRMMiniprogramGenerator", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def build_excel():
    wb = Workbook()
    rows_all = [case_to_row(c) for c in CASES]

    ws_meta = wb.active
    ws_meta.title = "需求追溯"
    ws_meta["A1"], ws_meta["B1"] = "需求文档", DOC_NAME
    ws_meta["A2"], ws_meta["B2"] = "功能范围", (
        "①客户详情新增「询价记录」Tab：卡片列表+询价详情（基本信息/询价要求/出厂报价/平台报价/流转记录）；"
        "②业务员角色字段脱敏与PC一致；③新增「线索历史记录」Tab：多线索横向卡片切换+转换前动态/活动倒序时间轴"
    )
    ws_meta["A3"], ws_meta["B3"] = "首轮建议", (
        "Tab入口→询价列表→详情→脱敏→同步→多线索切换→轨迹→E2E；"
        "粉色行=阻塞场景，失败暂停并提缺陷"
    )
    ws_meta["A4"], ws_meta["B4"] = "用例状态", "下拉可选 PASS / FAIL / BLOCK / N/A"
    total = len(CASES)
    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    p0 = sum(1 for c in CASES if c[3] == "P0")
    ws_meta["A5"], ws_meta["B5"] = "统计", (
        f"合计{total}条；阻塞{block}条；首轮必测{first}条；P0={p0}条"
    )
    ws_meta.column_dimensions["A"].width = 14
    ws_meta.column_dimensions["B"].width = 88
    for r in range(1, 6):
        ws_meta[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws_all = wb.create_sheet("测试用例")
    _write_cases_sheet(ws_all, rows_all)

    ws_smoke = wb.create_sheet("首轮必测")
    _write_cases_sheet(ws_smoke, [r for r in rows_all if r[4] == "是"])

    ws_block_cases = wb.create_sheet("阻塞用例")
    _write_cases_sheet(ws_block_cases, [r for r in rows_all if r[3] == "是"])

    def rows_for_module(prefix):
        return [case_to_row(c) for c in CASES if c[0].startswith(prefix)]

    for sheet_name, prefix in [
        ("1-Tab导航", "1-"),
        ("2-询价列表", "2-"),
        ("3-询价详情", "3-"),
        ("4-脱敏", "4-"),
        ("5-数据同步", "5-"),
        ("6-多线索", "6-"),
        ("7-跟进轨迹", "7-"),
        ("8-权限异常", "8-"),
        ("9-非功能", "9-"),
        ("10-端到端", "10-"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_cases_sheet(ws, rows_for_module(prefix))

    ws_block = wb.create_sheet("阻塞场景清单")
    block_headers = ["用例ID", "模块", "场景", "优先级", "阻塞说明", "失败影响"]
    _style_header(ws_block, block_headers, fill_color="C00000")
    block_impact = {
        "MP": "Tab不可用，后续询价/线索模块无法测试",
        "IN": "询价列表核心展示失败",
        "ID": "询价详情主流程不可用",
        "DS": "脱敏合规风险，阻塞发布",
        "SY": "数据同步错误影响业务判断",
        "LH": "多线索切换不可用",
        "LT": "线索历史轨迹核心不可用",
        "PE": "权限漏洞",
        "E2E": "全链路验收失败",
    }
    for i, c in enumerate([x for x in CASES if x[4] == "是"], 2):
        prefix = c[1].split("-")[0]
        impact = block_impact.get(prefix, "阻塞关联模块测试")
        vals = [c[1], c[0], c[2], c[3], c[9] or c[2], impact]
        for col, v in enumerate(vals, 1):
            ws_block.cell(row=i, column=col, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    for col in range(1, 7):
        ws_block.column_dimensions[get_column_letter(col)].width = 22 if col >= 5 else 18
    ws_block.freeze_panes = "A2"

    ws_pri = wb.create_sheet("优先级说明")
    pri_rows = [
        ["字段", "定义", "说明"],
        ["P0", "核心功能/主流程/数据正确性", "必须首轮全部执行"],
        ["P1", "重要分支、同步、交互细节", "首轮时间允许则执行"],
        ["P2", "边界、性能、UI细节", "回归轮次执行"],
        ["是否阻塞=是", "失败会导致后续模块测试无意义或合规风险", "失败即停；见「阻塞场景清单」"],
        ["首轮必测=是", "建议第一遍测试必须执行", "见Sheet「首轮必测」"],
        ["用例状态", "PASS/FAIL/BLOCK/N/A", "BLOCK=环境/依赖阻塞无法测；数据验证用下拉选择"],
    ]
    for r, row in enumerate(pri_rows, 1):
        for col, v in enumerate(row, 1):
            ws_pri.cell(row=r, column=col, value=v)
    ws_pri.column_dimensions["A"].width = 18
    ws_pri.column_dimensions["B"].width = 28
    ws_pri.column_dimensions["C"].width = 48

    ws_cov = wb.create_sheet("需求覆盖")
    cov_headers = ["需求点", "对应用例ID", "覆盖说明"]
    _style_header(ws_cov, cov_headers, fill_color="548235")
    coverage = [
        ("需求背景-PC数据同步至小程序", "SY-001~003, E2E-001~003", "询价与线索历史展示同步"),
        ("Tab1 询价记录-卡片列表", "IN-001~007", "卡片字段、详情入口、同步"),
        ("Tab1 询价详情页", "ID-001~009", "各区块+流转记录步骤卡片"),
        ("业务员脱敏与PC一致", "DS-001~003", "列表/详情/角色"),
        ("Tab2 线索历史-多线索卡片", "LH-001~005", "横向切换、字段、过滤"),
        ("Tab2 转换前轨迹倒序合并", "LT-001~006", "动态+活动、边界、附件"),
        ("Tab入口与空态", "MP-001~004", "导航、无数据、非线索客户"),
        ("权限与异常", "PE-001~003", "无权限、弱网、长文本"),
        ("非功能", "NF-001~002", "适配、性能"),
    ]
    for i, row in enumerate(coverage, 2):
        for col, v in enumerate(row, 1):
            ws_cov.cell(row=i, column=col, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws_cov.column_dimensions["A"].width = 32
    ws_cov.column_dimensions["B"].width = 22
    ws_cov.column_dimensions["C"].width = 40
    ws_cov.freeze_panes = "A2"

    wb.save(XLSX_PATH)


def main():
    build_excel()
    build_xmind()
    for src, dst in ((XLSX_PATH, ROOT_XLSX), (XMIND_PATH, ROOT_XMIND)):
        try:
            shutil.copy2(src, dst)
        except OSError as e:
            print(f"skip copy {dst}: {e}")

    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    print(f"Generated {len(CASES)} cases (阻塞={block}, 首轮必测={first})")
    print(" ", XLSX_PATH)
    print(" ", ROOT_XLSX)
    print(" ", XMIND_PATH)
    print(" ", ROOT_XMIND)


if __name__ == "__main__":
    main()
