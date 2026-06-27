# -*- coding: utf-8 -*-
"""中文商城-支持微信注册 — 测试用例 Excel + XMind."""
import json
import shutil
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cases_esbao_wechat_register import CASES

OUT_DIR = Path(__file__).resolve().parent
TESTCASES_DIR = OUT_DIR / "testcases"
TESTCASES_DIR.mkdir(parents=True, exist_ok=True)
DOC_NAME = "中文商城-支持微信注册(1).docx"
XLSX_PATH = TESTCASES_DIR / "Esbao_WeChat_Register_TestCases.xlsx"
XMIND_PATH = TESTCASES_DIR / "Esbao_WeChat_Register_TestMindMap.xmind"
ROOT_XLSX = OUT_DIR / "Esbao_WeChat_Register_TestCases.xlsx"
ROOT_XMIND = OUT_DIR / "Esbao_WeChat_Register_TestMindMap.xmind"

HEADERS = [
    "用例ID",
    "模块",
    "优先级",
    "场景",
    "前置条件",
    "测试步骤",
    "预期结果",
    "实际结果",
    "备注",
    "用例状态",
]

STATUS_OPTIONS = "PASS,FAIL,BLOCK,N/A"


def tid():
    return str(uuid.uuid4()).replace("-", "")[:16]


def add_topic(parent_attached, title, children_titles=None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    if children_titles:
        for ct in children_titles:
            add_topic(node["children"]["attached"], ct)
    return node


def case_to_row(c):
    module, cid, scene, pri, pre, steps, expect, note = c
    return (cid, module, pri, scene, pre, steps, expect, "", note, "")


def _style_header(ws, headers, fill_color="4472C4"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_cases_sheet(ws, rows):
    _style_header(ws, HEADERS)
    p0_fill = PatternFill("solid", fgColor="DDEBF7")

    for i, row in enumerate(rows, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row[2] == "P0":
            for col in range(1, len(HEADERS) + 1):
                fill = ws.cell(row=i, column=col).fill
                if not fill or fill.fgColor.rgb in ("00000000", "FFFFFFFF", None):
                    ws.cell(row=i, column=col).fill = p0_fill

    widths = {
        1: 12,
        2: 22,
        3: 8,
        4: 32,
        5: 28,
        6: 46,
        7: 38,
        8: 18,
        9: 20,
        10: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    status_col = get_column_letter(HEADERS.index("用例状态") + 1)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "请选择 PASS、FAIL、BLOCK 或 N/A"
    dv.errorTitle = "用例状态无效"
    dv.prompt = "执行后填写：PASS / FAIL / BLOCK / N/A"
    dv.promptTitle = "用例状态"
    ws.add_data_validation(dv)
    dv.add(f"{status_col}2:{status_col}{max(len(rows) + 1, 500)}")


def build_xmind():
    root_attached = []

    m1 = add_topic(root_attached, "1 商城登录页｜微信入口")
    add_topic(
        m1["children"]["attached"],
        "入口与二维码",
        ["微信登录展示", "UI 优化", "点击出码", "qrconnect/WxLogin", "过期刷新"],
    )

    m2 = add_topic(root_attached, "2 已绑定微信｜直接登录")
    add_topic(
        m2["children"]["attached"],
        "快捷登录",
        ["扫码直登", "跳转目标页", "code/state 回调"],
    )

    m3 = add_topic(root_attached, "3 绑定手机号页")
    add_topic(
        m3["children"]["attached"],
        "必填校验",
        ["进入绑定页", "手机号验证码"],
    )
    add_topic(
        m3["children"]["attached"],
        "企业名称",
        ["非必填", "企查查下拉", "有效校验", "无效红字", "名称不一致 toast"],
    )
    add_topic(
        m3["children"]["attached"],
        "邮箱",
        ["非必填", "格式有效", "格式无效红字", "多字段同时无效"],
    )
    add_topic(
        m3["children"]["attached"],
        "提交与中断",
        ["完成注册并登录", "建立绑定", "中断重扫", "手机号占用"],
    )

    m4 = add_topic(root_attached, "4 注册结果页")
    add_topic(
        m4["children"]["attached"],
        "0518 引导",
        ["结果页展示", "完善资料", "前往首页"],
    )

    m5 = add_topic(root_attached, "5 商城异常边界")
    add_topic(
        m5["children"]["attached"],
        "异常",
        ["取消授权", "非法回调", "重复绑定幂等"],
    )

    m6 = add_topic(root_attached, "6 供应商｜跳转路径")
    add_topic(
        m6["children"]["attached"],
        "路径独立",
        ["双登录页独立", "注册返回供应商", "忘记密码返回", "商城回归"],
    )

    m7 = add_topic(root_attached, "7 供应商｜注册页")
    add_topic(
        m7["children"]["attached"],
        "按钮与流程",
        ["返回登录文案", "完成注册文案", "注册自动登录", "UI  redesign", "必填校验"],
    )

    m8 = add_topic(root_attached, "8 端到端")
    add_topic(
        m8["children"]["attached"],
        "全链路",
        ["新用户微信注册", "老用户微信登录", "供应商与商城隔离"],
    )

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "微信注册",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "中文商城-支持微信注册 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "EsbaoWeChatRegisterGenerator", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def build_excel():
    wb = Workbook()
    rows_all = [case_to_row(c) for c in CASES]

    ws_meta = wb.active
    ws_meta.title = "需求追溯"
    ws_meta["A1"], ws_meta["B1"] = "需求文档", DOC_NAME
    ws_meta["A2"], ws_meta["B2"] = "功能范围", (
        "商城端：登录页新增微信扫码登录（已绑定直登/首次绑定手机号）；"
        "企业名称企查查联想与校验、邮箱格式校验、完成注册并登录建立绑定、注册结果页引导；"
        "供应商端：登录/注册/忘记密码返回路径独立、按钮文案调整、注册成功自动登录、UI 优化"
    )
    ws_meta["A3"], ws_meta["B3"] = "UI 稿", "https://lanhuapp.com/link/#/invite?sid=lX0QPsUR"
    ws_meta["A4"], ws_meta["B4"] = "用例状态", "下拉可选 PASS / FAIL / BLOCK / N/A"
    total = len(CASES)
    p0 = sum(1 for c in CASES if c[3] == "P0")
    ws_meta["A5"], ws_meta["B5"] = "统计", f"合计{total}条；P0={p0}条"
    ws_meta.column_dimensions["A"].width = 14
    ws_meta.column_dimensions["B"].width = 88
    for r in range(1, 6):
        ws_meta[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws_all = wb.create_sheet("测试用例")
    _write_cases_sheet(ws_all, rows_all)

    ws_p0 = wb.create_sheet("P0用例")
    _write_cases_sheet(ws_p0, [r for r in rows_all if r[2] == "P0"])

    def rows_for_prefix(prefix):
        return [case_to_row(c) for c in CASES if c[0].startswith(prefix)]

    for sheet_name, prefix in [
        ("1-商城登录页", "1-"),
        ("2-已绑定登录", "2-"),
        ("3-绑定手机号", "3-"),
        ("4-注册结果页", "4-"),
        ("5-商城异常", "5-"),
        ("6-供应商跳转", "6-"),
        ("7-供应商注册", "7-"),
        ("8-端到端", "8-"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_cases_sheet(ws, rows_for_prefix(prefix))

    ws_cov = wb.create_sheet("需求覆盖")
    cov_headers = ["需求点", "对应用例ID", "覆盖说明"]
    _style_header(ws_cov, cov_headers, fill_color="548235")
    coverage = [
        ("商城-微信登录入口", "WR-M-001~005", "入口、UI、二维码、实现方式、过期刷新"),
        ("商城-已绑定直登", "WR-M-006~008", "扫码登录、跳转、回调校验"),
        ("商城-绑定手机号", "WR-M-009~022", "首次绑定、必填/选填校验、企查查、错误提示、绑定关系"),
        ("商城-注册结果页", "WR-M-023~025", "结果页、完善资料、前往首页"),
        ("商城-异常边界", "WR-M-026~028", "取消授权、非法回调、幂等"),
        ("供应商-跳转独立", "WR-S-001~004", "双登录独立、返回路径修复、商城回归"),
        ("供应商-注册改造", "WR-S-005~009", "按钮文案、自动登录、UI、校验"),
        ("端到端", "WR-E-001~003", "新用户/老用户/双端隔离"),
    ]
    for i, row in enumerate(coverage, 2):
        for col, v in enumerate(row, 1):
            ws_cov.cell(row=i, column=col, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws_cov.column_dimensions["A"].width = 28
    ws_cov.column_dimensions["B"].width = 22
    ws_cov.column_dimensions["C"].width = 44
    ws_cov.freeze_panes = "A2"

    ws_pri = wb.create_sheet("优先级说明")
    ws_pri.append(["字段", "定义", "说明"])
    ws_pri.append(["P0", "核心主流程/数据正确性", "微信登录、绑定、供应商路径修复等必测"])
    ws_pri.append(["P1", "重要分支与 UI/校验细节", "企查查、toast、按钮文案等"])
    ws_pri.append(["P2", "边界与异常模拟", "回调篡改、幂等等"])
    ws_pri.append(["用例状态", "PASS/FAIL/BLOCK/N/A", "BLOCK=环境/依赖阻塞无法测"])

    wb.save(XLSX_PATH)


def main():
    build_excel()
    build_xmind()
    for src, dst in ((XLSX_PATH, ROOT_XLSX), (XMIND_PATH, ROOT_XMIND)):
        try:
            shutil.copy2(src, dst)
        except OSError as e:
            print(f"skip copy {dst}: {e}")

    p0 = sum(1 for c in CASES if c[3] == "P0")
    print(f"Generated {len(CASES)} cases (P0={p0})")
    print(" ", XLSX_PATH)
    print(" ", ROOT_XLSX)
    print(" ", XMIND_PATH)
    print(" ", ROOT_XMIND)


if __name__ == "__main__":
    main()
