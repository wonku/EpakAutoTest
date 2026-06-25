# -*- coding: utf-8 -*-
"""CRM3.13 优化需求 — 测试用例 Excel + XMind."""
import json
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cases_crm313 import CASES

OUT_DIR = Path(__file__).resolve().parent
TESTCASES_DIR = OUT_DIR / "testcases"
TESTCASES_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = TESTCASES_DIR / "CRM3.13_Optimization_TestCases.xlsx"
XMIND_PATH = TESTCASES_DIR / "CRM3.13_Optimization_TestMindMap.xmind"
ROOT_XLSX = OUT_DIR / "CRM3.13_Optimization_TestCases.xlsx"
ROOT_XMIND = OUT_DIR / "CRM3.13_Optimization_TestMindMap.xmind"

HEADERS = [
    "用例ID",
    "模块",
    "优先级",
    "是否阻塞",
    "首轮必测",
    "场景",
    "前置条件",
    "测试步骤",
    "预期结果",
    "实际结果",
    "备注",
    "用例状态",
]

STATUS_OPTIONS = "PASS,FAIL,BLOCK,N/A"


def tid():
    return str(uuid.uuid4()).replace("-", "")[:16]


def add_topic(parent_attached, title, children_titles=None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    if children_titles:
        for ct in children_titles:
            add_topic(node["children"]["attached"], ct)
    return node


def case_to_row(c):
    module, cid, scene, pri, blocking, first, pre, steps, expect, note = c
    return (cid, module, pri, blocking, first, scene, pre, steps, expect, "", note, "")


def _style_header(ws, headers, fill_color="4472C4"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _write_cases_sheet(ws, rows):
    _style_header(ws, HEADERS)
    block_fill = PatternFill("solid", fgColor="FFC7CE")
    p0_fill = PatternFill("solid", fgColor="DDEBF7")

    for i, row in enumerate(rows, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if row[3] == "是":
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=i, column=col).fill = block_fill
        elif row[2] == "P0":
            for col in range(1, len(HEADERS) + 1):
                if ws.cell(row=i, column=col).fill.fgColor.rgb in ("00000000", "FFFFFFFF", None):
                    ws.cell(row=i, column=col).fill = p0_fill

    widths = {
        1: 12,
        2: 22,
        3: 8,
        4: 10,
        5: 10,
        6: 28,
        7: 24,
        8: 44,
        9: 36,
        10: 18,
        11: 18,
        12: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    status_col = get_column_letter(HEADERS.index("用例状态") + 1)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "请选择 PASS、FAIL、BLOCK 或 N/A"
    dv.errorTitle = "用例状态无效"
    dv.prompt = "执行后填写：PASS / FAIL / BLOCK / N/A"
    dv.promptTitle = "用例状态"
    ws.add_data_validation(dv)
    dv.add(f"{status_col}2:{status_col}{max(len(rows) + 1, 500)}")


def build_xmind():
    root_attached = []

    m1 = add_topic(root_attached, "1 外呼手机号解绑")
    add_topic(
        m1["children"]["attached"],
        "解绑按钮展示",
        ["已绑定展示", "未绑定不展示"],
    )
    add_topic(
        m1["children"]["attached"],
        "解绑流程",
        ["二次确认文案", "取消", "确认解绑", "重新绑定"],
    )

    m2 = add_topic(root_attached, "2 外呼未接通活动记录")
    add_topic(
        m2["children"]["attached"],
        "自动创建",
        ["未接通触发", "名称格式", "备注文案", "挂载对象"],
    )
    add_topic(m2["children"]["attached"], "异常", ["接通不生成未接通", "幂等"])

    m3 = add_topic(root_attached, "3 客户活动记录与跟进")
    add_topic(
        m3["children"]["attached"],
        "关联对象筛选",
        ["子Tab", "来自列", "直接跟进", "联系人动态", "销售机会动态"],
    )
    add_topic(
        m3["children"]["attached"],
        "子对象同步",
        ["联系人活动展示", "销售机会活动展示", "最新活动时间", "回收逻辑"],
    )
    add_topic(
        m3["children"]["attached"],
        "图片上传",
        ["Ctrl+V粘贴", "拖拽", "删除"],
    )

    m4 = add_topic(root_attached, "4 拜访记录")
    add_topic(m4["children"]["attached"], "状态样式", ["已过期灰色"])
    add_topic(
        m4["children"]["attached"],
        "关联活动",
        ["按钮展示条件", "绑定弹窗", "权限过滤", "已关联过滤", "完成状态"],
    )
    add_topic(
        m4["children"]["attached"],
        "已完成操作",
        ["查看记录", "解绑确认", "状态回退", "按钮权限"],
    )

    m5 = add_topic(root_attached, "5 无活动记录筛选")
    add_topic(
        m5["children"]["attached"],
        "客户列表",
        ["时间筛选", "无活动勾选", "置灰", "结果准确", "子对象计入"],
    )
    add_topic(
        m5["children"]["attached"],
        "线索列表",
        ["时间筛选", "无活动勾选", "与客户一致"],
    )

    m6 = add_topic(root_attached, "6 联系人职务")
    add_topic(
        m6["children"]["attached"],
        "筛选",
        ["多选下拉", "枚举字典", "单选/多选", "组合筛选"],
    )
    add_topic(m6["children"]["attached"], "列表展示", ["职务列", "空值"])

    m7 = add_topic(root_attached, "7 端到端")
    add_topic(
        m7["children"]["attached"],
        "全链路",
        ["外呼→活动→来源", "拜访关联解绑", "筛选与时间一致"],
    )

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "CRM3.13",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "CRM3.13 优化需求 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "CRM313Generator", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def build_excel():
    wb = Workbook()
    rows_all = [case_to_row(c) for c in CASES]

    ws_meta = wb.active
    ws_meta.title = "需求追溯"
    ws_meta["A1"], ws_meta["B1"] = "需求文档", "CRM3.13优化需求.docx"
    ws_meta["A2"], ws_meta["B2"] = "功能范围", (
        "①外呼手机号解绑；②外呼未接通自动创建活动记录；"
        "③客户联系人/销售机会活动展示在客户跟进记录（含子Tab筛选、来自列、回收逻辑、PC图片粘贴/拖拽）；"
        "④拜访记录（已过期灰色、关联/解绑活动、查看记录及权限）；"
        "⑤线索/客户无活动记录快速筛选；⑥联系人职务筛选与列表展示"
    )
    ws_meta["A3"], ws_meta["B3"] = "首轮建议", (
        "1外呼解绑→2外呼活动→3客户活动/图片→4拜访→5筛选→6联系人→7 E2E；"
        "阻塞项失败暂停并提缺陷；粉色行=阻塞场景"
    )
    ws_meta["A4"], ws_meta["B4"] = "用例状态", "下拉可选 PASS / FAIL / BLOCK / N/A"
    total = len(CASES)
    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    p0 = sum(1 for c in CASES if c[3] == "P0")
    ws_meta["A5"], ws_meta["B5"] = "统计", (
        f"合计{total}条；阻塞{block}条；首轮必测{first}条；P0={p0}条"
    )
    ws_meta.column_dimensions["A"].width = 14
    ws_meta.column_dimensions["B"].width = 88
    for r in range(1, 6):
        ws_meta[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws_all = wb.create_sheet("测试用例")
    _write_cases_sheet(ws_all, rows_all)

    ws_smoke = wb.create_sheet("首轮冒烟")
    _write_cases_sheet(ws_smoke, [r for r in rows_all if r[4] == "是"])

    def rows_for_prefix(prefix):
        return [case_to_row(c) for c in CASES if c[0].startswith(prefix)]

    for sheet_name, prefix in [
        ("1-外呼解绑", "1-"),
        ("2-外呼活动", "2-"),
        ("3-客户活动", "3-"),
        ("4-拜访记录", "4-"),
        ("5-无活动筛选", "5-"),
        ("6-联系人职务", "6-"),
        ("7-端到端", "7-"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_cases_sheet(ws, rows_for_prefix(prefix))

    ws_block = wb.create_sheet("阻塞场景清单")
    block_headers = ["用例ID", "模块", "场景", "优先级", "阻塞说明", "失败影响"]
    _style_header(ws_block, block_headers, fill_color="C00000")
    block_impact = {
        "OB": "外呼解绑不可用，影响外呼账号管理",
        "OC": "未接通活动缺失，影响跟进完整性",
        "AR": "客户活动展示/回收/图片上传异常，影响核心跟进",
        "VR": "拜访关联闭环不可用",
        "FIL": "无活动记录筛选不可用",
        "CT": "联系人职务筛选不可用",
        "E2E": "全链路验收失败",
    }
    for i, c in enumerate([x for x in CASES if x[4] == "是"], 2):
        prefix = c[1].split("-")[0]
        impact = block_impact.get(prefix, "阻塞关联模块测试")
        vals = [c[1], c[0], c[2], c[3], c[9] or c[2], impact]
        for col, v in enumerate(vals, 1):
            ws_block.cell(row=i, column=col, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    for col in range(1, 7):
        ws_block.column_dimensions[get_column_letter(col)].width = 22 if col >= 5 else 18
    ws_block.freeze_panes = "A2"

    ws_pri = wb.create_sheet("优先级说明")
    ws_pri.append(["字段", "定义", "说明"])
    ws_pri.append(["P0", "核心功能/主流程/数据正确性", "必须首轮全部执行"])
    ws_pri.append(["P1", "重要分支、权限、筛选、补偿逻辑", "首轮时间允许则执行"])
    ws_pri.append(["P2", "边界、UI细节、非关键异常", "回归轮次执行"])
    ws_pri.append(["是否阻塞=是", "失败会导致后续模块测试无意义", "失败即停，修复后从该模块重测"])
    ws_pri.append(["首轮必测=是", "建议第一遍测试必须执行", "见Sheet「首轮冒烟」"])
    ws_pri.append(["用例状态", "PASS/FAIL/BLOCK/N/A", "BLOCK=环境/依赖阻塞无法测"])

    ws_cov = wb.create_sheet("需求覆盖")
    cov_headers = ["需求点", "对应用例ID", "覆盖说明"]
    _style_header(ws_cov, cov_headers, fill_color="548235")
    coverage = [
        ("1 外呼手机号解绑", "OB-001~008", "按钮、确认、解绑、重绑、权限"),
        ("2 外呼未接通活动记录", "OC-001~007", "自动创建、命名、备注、挂载"),
        ("3.1 关联对象子Tab筛选", "AR-001~002", "子Tab及筛选项"),
        ("3.2 来自（关联对象）列", "AR-003~008", "三来源展示及同步"),
        ("3.3 客户回收/最新活动时间", "AR-009~011", "子对象活动更新时间"),
        ("3.4 PC图片Ctrl+V/拖拽", "AR-012~016", "粘贴、拖拽、删除"),
        ("4.1 已过期灰色", "VR-001", "样式"),
        ("4.2 关联活动记录", "VR-002~009", "按钮、弹窗、过滤、完成"),
        ("4.3 查看记录/解绑", "VR-010~019", "详情、权限、回退"),
        ("5 客户/线索无活动筛选", "FIL-001~012", "置灰、筛选、线索新增"),
        ("6 联系人职务", "CT-001~008", "多选、枚举、列表展示"),
        ("端到端", "E2E-001~003", "跨模块联调"),
    ]
    for i, row in enumerate(coverage, 2):
        for col, v in enumerate(row, 1):
            ws_cov.cell(row=i, column=col, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws_cov.column_dimensions["A"].width = 28
    ws_cov.column_dimensions["B"].width = 18
    ws_cov.column_dimensions["C"].width = 40
    ws_cov.freeze_panes = "A2"

    wb.save(XLSX_PATH)


def main():
    build_excel()
    build_xmind()
    import shutil

    for src, dst in ((XLSX_PATH, ROOT_XLSX), (XMIND_PATH, ROOT_XMIND)):
        try:
            shutil.copy2(src, dst)
        except OSError as e:
            print(f"skip copy {dst}: {e}")

    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    print(f"Generated {len(CASES)} cases (阻塞={block}, 首轮必测={first})")
    print(" ", XLSX_PATH)
    print(" ", ROOT_XLSX)
    print(" ", XMIND_PATH)
    print(" ", ROOT_XMIND)


if __name__ == "__main__":
    main()
