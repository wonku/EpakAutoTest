"""生成 CRM 线索看板·人员模块「责任周期 + 6 种结束类型」测试数据准备 Excel。

输出：
  testcases/CRM_Lead_Dashboard_EMP_TestData.xlsx
  根目录副本：CRM_Lead_Dashboard_EMP_TestData.xlsx
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

OUT_DIR = _ROOT / "testcases"
OUT_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = OUT_DIR / "CRM_Lead_Dashboard_EMP_TestData.xlsx"
ROOT_COPY = _ROOT / "CRM_Lead_Dashboard_EMP_TestData.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
WARN_FILL = PatternFill("solid", fgColor="FFE699")
HOT_FILL = PatternFill("solid", fgColor="F4B084")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, headers, fill=HEADER_FILL):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_rows(ws, rows, start_row=2):
    for ri, row in enumerate(rows, start_row):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = ALIGN_WRAP


# ---------- Sheet 1: 总览与执行指引 ----------
def build_overview(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "总览与执行指引"
    ws["A1"] = "用途"
    ws["B1"] = (
        "为「CRM 线索看板·人员处理与工作量」模块的责任周期口径与 6 种责任周期结束类型，"
        "准备一组可重复使用的测试数据；支撑 EMP-004~EMP-079 全部用例的对账与导出校验。"
    )
    ws["A2"] = "数据范围"
    ws["B2"] = (
        "1) 责任周期=员工承接某条线索从开始到结束的连续时间段；\n"
        "2) 6 种结束类型：当前持有 / 分配 / 主动移入公海 / 系统回收 / 转客户 / 作废；\n"
        "3) 覆盖与筛选时间段「完全在内 / 跨左 / 跨右 / 跨全段 / 完全在外」共 5 种交集形态；\n"
        "4) 覆盖「同一线索多次承接（多周期）」与「不同员工承接同一线索」。"
    )
    ws["A3"] = "执行原则"
    ws["B3"] = (
        "脚本仅自动化：线索创建 + 活动记录；其余（转分配/移入公海/系统回收/转客户/作废）"
        "由 QA 按场景矩阵在 CRM 后台手工触发并记录时间。脚本会生成 manifest.md 作为手工执行清单。"
    )
    ws["A4"] = "脚本入口"
    ws["B4"] = (
        "python scripts/create_lead_dashboard_emp_dataset.py [--config xxx.json] [--dry-run]\n"
        "  --config       自定义员工与场景，未提供则使用默认 15 个场景；\n"
        "  --dry-run      只输出计划，不发起接口请求；\n"
        "  --tag <name>   指定输出目录名，默认时间戳。"
    )
    ws["A5"] = "输出产物"
    ws["B5"] = (
        "testcases/_emp_dataset/<tag>/lead_ids.json   全部新建线索的 ID 与归属\n"
        "testcases/_emp_dataset/<tag>/manifest.md     手工触发清单（按场景顺序）\n"
        "testcases/_emp_dataset/<tag>/run.log         执行日志"
    )
    ws["A6"] = "依赖账号"
    ws["B6"] = (
        "至少 3 个销售员工账号（甲/乙/丙），共属于同一管理员可见范围；\n"
        "至少 1 个特殊员工/全量权限账号（用于基准对账）；\n"
        "1 个无导出权限账号（验证 EMP-037）。"
    )
    ws["A7"] = "关键说明"
    ws["B7"] = (
        "1) 涉及历史时间的场景（back_days_for_assign / void_after_days）需要 DB 写入或测试库辅助接口，脚本仅记录"
        "需求；\n"
        "2) 系统回收依赖定时任务；建议测试环境提供「立即触发回收」运维接口；\n"
        "3) 导出对账以 c:\\Users\\caitiantian\\Downloads\\线索-人员看板导出.xlsx Sheet2 为字段基准。"
    )
    ws["A8"] = "对账参考"
    ws["B8"] = (
        "本 Excel Sheet「Sheet4 导出样例（28列）」与导出模板 Sheet2 列头一致；"
        "Sheet「Sheet5 字段映射」记录每列与 EMP 用例的对应关系。"
    )
    for r in range(1, 9):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].fill = SECTION_FILL
        ws[f"A{r}"].alignment = ALIGN_WRAP
        ws[f"B{r}"].alignment = ALIGN_WRAP
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 92


# ---------- Sheet 2: 测试数据准备清单 ----------
def build_checklist(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet2 数据准备清单")
    headers = ["编号", "类型", "名称/说明", "数量/要求", "建议来源", "对应用例", "状态"]
    _style_header(ws, headers)
    rows = [
        # 账号
        ("ACC-01", "账号", "全量看板权限账号（特殊员工 / 18143073992）", "1", "管理员后台配置", "MNU-004/005, DAT-001", "☐"),
        ("ACC-02", "账号", "销售员工·甲（基础归属）", "1", "测试环境创建", "EMP-002~014", "☐"),
        ("ACC-03", "账号", "销售员工·乙（转出目标）", "1", "测试环境创建", "EMP-003/EMP-057/EMP-055", "☐"),
        ("ACC-04", "账号", "销售员工·丙（多周期复用）", "1", "测试环境创建", "EMP-054/EMP-055", "☐"),
        ("ACC-05", "账号", "无导出权限账号", "1", "权限不配置「人员明细导出」", "EMP-037", "☐"),
        ("ACC-06", "账号", "管理者·部门负责人（部分员工权限）", "1", "测试环境创建", "EMP-076", "☐"),
        # 组织
        ("ORG-01", "组织", "员工归属部门设置（甲/乙/丙各属不同部门）", "3部门", "测试库", "明细表所属部门", "☐"),
        # 线索批次
        ("LEAD-01", "线索", "默认场景批 15 条（脚本造数）", "≥15", "API 脚本", "EMP-051~062", "☐"),
        ("LEAD-02", "线索", "历史时间段补造 2 条（DB 改写 dispatch_time）", "2", "DB 写入", "EMP-052/053", "☐"),
        ("LEAD-03", "线索", "数据量线索 ≥1000 用于性能与导出（可选）", "≥1000", "API/工具批量", "EMP-079", "☐"),
        # 活动
        ("ACT-01", "活动", "电话/微信/拜访/邮件四类活动（每类≥10条）", "≥40", "API 脚本", "EMP-005/034/070", "☐"),
        ("ACT-02", "活动", "系统自动活动（如自动外呼）", "若干", "测试库自动", "EMP-005 排除自动", "☐"),
        # 责任周期 6 类
        ("CYC-01", "责任周期·当前持有", "甲名下保持持有（脚本自动）", "≥2", "API 脚本", "EMP-056", "☐"),
        ("CYC-02", "责任周期·分配", "甲→乙 转分配（CRM 后台）", "≥2", "手工触发", "EMP-057", "☐"),
        ("CYC-03", "责任周期·主动移入公海", "甲主动移入公海", "≥1", "手工触发", "EMP-058", "☐"),
        ("CYC-04", "责任周期·系统回收", "超期未跟进系统回收（定时/运维接口）", "≥1", "运维触发", "EMP-059", "☐"),
        ("CYC-05", "责任周期·转客户", "甲名下转客户", "≥1", "手工触发", "EMP-060", "☐"),
        ("CYC-06", "责任周期·作废", "甲名下作废", "≥1", "手工触发", "EMP-061", "☐"),
        # 多周期 / 多人
        ("MUL-01", "多周期", "同一线索丙→甲→丙（覆盖多次计入）", "1", "脚本+手工", "EMP-003/054", "☐"),
        ("MUL-02", "多人承接", "同一线索甲→乙→丙（覆盖多人各一条）", "1", "脚本+手工", "EMP-055", "☐"),
        # 基准
        ("BASE-01", "基准对账表", "默认条件 + 5 组组合条件的 SQL/BI 导出", "≥6组", "开发/BI", "EMP-002~009/074", "☐"),
        ("BASE-02", "基准对账表", "导出明细 28 列基准（单员工 + 全部员工）", "2份", "开发/BI", "EMP-063~073", "☐"),
        ("BASE-03", "时区/工时", "work_hours_between 节假日表与工作日定义", "1份", "开发", "EMP-015/069", "☐"),
    ]
    _write_rows(ws, rows)
    _autosize(ws, [12, 18, 50, 18, 22, 32, 8])
    # 高亮 6 个 CYC 行
    for i, row in enumerate(rows, 2):
        if str(row[0]).startswith("CYC-"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = WARN_FILL


# ---------- Sheet 3: 责任周期场景矩阵 ----------
def build_scenario_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet3 责任周期场景矩阵")
    headers = [
        "场景ID",
        "归属员工",
        "活动数(有效人工)",
        "结束类型",
        "转出至/对方",
        "复用线索",
        "时间偏移说明",
        "与筛选时段交集",
        "对应用例",
        "自动执行步骤",
        "手工执行步骤",
        "预期结束时间字段",
    ]
    _style_header(ws, headers)
    rows = [
        ("CUR-001", "甲", 1, "当前持有", "-", "-", "-", "完全在内", "EMP-056/EMP-067", "API 建线索+1条活动", "无", "当前时间"),
        ("CUR-002", "甲", 0, "当前持有", "-", "-", "-", "完全在内", "EMP-013/067", "API 建线索", "无", "当前时间"),
        ("ASN-001", "甲", 1, "分配", "乙", "-", "-", "完全在内", "EMP-057", "API 建线索+活动", "CRM：甲名下转分配给乙", "转出时间"),
        ("ASN-002", "甲", 0, "分配", "乙", "-", "-", "完全在内", "EMP-014/057", "API 建线索", "CRM：甲名下转分配给乙", "转出时间"),
        ("MAN-POOL-001", "甲", 1, "主动移入公海", "-", "-", "-", "完全在内", "EMP-058", "API 建线索+活动", "甲在 CRM 主动移入公海", "移入时间"),
        ("AUTO-POOL-001", "甲", 0, "系统回收", "-", "-", "-", "完全在内", "EMP-059", "API 建线索", "运维触发/等待定时回收", "系统回收时间"),
        ("WIN-CUST-001", "甲", 2, "转客户", "-", "-", "-", "完全在内", "EMP-060/EMP-071", "API 建线索+2条活动", "甲在 CRM 转客户", "转客户时间"),
        ("VOID-001", "甲", 1, "作废", "-", "-", "-", "完全在内", "EMP-061", "API 建线索+活动", "甲在 CRM 作废线索", "作废时间"),
        (
            "MULTI-CYCLE-001",
            "丙",
            1,
            "分配",
            "甲",
            "-",
            "-",
            "完全在内",
            "EMP-054 段1",
            "API 建线索（归属丙）+ 活动",
            "CRM：丙名下转分配给甲",
            "转出时间",
        ),
        (
            "MULTI-CYCLE-002",
            "甲",
            1,
            "分配",
            "丙",
            "MULTI-CYCLE-001",
            "-",
            "完全在内",
            "EMP-054 段2",
            "复用线索 + 甲创建活动",
            "CRM：甲名下再次转回丙",
            "转出时间",
        ),
        (
            "MULTI-USER-001",
            "甲",
            1,
            "分配",
            "乙",
            "-",
            "-",
            "完全在内",
            "EMP-055 段1",
            "API 建线索（甲）+ 活动",
            "CRM：甲→乙",
            "转出时间",
        ),
        (
            "MULTI-USER-002",
            "乙",
            1,
            "分配",
            "丙",
            "MULTI-USER-001",
            "-",
            "完全在内",
            "EMP-055 段2",
            "复用线索 + 乙创建活动",
            "CRM：乙→丙",
            "转出时间",
        ),
        (
            "MULTI-USER-003",
            "丙",
            0,
            "当前持有",
            "-",
            "MULTI-USER-001",
            "-",
            "完全在内",
            "EMP-055 段3",
            "复用线索（丙持有）",
            "无",
            "当前时间",
        ),
        ("SPAN-001", "甲", 1, "当前持有", "-", "-", "分配时间=T-365天", "跨左（开始<筛选起点）", "EMP-052", "API 建线索+活动", "DB 改写 dispatch_time=今日-365天", "当前时间"),
        ("OUT-001", "甲", 0, "作废", "-", "-", "分配=T-400天，作废=T-380天", "完全在外", "EMP-053", "API 建线索", "DB 改写时间 + CRM 作废", "作废时间"),
    ]
    _write_rows(ws, rows)
    _autosize(ws, [16, 12, 14, 18, 14, 16, 26, 18, 22, 32, 32, 18])
    # 标记不同结束类型
    color_map = {
        "当前持有": "BDD7EE",
        "分配": "C6E0B4",
        "主动移入公海": "F8CBAD",
        "系统回收": "FFD966",
        "转客户": "B4A7D6",
        "作废": "F4B084",
    }
    for i, row in enumerate(rows, 2):
        end_type = row[3]
        c = ws.cell(row=i, column=4)
        c.fill = PatternFill("solid", fgColor=color_map.get(end_type, "FFFFFF"))


# ---------- Sheet 4: 导出样例（28列，与模板 Sheet2 一致） ----------
def build_export_sample(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet4 导出样例(28列)")
    headers = [
        "统计归属员工",
        "所属部门",
        "线索ID",
        "线索名称",
        "线索来源",
        "线索评分",
        "线索等级",
        "线索创建时间",
        "线索团队成员",
        "分配到员工时间",
        "责任周期结束时间",
        "责任周期结束类型",
        "当前跟进人",
        "是否当前仍由该员工负责",
        "该员工是否已处理",
        "该员工首次有效活动时间",
        "该员工首跟工作时长",
        "该员工活动总数",
        "电话活动数",
        "微信活动数",
        "拜访活动数",
        "邮件活动数",
        "责任期内是否转换客户",
        "转客户时间",
        "转换客户名称",
        "客户跟进人",
        "客户首次跟进时间",
        "客户最近一次跟进时间",
    ]
    _style_header(ws, headers)
    rows = [
        (
            "甲", "销售一部", "LD100001", "tinker线索-001", "展会", 68, "A",
            "2026-05-18 10:00:00", "甲,乙", "2026-05-18 10:01:30", "2026-05-20 16:11:00",
            "当前持有", "甲", "是", "是", "2026-05-18 14:20:00", "4.1h", 2, 1, 0, 1, 0,
            "否", "", "", "", "", "",
        ),
        (
            "甲", "销售一部", "LD100002", "tinker线索-002", "展会", 60, "B",
            "2026-05-18 11:00:00", "甲", "2026-05-18 11:01:00", "2026-05-19 09:10:00",
            "分配", "乙", "否", "是", "2026-05-18 16:40:00", "5.5h", 1, 1, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
        (
            "甲", "销售一部", "LD100003", "tinker线索-003", "百度", 50, "C",
            "2026-04-18 09:00:00", "甲", "2026-04-18 09:02:00", "2026-05-15 18:00:00",
            "主动移入公海", "-", "否", "否", "", "—", 0, 0, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
        (
            "甲", "销售一部", "LD100004", "tinker线索-004", "官网", 40, "D",
            "2026-04-01 12:00:00", "甲", "2026-04-01 12:00:00", "2026-05-12 03:00:00",
            "系统回收", "-", "否", "否", "", "—", 0, 0, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
        (
            "甲", "销售一部", "LD100005", "tinker线索-005", "展会", 88, "A",
            "2026-05-10 10:00:00", "甲", "2026-05-10 10:00:00", "2026-05-19 17:30:00",
            "转客户", "丁", "否", "是", "2026-05-10 11:15:00", "1.3h", 5, 3, 1, 0, 1,
            "是", "2026-05-19 17:30:00", "客户A", "丁", "2026-05-19 18:00:00", "2026-05-20 09:30:00",
        ),
        (
            "甲", "销售一部", "LD100006", "tinker线索-006", "百度", 30, "D",
            "2026-05-15 09:00:00", "甲", "2026-05-15 09:00:00", "2026-05-16 10:00:00",
            "作废", "-", "否", "是", "2026-05-15 11:00:00", "2.0h", 1, 1, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
        (
            "丙", "销售三部", "LD100007", "tinker线索-007", "展会", 70, "B",
            "2026-05-01 09:00:00", "丙,甲", "2026-05-01 09:01:00", "2026-05-08 09:00:00",
            "分配", "甲", "否", "是", "2026-05-01 10:00:00", "1.0h", 1, 1, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
        (
            "甲", "销售一部", "LD100007", "tinker线索-007", "展会", 70, "B",
            "2026-05-01 09:00:00", "丙,甲", "2026-05-08 09:00:00", "2026-05-20 16:11:00",
            "分配", "丙", "否", "是", "2026-05-08 14:00:00", "5.0h", 2, 2, 0, 0, 0,
            "否", "", "", "", "", "",
        ),
    ]
    _write_rows(ws, rows)
    widths = [14, 12, 14, 24, 12, 10, 10, 18, 16, 18, 18, 16, 14, 18, 14, 18, 14, 12, 12, 12, 12, 12, 18, 18, 16, 12, 18, 18]
    _autosize(ws, widths)
    # 高亮责任周期结束类型列
    for r in range(2, 2 + len(rows)):
        ws.cell(row=r, column=12).font = Font(bold=True)


# ---------- Sheet 5: 字段口径与用例映射 ----------
def build_field_mapping(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet5 字段映射与口径")
    headers = ["#", "字段(导出列)", "取值逻辑", "对应用例", "备注"]
    _style_header(ws, headers)
    rows = [
        (1, "统计归属员工", "承接该责任周期的员工姓名（按责任周期员工归属）", "EMP-064", "—"),
        (2, "所属部门", "员工所属部门", "EMP-026", "—"),
        (3, "线索ID / 名称 / 来源 / 评分 / 等级 / 创建时间 / 团队成员", "线索主数据", "EMP-073", "字段名以 CRM 主数据为准"),
        (4, "分配到员工时间", "该责任周期开始时间", "EMP-065", "—"),
        (5, "责任周期结束时间", "按结束类型：当前持有=当前时间；分配=转出时间；主动移入公海=移入时间；系统回收=回收时间；转客户=转客户时间；作废=作废时间", "EMP-056~061", "枚举见 Sheet3"),
        (6, "责任周期结束类型", "当前持有 / 分配 / 主动移入公海 / 系统回收 / 转客户 / 作废", "EMP-062", "枚举完备性"),
        (7, "当前跟进人", "线索当前跟进人（与统计归属员工可能不同）", "EMP-066", "—"),
        (8, "是否当前仍由该员工负责", "当前持有=是；其它=否", "EMP-067", "—"),
        (9, "该员工是否已处理", "责任周期内是否产生过有效人工活动", "EMP-068", "排除系统自动"),
        (10, "该员工首次有效活动时间", "责任周期内首次有效人工跟进时间", "EMP-068", "—"),
        (11, "该员工首跟工作时长", "已跟进=work_hours_between(开始, 首次活动)；未跟进且持有=算到当前时间；未跟进且已结束=算到结束时间", "EMP-069 / EMP-012~014", "剔除周末与法定节日"),
        (12, "该员工活动总数 / 电话 / 微信 / 拜访 / 邮件", "责任周期内分类活动数；总数=电+微+拜+邮", "EMP-070", "—"),
        (13, "责任期内是否转换客户", "在该员工责任周期内是否发生转客户事件", "EMP-071", "—"),
        (14, "转客户时间 / 转换客户名称", "在该员工责任期内转客户时填入", "EMP-071", "未在该周期转客户为空"),
        (15, "客户跟进人 / 客户首次跟进时间 / 客户最近一次跟进时间", "线索关联客户的最新数据；未转客户为空", "EMP-072", "和员工无关，按客户最新数据"),
    ]
    _write_rows(ws, rows)
    _autosize(ws, [6, 36, 80, 18, 32])


# ---------- Sheet 6: 执行 SOP ----------
def build_sop(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet6 执行SOP")
    headers = ["步骤", "操作", "执行人", "工具/入口", "校验点"]
    _style_header(ws, headers)
    rows = [
        (1, "准备环境：3 个销售账号、2 个特殊账号、1 个无导出权限账号", "测试经理", "管理员后台", "Sheet2 ACC-01~06 全部就绪"),
        (2, "整理基准对账表（默认 + 5 组组合）", "开发/BI", "SQL / BI", "Sheet2 BASE-01 通过"),
        (3, "执行造数脚本 dry-run 预演", "测试", "python scripts/create_lead_dashboard_emp_dataset.py --dry-run", "plan.json 与场景矩阵一致"),
        (4, "执行造数脚本（线索 + 活动）", "测试", "python scripts/create_lead_dashboard_emp_dataset.py", "lead_ids.json 全部 success"),
        (5, "按 manifest.md 在 CRM 触发 6 类结束事件", "测试", "CRM 后台手工", "每个场景结束时间已记录"),
        (6, "为 SPAN-001/OUT-001 调整历史时间字段", "DBA/开发", "DB 写入或测试库辅助接口", "实际 dispatch_time/void_time 一致"),
        (7, "触发系统回收（运维接口或等待定时任务）", "运维/测试", "测试库定时任务", "AUTO-POOL-001 状态为系统回收"),
        (8, "进入线索看板·人员模块进行对账", "测试", "线索看板", "EMP-004~073 全部对账通过"),
        (9, "执行批量导出与行内导出对账", "测试", "线索看板·人员模块", "导出文件 28 列、行数与基准一致"),
        (10, "权限/边界用例：无权限、>90天、空数据", "测试", "线索看板", "EMP-037/044/045/078 通过"),
        (11, "记录缺陷 + 复测", "测试", "缺陷平台", "0 P0 未关闭"),
        (12, "归档：lead_ids.json + 导出对账文件 + 用例执行单", "测试经理", "Git/网盘", "回归集稳定可复用"),
    ]
    _write_rows(ws, rows)
    _autosize(ws, [6, 56, 14, 36, 36])


# ---------- Sheet 7: 默认 scenarios.json ----------
def build_scenarios_json(wb: Workbook) -> None:
    ws = wb.create_sheet("Sheet7 scenarios.json 模板")
    ws["A1"] = "复制以下内容到 scenarios.json，并通过 --config 指定。"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = SECTION_FILL
    ws["A1"].alignment = ALIGN_WRAP
    tpl = """{
  "members": [
    {"name": "甲", "phone": "13800000001", "password_encrypted": "<base64-encoded>"},
    {"name": "乙", "phone": "13800000002", "password_encrypted": "<base64-encoded>"},
    {"name": "丙", "phone": "13800000003", "password_encrypted": "<base64-encoded>"}
  ],
  "scenarios": [
    {"id": "CUR-001", "owner": "甲", "activities": 1, "end_type": "当前持有"},
    {"id": "CUR-002", "owner": "甲", "activities": 0, "end_type": "当前持有"},
    {"id": "ASN-001", "owner": "甲", "activities": 1, "end_type": "分配", "transfer_to": "乙"},
    {"id": "ASN-002", "owner": "甲", "activities": 0, "end_type": "分配", "transfer_to": "乙"},
    {"id": "MAN-POOL-001", "owner": "甲", "activities": 1, "end_type": "主动移入公海"},
    {"id": "AUTO-POOL-001", "owner": "甲", "activities": 0, "end_type": "系统回收"},
    {"id": "WIN-CUST-001", "owner": "甲", "activities": 2, "end_type": "转客户"},
    {"id": "VOID-001", "owner": "甲", "activities": 1, "end_type": "作废"},
    {"id": "MULTI-CYCLE-001", "owner": "丙", "activities": 1, "end_type": "分配", "transfer_to": "甲"},
    {"id": "MULTI-CYCLE-002", "owner": "甲", "activities": 1, "end_type": "分配", "transfer_to": "丙", "reuse_lead_of": "MULTI-CYCLE-001"},
    {"id": "MULTI-USER-001", "owner": "甲", "activities": 1, "end_type": "分配", "transfer_to": "乙"},
    {"id": "MULTI-USER-002", "owner": "乙", "activities": 1, "end_type": "分配", "transfer_to": "丙", "reuse_lead_of": "MULTI-USER-001"},
    {"id": "MULTI-USER-003", "owner": "丙", "activities": 0, "end_type": "当前持有", "reuse_lead_of": "MULTI-USER-001"},
    {"id": "SPAN-001", "owner": "甲", "activities": 1, "end_type": "当前持有", "back_days_for_assign": 365},
    {"id": "OUT-001", "owner": "甲", "activities": 0, "end_type": "作废", "back_days_for_assign": 400, "void_after_days": 380}
  ]
}
"""
    ws["A3"] = tpl
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 130
    ws.row_dimensions[3].height = 460


def main() -> int:
    wb = Workbook()
    build_overview(wb)
    build_checklist(wb)
    build_scenario_matrix(wb)
    build_export_sample(wb)
    build_field_mapping(wb)
    build_sop(wb)
    build_scenarios_json(wb)
    wb.save(XLSX_PATH)
    try:
        shutil.copy2(XLSX_PATH, ROOT_COPY)
    except OSError as exc:
        print(f"[WARN] 复制到根目录失败: {exc}")
    print(f"[OK] 生成 {XLSX_PATH}")
    print(f"[OK] 副本 {ROOT_COPY}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
