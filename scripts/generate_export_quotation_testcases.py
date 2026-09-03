# -*- coding: utf-8 -*-
"""Generate test cases for CRM quotation export requirement."""
from __future__ import annotations

import json
import re
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "testcases"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TITLE = "CRM-询价单支持导出英文报价单"
DOC_URL = "https://tvd6quau8vr.feishu.cn/wiki/YjpvwS21ViNvyAktADFcjOpBn6b"
SUPPLEMENT_DOC_URL = "https://tvd6quau8vr.feishu.cn/wiki/QLE6woWF1iJZKak6djlcw9VPnVh"
CRM_FIELD_DOC_URL = "https://tvd6quau8vr.feishu.cn/wiki/YFlQwrbJfi8jKUkWO7AcL5yzn1d"
FIELD_SHEET_URL = (
    "https://tvd6quau8vr.feishu.cn/wiki/Orn9wUv93itKzMkiWcBcRygWnRh?sheet=zQuRgR"
)
XLSX_PATH = OUT_DIR / f"{TITLE}_TestCases.xlsx"
XMIND_PATH = OUT_DIR / f"{TITLE}_TestMindMap.xmind"

HEADERS = [
    "用例ID",
    "模块",
    "优先级",
    "场景",
    "前置条件",
    "测试步骤",
    "预期结果",
    "实际结果",
    "备注",
    "用例状态",
    "是否阻塞",
    "首轮必测",
]

STATUS_OPTIONS = "PASS,FAIL,BLOCK,N/A"

MODULES = [
    ("TAB", "1-询价单TAB页"),
    ("EXP", "2-导出入口与权限"),
    ("POP", "3-联系人选择弹窗"),
    ("TPL", "4-模板选择与Sheet分组"),
    ("HDR", "5-报价单表头取数"),
    ("COM", "6-商品通用字段"),
    ("BOX", "7-箱装模板字段"),
    ("ROL", "8-卷装模板字段"),
    ("OTH", "9-其他包装模板字段"),
    ("CAL", "10-公式空值与TOTAL"),
    ("FIL", "11-文件与导出记录"),
    ("ECF", "12-英文商城字段调整"),
    ("ADM", "13-后台页面兼容"),
    ("LNK", "14-字段联动与只读展示"),
    ("CRF", "15-CRM询价字段展示"),
    ("ACC", "16-异常与兼容回归"),
]

CASES: list[dict[str, str]] = []


def add(
    prefix: str,
    seq: int,
    priority: str,
    scene: str,
    precondition: str,
    steps: list[str],
    expected: str,
    note: str,
    blocking: str = "否",
    first: str = "否",
) -> None:
    module = dict(MODULES)[prefix]
    CASES.append(
        {
            "用例ID": f"{prefix}-{seq:03d}",
            "模块": module,
            "优先级": priority,
            "场景": scene,
            "前置条件": precondition,
            "测试步骤": "\n".join(f"{i}. {step}" for i, step in enumerate(steps, 1)),
            "预期结果": expected,
            "实际结果": "",
            "备注": note,
            "用例状态": "",
            "是否阻塞": blocking,
            "首轮必测": first,
        }
    )


def build_cases() -> None:
    add(
        "TAB",
        1,
        "P0",
        "按创建时间筛选询价单",
        "存在不同创建时间的询价主单；账号具备询价单TAB访问权限。",
        ["进入CRM询价单TAB页", "选择创建时间范围并查询", "与预置数据或接口返回进行核对", "清空时间筛选后再次查询"],
        "列表仅展示创建时间落在筛选范围内的询价单；清空后恢复展示可见范围内全部询价单。",
        "PRD：新增“创建时间”筛选项。",
        "是",
        "是",
    )
    add(
        "TAB",
        2,
        "P1",
        "创建时间筛选边界值正确",
        "存在创建时间等于开始时间、结束时间前后1秒的询价单。",
        ["设置包含开始与结束时间的筛选范围", "查询列表并记录命中的询价单", "分别调整开始/结束边界后复测"],
        "开始与结束边界命中规则一致且可解释；范围外数据不展示。",
        "待产品确认：创建时间筛选是否包含结束时刻。",
        "否",
        "否",
    )
    add(
        "TAB",
        3,
        "P0",
        "询价主单创建后默认生成报价单号",
        "具备创建询价主单权限；当天已有若干报价单号样例。",
        ["新建并保存一条询价主单", "返回询价单TAB页查看“报价单”列", "连续创建多条主单观察序列号"],
        "每个主单创建后默认生成报价单号；格式为 BJ+年月日+四位序列号，且同日序列递增不重复。",
        "PRD：单号生成规则 BJ+年月日+四位序列号。",
        "是",
        "是",
    )
    add(
        "TAB",
        4,
        "P0",
        "不同状态展示正确的当前操作人标签",
        "准备草稿、待提交技术方案、待出厂报价、待平台报价、待发起上架申请状态的询价单。",
        ["进入询价单TAB页", "分别查看各状态行的状态列", "核对当前操作人标签与对应角色人员"],
        "草稿=创建人；待提交技术方案=技术经理；待出厂报价=采购员；待平台报价/待发起上架申请=客户在商城的业务支撑。",
        "PRD：状态列新增“当前操作人”标签。",
        "是",
        "是",
    )
    add(
        "TAB",
        5,
        "P1",
        "角色人员为空时当前操作人展示未维护",
        "准备相关状态询价单，且技术经理/采购员/商城业务支撑等对应角色为空。",
        ["进入询价单TAB页", "查看状态列当前操作人标签", "补充角色人员后刷新列表复测"],
        "角色为空时展示“未维护”；维护人员后展示对应人员名称。",
        "PRD：对应角色人员为空展示未维护。",
        "否",
        "是",
    )
    add(
        "TAB",
        6,
        "P1",
        "非指定状态不显示当前操作人标签",
        "准备已完成、已取消或其他非PRD列举状态的询价单。",
        ["进入询价单TAB页", "查看非指定状态行", "切换筛选条件并复查"],
        "其他状态下不显示“当前操作人”标签，不占用异常空白布局。",
        "PRD：其他状态下不显示标签。",
    )

    add(
        "EXP",
        1,
        "P0",
        "满足条件且有权限时展示导出报价单按钮",
        "客户国家非中国；客户不在公海；存在状态为待平台报价/待发起上架申请/待关联报价商品/待转单/已完成的询价子单；账号有导出权限。",
        ["进入询价单TAB页", "定位满足条件的询价主单", "查看行操作或详情中的导出入口"],
        "展示“导出报价单”按钮，点击后进入联系人选择/导出流程。",
        "PRD：新增导出报价单按钮，权限单独配置。",
        "是",
        "是",
    )
    add(
        "EXP",
        2,
        "P0",
        "无导出权限时不展示导出报价单按钮",
        "同一询价单满足业务展示条件；账号无导出报价单权限。",
        ["使用无导出权限账号登录", "进入询价单TAB页", "查看满足业务条件的询价单操作区"],
        "不展示“导出报价单”按钮；直接访问导出接口也被拒绝且无文件生成。",
        "按钮权限需要支持单独配置。",
        "是",
        "是",
    )
    add(
        "EXP",
        3,
        "P1",
        "客户国家为中国时不展示导出报价单按钮",
        "客户所在国家=中国；询价子单状态满足导出状态；账号有导出权限。",
        ["进入询价单TAB页", "定位中国客户询价单", "查看导出入口"],
        "不展示导出报价单按钮；接口侧校验阻断导出。",
        "展示条件：客户当前所在国家不等于中国。",
        "否",
        "是",
    )
    add(
        "EXP",
        4,
        "P1",
        "客户在公海时不展示导出报价单按钮",
        "客户在公海；国家非中国；询价子单状态满足导出状态。",
        ["进入询价单TAB页", "定位公海客户询价单", "查看导出入口"],
        "不展示导出报价单按钮；接口侧校验阻断导出。",
        "展示条件：客户不在公海。",
    )
    add(
        "EXP",
        5,
        "P1",
        "无可导出状态子单时不展示按钮",
        "询价主单下全部子单状态不属于待平台报价、待发起上架申请、待关联报价商品、待转单、已完成。",
        ["进入询价单TAB页", "定位该询价主单", "查看导出入口"],
        "不展示导出报价单按钮；不会因主单存在而误展示。",
        "导出对象按主单维度合并导出。",
    )
    add(
        "EXP",
        6,
        "P0",
        "按主单维度合并导出全部询价明细",
        "同一主单下存在多个可导出子单和多条询价明细。",
        ["点击主单导出报价单", "选择联系人并确认导出", "打开下载的Excel核对明细"],
        "系统读取当前询价单的全部询价明细；按主单维度生成一个报价文件，不按子单拆多个文件。",
        "PRD：导出对象按主单维度合并导出。",
        "是",
        "是",
    )

    add(
        "POP",
        1,
        "P0",
        "多个联系人时打开选择联系人弹窗",
        "当前客户名下存在2个及以上联系人；询价单满足导出条件。",
        ["点击导出报价单", "观察页面弹窗", "查看联系人选择控件和只读字段"],
        "系统打开选择联系人弹窗；客户联系人为必选搜索单选；联系人姓名和邮箱为只读展示。",
        "PRD：有多个联系人则打开弹窗，用户选择一个客户联系人。",
        "是",
        "是",
    )
    add(
        "POP",
        2,
        "P1",
        "联系人搜索支持姓名电话邮箱",
        "客户下准备联系人A/B/C，分别通过姓名、电话、邮箱可唯一命中。",
        ["打开选择联系人弹窗", "分别输入姓名、电话、邮箱关键词搜索", "选择搜索结果"],
        "搜索结果来自当前客户名下联系人；支持按姓名、电话或邮箱命中；不展示其他客户联系人。",
        "弹窗展示姓名、公司名称、所在国家、电话和邮箱。",
        "否",
        "是",
    )
    add(
        "POP",
        3,
        "P1",
        "未选择联系人时确认导出不可点击",
        "打开联系人选择弹窗且未选择任何联系人。",
        ["观察确认导出按钮状态", "尝试点击确认导出", "选择联系人后再次观察"],
        "未选择联系人时确认导出不可点击；选择联系人后按钮变为可点击。",
        "规则：客户联系人必选。",
        "否",
        "是",
    )
    add(
        "POP",
        4,
        "P1",
        "选择联系人后只读字段同步展示",
        "联系人包含姓名与邮箱。",
        ["打开弹窗并选择联系人", "查看联系人姓名和联系人邮箱只读字段", "切换另一个联系人"],
        "只读字段实时展示所选联系人姓名和邮箱；切换联系人后同步刷新。",
        "CONTACT/EMAIL来源于导出弹窗。",
    )
    add(
        "POP",
        5,
        "P1",
        "联系人邮箱为空时允许导出并提示",
        "客户联系人姓名存在但邮箱为空；询价单满足导出条件。",
        ["打开弹窗", "选择邮箱为空的联系人", "点击确认导出", "打开导出文件查看EMAIL"],
        "弹窗提示“当前联系人邮箱为空，导出的报价单邮箱将为空”；导出不被阻断；Excel EMAIL为空。",
        "PRD：邮箱为空时允许导出。",
        "否",
        "是",
    )
    add(
        "POP",
        6,
        "P1",
        "取消按钮关闭弹窗且不导出",
        "已打开联系人选择弹窗。",
        ["点击取消按钮", "观察弹窗状态", "检查浏览器下载与导出记录"],
        "弹窗关闭；不触发文件下载；不新增成功导出记录。",
        "PRD：取消关闭弹窗，不导出。",
    )
    add(
        "POP",
        7,
        "P1",
        "确认导出期间按钮进入加载态防重复提交",
        "导出接口可模拟延迟响应。",
        ["选择联系人后点击确认导出", "在生成期间连续点击确认导出", "查看网络请求与导出记录"],
        "确认按钮进入加载/禁用态；只发起一次导出请求；生成完成或失败后恢复。",
        "PRD：生成期间禁用确认按钮，防止重复提交。",
        "否",
        "是",
    )

    add(
        "TPL",
        1,
        "P0",
        "箱装不打托进入箱装不打托报价单Sheet",
        "询价明细包装类型=箱装，是否打托=否。",
        ["导出报价单", "打开Excel文件", "查看Sheet名称与该明细所在Sheet"],
        "生成Sheet“箱装不打托报价单”；该明细只出现在该Sheet，Packing为Carton。",
        "模板规则：箱装+否。",
        "是",
        "是",
    )
    add(
        "TPL",
        2,
        "P0",
        "箱装打托进入箱装打托报价单Sheet",
        "询价明细包装类型=箱装，是否打托=是。",
        ["导出报价单", "打开Excel文件", "查看Sheet名称与该明细所在Sheet"],
        "生成Sheet“箱装打托报价单”；该明细只出现在该Sheet，Packing为Pallet。",
        "模板规则：箱装+是。",
        "是",
        "是",
    )
    add(
        "TPL",
        3,
        "P0",
        "卷装不打托进入卷装不打托报价单Sheet",
        "询价明细包装类型=卷装，是否打托=否。",
        ["导出报价单", "打开Excel文件", "查看Sheet名称与该明细所在Sheet"],
        "生成Sheet“卷装不打托报价单”；该明细只出现在该Sheet，Packing为Roll。",
        "模板规则：卷装+否。",
        "是",
        "是",
    )
    add(
        "TPL",
        4,
        "P0",
        "卷装打托进入卷装打托报价单Sheet",
        "询价明细包装类型=卷装，是否打托=是。",
        ["导出报价单", "打开Excel文件", "查看Sheet名称与该明细所在Sheet"],
        "生成Sheet“卷装打托报价单”；该明细只出现在该Sheet，Packing为Pallet。",
        "模板规则：卷装+是。",
        "是",
        "是",
    )
    add(
        "TPL",
        5,
        "P0",
        "其他包装不打托进入其他不打托Sheet",
        "询价明细包装类型=其他，是否打托=否；供应链维护其他包装英文名称。",
        ["导出报价单", "打开Excel文件", "查看Sheet名称、Packing与该明细所在Sheet"],
        "生成Sheet“其他-不打托报价单”；Packing取供应链询价结果中的其他包装英文名称。",
        "模板规则：其他+否。",
        "是",
        "是",
    )
    add(
        "TPL",
        6,
        "P0",
        "其他包装打托阻断导出",
        "询价明细包装类型=其他，是否打托=是。",
        ["点击导出报价单", "选择联系人并确认", "观察页面提示和下载结果"],
        "系统阻断导出并明确提示不允许导出；不生成Excel，不新增成功导出记录。",
        "模板规则：其他+是=不允许，阻断导出。",
        "是",
        "是",
    )
    add(
        "TPL",
        7,
        "P0",
        "同一询价单多模板类型生成一个Excel多个Sheet",
        "同一主单包含箱装不打托、箱装打托、卷装不打托、卷装打托、其他不打托明细。",
        ["导出报价单", "打开下载的Excel", "检查文件数量、Sheet数量和每个Sheet明细"],
        "只下载一个Excel；按模板类型生成多个非空Sheet；每条明细进入唯一正确Sheet。",
        "PRD：一个Excel，多个Sheet。",
        "是",
        "是",
    )
    add(
        "TPL",
        8,
        "P1",
        "Sheet顺序与明细顺序符合规则",
        "同一主单包含多种模板类型，明细录入顺序可识别。",
        ["导出报价单", "查看Sheet排列顺序", "查看每个Sheet内商品行顺序"],
        "Sheet顺序按PRD模板表顺序排列；每个Sheet内商品顺序与询价单中的明细顺序一致。",
        "PRD：Sheet顺序按上表；商品顺序与明细顺序一致。",
        "否",
        "是",
    )
    add(
        "TPL",
        9,
        "P1",
        "所有Sheet共用同一报价单基础信息",
        "同一文件包含多个Sheet。",
        ["导出报价单", "逐个Sheet查看Quotation No.、Date、BUYER、ADDRESS、EMAIL、CONTACT", "与客户和联系人数据核对"],
        "同一文件内所有Sheet上述字段完全一致，取值符合所选客户和联系人。",
        "PRD：同一文件所有Sheet共用基础信息。",
        "否",
        "是",
    )

    add(
        "HDR",
        1,
        "P0",
        "报价单表头客户与联系人字段取数正确",
        "客户维护公司名称、详细地址；选择联系人有姓名和邮箱。",
        ["导出报价单", "打开Excel任一Sheet", "核对BUYER、ADDRESS、EMAIL、CONTACT"],
        "BUYER=CRM公司名称；ADDRESS=CRM详细地址；EMAIL=所选联系人邮箱；CONTACT=所选联系人姓名。",
        "PRD 6.1 报价单表头。",
        "是",
        "是",
    )
    add(
        "HDR",
        2,
        "P1",
        "客户详细地址为空时ADDRESS导出为空",
        "客户详细地址为空；其他导出条件满足。",
        ["导出报价单", "打开Excel查看ADDRESS字段"],
        "ADDRESS单元格为空，不填默认值，不报错。",
        "ADDRESS为空时导出为空。",
    )
    add(
        "HDR",
        3,
        "P1",
        "SELLER和SELLER ADDRESS固定值正确",
        "任意满足导出条件的询价单。",
        ["导出报价单", "打开Excel查看SELLER与SELLER ADDRESS"],
        "SELLER固定为EPAK (JIANG SU) CO., LTD；SELLER ADDRESS固定为PRD指定英文地址。",
        "固定字段按PRD原文导出。",
        "否",
        "是",
    )
    add(
        "HDR",
        4,
        "P1",
        "Quotation No和Date格式正确",
        "系统日期已知；询价单已有报价单号。",
        ["导出报价单", "查看Quotation No.与Date", "跨天或模拟日期后复测Date"],
        "Quotation No.取系统生成报价单号；Date取导出当天日期，格式mm/DD/YY。",
        "PRD：Date格式 mm/DD/YY。",
    )
    add(
        "HDR",
        5,
        "P1",
        "Notes与红色条款按模板原样保留",
        "最新导出模板可用。",
        ["导出报价单", "打开Excel查看Notes区域", "与模板源文件比对红色贸易条款和红色提示"],
        "Notes按最新模板原样导出；红色贸易条款和红色提示保留，系统不删减、不判断。",
        "PRD：销售导出后自行处理。",
    )

    add(
        "COM",
        1,
        "P0",
        "已上架商品通用字段取上架信息",
        "询价明细关联已上架商品，SKU、商品名称、图片存在。",
        ["导出报价单", "查看商品行Item No.、Product Name、Picture", "与商品上架信息核对"],
        "Item No.取SKU；Product Name取商品名称；Picture嵌入商品图片单元格。",
        "PRD 6.2 商品通用字段。",
        "是",
        "是",
    )
    add(
        "COM",
        2,
        "P1",
        "未上架商品通用字段按询价明细取数",
        "询价明细未关联上架商品，维护物料名称、规格、材质和首单数量。",
        ["导出报价单", "查看商品通用字段", "与询价明细核对"],
        "Item No.为空；Product Name取询价明细物料名称；Picture为空；规格、材质、首单数量按询价明细导出。",
        "未上架导出规则。",
        "否",
        "是",
    )
    add(
        "COM",
        3,
        "P1",
        "Specification必须带合法单位",
        "准备规格带mm、cm、inch以及缺少单位的明细。",
        ["分别导出合法单位明细", "导出缺少单位的明细", "观察导出结果或校验提示"],
        "合法单位按原规格导出；缺少mm/cm/inch单位时应按系统校验策略提示或在备注中明确，不应静默生成错误规格。",
        "PRD：Specification必须带 mm、cm 或 inch 单位。待确认：缺单位时是否阻断。",
    )
    add(
        "COM",
        4,
        "P1",
        "Unit Price与Remarks导出为空",
        "任意满足导出条件的商品明细。",
        ["导出报价单", "查看Unit Price (USD/pcs)和Remarks列"],
        "Unit Price导出为空且不填0；Remarks导出为空，供销售后续填写。",
        "销售填写字段。",
        "否",
        "是",
    )
    add(
        "COM",
        5,
        "P1",
        "Order Amount公式随销售填写单价自动计算",
        "导出文件中Order Quantity有值，Unit Price为空。",
        ["打开导出的Excel", "在Unit Price中输入数值", "查看Order Amount计算结果"],
        "Order Amount按Unit Price × Order Quantity计算，保留2位小数；输入变化后自动重算。",
        "PRD：Excel公式计算。",
        "否",
        "是",
    )

    add("BOX", 1, "P0", "箱装不打托字段和公式正确", "箱装不打托明细维护Qty/Ctn、纸箱尺寸、单箱重量。", ["导出报价单", "查看箱装不打托Sheet字段", "核对Carton、Volume、Gross weight公式"], "Qty/Ctn和Carton size直接取供应链询价结果；Carton=Order Quantity÷Qty/Ctn；Volume=长×宽×高÷1,000,000×Carton；Gross weight=单箱重量×Carton，均保留2位小数。", "PRD 6.3。", "是", "是")
    add("BOX", 2, "P0", "箱装打托字段和公式正确", "箱装打托明细维护每箱装数、纸箱尺寸、每托箱数、打托总数、每托重量、每托体积。", ["导出报价单", "查看箱装打托Sheet", "核对Pallet与重量体积公式"], "Qty/Ctn、Cartons Per Pallet、Pallet、Weight Per Pallet、Volume Per Pallet直接取供应链结果；Gross weight=Weight Per Pallet×Pallet；Volume=Volume Per Pallet×Pallet。", "PRD 6.4。", "是", "是")
    add("ROL", 1, "P0", "卷装不打托字段和公式正确", "卷装不打托明细维护每卷重量、每卷体积、卷数。", ["导出报价单", "查看卷装不打托Sheet", "核对Roll size、Roll、Volume、Gross weight"], "Weight/Roll、Roll size、Roll直接取供应链结果；Volume=每卷体积÷1,000,000×Roll；Gross weight=Weight/Roll×Roll，保留2位小数。", "PRD 6.5。", "是", "是")
    add("ROL", 2, "P0", "卷装打托字段和公式正确", "卷装打托明细维护每卷重量、每卷体积、卷数、托数、每托重量、每托体积、每托盘装卷数。", ["导出报价单", "查看卷装打托Sheet", "核对打托字段与公式"], "Roll/Pallet/Weight Per Pallet/Volume Per Pallet/Rolls Per Pallet直接取供应链结果；Gross weight=Weight Per Pallet×Pallet；Volume=Volume Per Pallet×Pallet。", "PRD 6.6。", "是", "是")
    add("ROL", 3, "P1", "Roll size字段按体积含义取值", "卷装明细维护直径、高度并产生供应链每卷体积值。", ["导出卷装报价单", "查看Roll size字段", "与供应链询价结果中的体积值核对"], "Roll size取供应链提供的每卷体积数值，不按尺寸文本拼接；后续CBM换算正确。", "PRD说明字段实际含义为每卷体积。", "否", "是")
    add("OTH", 1, "P0", "其他不打托字段和公式正确", "其他包装不打托明细维护Qty/Pac、单个包装体积、单个包装重量。", ["导出报价单", "查看其他-不打托Sheet", "核对Pack、Volume、Gross weight公式"], "Qty/Pac、Pac size、Pac weight直接取供应链结果；Pack=Order Quantity÷Qty/Pac；Volume=单个包装体积×Pack；Gross weight=Pac weight×Pack。", "PRD 6.7。", "是", "是")
    add("OTH", 2, "P1", "其他包装英文名称用于Packing", "其他包装不打托明细维护其他包装英文名称。", ["导出报价单", "查看其他-不打托Sheet的Packing字段"], "Packing取供应链询价结果中的其他包装英文名称，未错误固定为Other或空值。", "模板选择规则表。")

    add(
        "CAL",
        1,
        "P0",
        "除法分母为空或为0时公式返回空值",
        "准备Qty/Ctn或Qty/Pac为空、0的明细。",
        ["导出报价单", "打开Excel查看Carton/Pack及依赖字段", "检查单元格错误值"],
        "分母为空或0时相关公式返回空值；文件不出现#DIV/0!、#VALUE!等错误。",
        "PRD：除法公式在分母为空或0时返回空值。",
        "是",
        "是",
    )
    add(
        "CAL",
        2,
        "P0",
        "公式依赖值为空时不产生错误值",
        "准备尺寸、重量、体积等依赖字段为空的明细。",
        ["导出报价单", "打开Excel查看体积、重量、金额等公式字段", "全工作簿搜索错误值"],
        "其他公式任一依赖值为空时返回空值；最终文件无#DIV/0!、#VALUE!、#N/A、#REF!等错误。",
        "PRD：最终文件不得出现错误值。",
        "是",
        "是",
    )
    add(
        "CAL",
        3,
        "P1",
        "Carton和Pack不向上取整并保留2位",
        "准备Order Quantity不能被Qty/Ctn或Qty/Pac整除的明细。",
        ["导出报价单", "查看Carton/Pack计算结果", "与手工计算值比对"],
        "Carton、Pack按除法结果保留2位小数，不向上取整。",
        "PRD：Carton、Pack不向上取整。",
        "否",
        "是",
    )
    add(
        "CAL",
        4,
        "P0",
        "TOTAL行以公式汇总当前Sheet商品行",
        "每个模板Sheet至少包含2条商品明细。",
        ["导出报价单", "查看每个Sheet的TOTAL行", "新增/修改商品行数值后观察汇总"],
        "TOTAL行以公式汇总当前Sheet适用字段：Order Quantity、Order Amount、Carton/Roll/Pack/Pallet、Gross weight、Volume；不跨Sheet汇总。",
        "PRD 6.8与第七章。",
        "是",
        "是",
    )
    add(
        "CAL",
        5,
        "P1",
        "新增商品行时公式可同步填充",
        "导出后的Excel保留模板和公式。",
        ["在商品区域插入或复制新增商品行", "填写必要数量、包装、重量体积字段", "查看金额和TOTAL公式"],
        "新增商品行公式可正常计算，TOTAL汇总范围覆盖新增行或通过模板约定可正确扩展。",
        "PRD：新增商品行时同步填充公式。待确认：用户手工新增行的支持边界。",
    )

    add(
        "FIL",
        1,
        "P0",
        "导出文件名格式正确并替换非法字符",
        "客户名称包含文件名非法字符；询价单号已知。",
        ["导出报价单", "查看下载文件名", "核对客户名非法字符处理"],
        "文件格式为.xlsx；文件名为Quotation_客户名称_询价单号.xlsx；客户名称中的非法字符替换为下划线。",
        "PRD：文件与记录规则。",
        "是",
        "是",
    )
    add(
        "FIL",
        2,
        "P1",
        "导出保留模板样式和格式",
        "最新模板Products list for xxxx 20260623.xlsx已配置。",
        ["导出报价单", "与模板源文件对比样式、公司信息、Notes、红色字体、合并单元格、金额格式"],
        "导出文件保留模板样式、公司信息、Notes、红色字体、合并单元格和金额格式；不修改模板源文件。",
        "PRD：不修改模板源文件。",
        "否",
        "是",
    )
    add(
        "FIL",
        3,
        "P0",
        "成功导出后写入导出记录",
        "导出流程成功完成；可查询数据库或审计表。",
        ["导出报价单", "查询导出记录", "核对记录字段"],
        "记录Quotation No.、询价单号、客户、联系人、操作人、导出时间、Sheet名称、文件名和结果。",
        "PRD：每次成功导出记录。",
        "是",
        "是",
    )
    add(
        "FIL",
        4,
        "P0",
        "成功导出后保存完整字段取值JSON",
        "导出流程成功完成；可访问JSON存储位置。",
        ["导出报价单", "查找本次导出对应JSON文件", "核对JSON字段与Excel字段取值"],
        "系统记录导出时整个文件里的详细数据；每个字段取值完整写入JSON，且与Excel一致。",
        "0724补充：存json文件。",
        "是",
        "是",
    )
    add(
        "FIL",
        5,
        "P1",
        "导出失败不写成功记录且按钮恢复",
        "模拟模板缺失、供应链取数异常或下载失败。",
        ["点击导出报价单", "触发失败场景", "查看页面状态、下载结果和导出记录"],
        "页面给出失败提示；确认按钮恢复；不下载坏文件；不写成功记录，失败记录按系统约定可追踪。",
        "异常路径。",
    )

    add(
        "ECF",
        1,
        "P0",
        "G单包装方式与是否打托必填及枚举正确",
        "英文商城供应商端/商家端可编辑G单；表单字段-2.2已生效。",
        [
            "进入G单编辑页报价信息区域",
            "核对包装方式选项：纸箱包装/卷类包装/其他（手工输入）",
            "分别保存包装方式为空、是否打托为空",
            "选择合法组合后保存",
        ],
        "包装方式、是否打托均为必填；包装方式枚举为纸箱包装/卷类包装/其他；其他可手工输入且限200字；合法保存成功。",
        "来源：表单字段-2.2 供应商单（G）（英文）报价信息。",
        "是",
        "是",
    )
    add(
        "ECF",
        2,
        "P0",
        "包装方式=其他时是否打托固定为否",
        "可编辑G单报价信息。",
        [
            "将包装方式改为其他并手工输入英文包装名",
            "观察是否打托字段",
            "尝试改为是并保存",
            "保存后重新打开核对",
        ],
        "包装方式=其他时是否打托固定为否且不可改为是；保存后仍为否，对应导出模板走其他-不打托。",
        "备注：包装方式=其他时，固定为“否”。",
        "是",
        "是",
    )
    add(
        "ECF",
        3,
        "P0",
        "纸箱包装字段显隐与展示时必填校验",
        "G单编辑页；准备纸箱包装+打托是/否两套数据。",
        [
            "选择包装方式=纸箱包装、是否打托=否",
            "检查每箱装数Qty/Ctn、纸箱尺寸Carton size、每纸箱重量是否展示且必填",
            "清空上述字段尝试保存",
            "改为是否打托=是，检查打托总数/每托重量/每托体积/每托盘装箱数Ctn/PLT/20gp/40hq展示与必填",
        ],
        "仅纸箱包装时展示箱规字段；仅打托=是时展示打托相关字段；展示时必填未填则阻断保存。",
        "表单字段-2.2：纸箱包装联动字段。",
        "是",
        "是",
    )
    add(
        "ECF",
        4,
        "P0",
        "卷类包装字段显隐与展示时必填校验",
        "G单编辑页；准备卷类包装+打托是/否两套数据。",
        [
            "选择包装方式=卷类包装、是否打托=否",
            "检查每卷重量、卷装直径、卷装高度、总卷数、卷类包装方式展示与必填",
            "改为是否打托=是",
            "检查直径高度是否隐藏，并展示每托盘装卷数/20gp/40HQ及打托字段必填",
        ],
        "卷类不打托展示直径高度；卷类打托展示托盘装卷数字段；展示时必填生效；非卷类时这些字段不展示。",
        "表单字段-2.2：卷类包装联动字段。",
        "是",
        "是",
    )
    add(
        "ECF",
        5,
        "P0",
        "其他包装字段显隐与展示时必填校验",
        "G单编辑页。",
        [
            "选择包装方式=其他",
            "检查单个包装商品数量、单个包装体积、单个包装重量、总包装数展示",
            "清空展示字段尝试保存",
            "填写合法数字后保存",
        ],
        "仅其他包装展示上述4个字段且展示时必填；是否打托固定否；不出现箱规/卷规字段。",
        "表单字段-2.2：其他包装字段。",
        "是",
        "是",
    )
    add(
        "ECF",
        6,
        "P1",
        "定制品G单包装相关字段展示预估ICON",
        "询价单类型=定制品；商家端和供应商端G单填写/查看页可访问。",
        [
            "打开定制品G单填写页",
            "核对打托总数、每托重量/体积、Qty/Ctn、纸箱尺寸、单箱重量、Ctn/PLT、卷规/其他包装等字段旁预估ICON",
            "保存后进入商家端与供应商端查看页复查",
            "切换为通用品对照无预估ICON字段",
        ],
        "表单字段-2.2标记“定制品时给到预估标识=√”的字段在定制品G单填写/查看时展示预估ICON；商家端和供应商端一致。",
        "PRD-3.89 + 表单字段-2.2 预估标识列。",
        "否",
        "是",
    )
    add(
        "ECF",
        7,
        "P0",
        "G单包装字段上推S单并供英文报价单取数",
        "G单填写完整纸箱/卷类/其他包装组合并上推；CRM可导出英文报价单。",
        [
            "在G单保存包装方式、是否打托及对应明细字段",
            "确认字段上推到S单出厂报价信息",
            "在CRM询价单导出英文报价单",
            "核对模板Sheet与Qty/Ctn、Carton size、Pallet、Roll、Pac等取数",
        ],
        "G单包装字段成功上推S单；导出报价单按包装方式+是否打托进入正确模板，供应链取数字段与表单一致。",
        "G表：是否上推子单（S）=√；主需求导出取供应链询价结果。",
        "是",
        "是",
    )
    add(
        "ECF",
        8,
        "P1",
        "规格长宽高必须带mm/cm/inch单位",
        "G单/S单可编辑规格（长*宽*高）字段。",
        [
            "输入无单位数字规格并保存",
            "分别输入带mm、cm、inch的规格并保存",
            "导出英文报价单查看Specification",
        ],
        "无单位时按表单规则阻断或提示；带mm/cm/inch可保存；导出Specification保留单位。",
        "G表格式：长宽高，数字+单位（mm/cm/inch）。",
        "否",
        "是",
    )

    add(
        "ADM",
        1,
        "P0",
        "商家后台S单详情展示包装字段且是否打托独立成列",
        "S单已有纸箱/卷类/其他包装出厂报价信息。",
        [
            "进入商家后台S单详情",
            "查看是否打托是否独立成列",
            "查看包装方式说明区域字段顺序与有值字段",
            "核对G单列表中的包装摘要",
        ],
        "是否打托单独成列；其余包装明细整合在包装方式说明区域并按表单字段顺序展示有值字段。",
        "PRD-3.89 只读展示规则。",
        "是",
        "是",
    )
    add(
        "ADM",
        2,
        "P0",
        "线下单模板与导入覆盖包装联动字段",
        "商家后台线下单导入可用。",
        [
            "下载线下单导入模板",
            "核对包装方式、是否打托及箱/卷/其他联动列",
            "导入合法纸箱打托、卷类不打托、其他包装样本",
            "导入缺展示时必填字段样本并查看错误",
        ],
        "模板含新增包装字段；合法数据落库；缺联动必填字段时给出明确错误且不写脏数据。",
        "PRD：导入/模板下载兼容新增字段与联动。",
        "是",
        "是",
    )
    add(
        "ADM",
        3,
        "P1",
        "商家端与供应商端G单详情包装字段一致",
        "同一G单已保存完整包装字段。",
        [
            "分别打开商家端G单详情和供应商端G单详情",
            "核对包装方式、是否打托及明细字段",
            "核对定制品预估ICON展示是否一致",
        ],
        "两端只读值一致；是否打托独立成列；预估ICON规则一致。",
        "涉及商家后台G单详情、供应商后台G单详情。",
        "否",
        "是",
    )
    add(
        "ADM",
        4,
        "P1",
        "S单出厂报价信息透传展示G单包装字段",
        "G单已上推包装字段到S单。",
        [
            "打开商家后台S单详情出厂报价信息",
            "逐项核对包装方式到总包装数字段",
            "确认S单侧为透传展示而非重新编辑丢失",
        ],
        "S单出厂报价信息完整展示G单上推的包装字段；与表单字段-2.2 S表清单一致。",
        "S表：出厂报价包装字段是否透传给供应商/上推关系。",
        "否",
        "是",
    )
    add(
        "ADM",
        5,
        "P2",
        "历史S/G单无新包装字段时可打开和保存",
        "存在字段调整前的历史单据。",
        [
            "打开历史S单/G单详情与编辑页",
            "不改包装字段直接保存",
            "必要时补齐新字段后再导出报价单",
        ],
        "历史单不报错；空值兼容；补齐后可正常导出。",
        "兼容历史数据。",
    )

    add(
        "LNK",
        1,
        "P0",
        "编辑态按包装方式切换联动显隐",
        "G单编辑页。",
        [
            "在纸箱包装、卷类包装、其他之间切换",
            "每次切换观察字段显隐",
            "确认隐藏字段不会残留为必填阻断",
            "保存后重新打开确认结果",
        ],
        "编辑态严格按表单字段-2.2备注联动显隐；保存后再次打开仍符合当前包装方式。",
        "字段联动关系在保存时处理。",
        "是",
        "是",
    )
    add(
        "LNK",
        2,
        "P0",
        "只读态有值字段直接展示不重算联动隐藏",
        "构造历史或特殊数据：当前包装方式下本不应展示但字段已有值。",
        [
            "打开S单/G单只读详情",
            "查看包装方式说明区域",
            "与编辑态联动结果对比",
        ],
        "只读态不按联动重新隐藏；有数据字段按表单顺序直接展示；是否打托仍独立成列。",
        "PRD：展示时不需要做逻辑判断，直接展示有数据的字段。",
        "是",
        "是",
    )
    add(
        "LNK",
        3,
        "P0",
        "包装方式映射到英文报价单模板类型",
        "准备纸箱不打托、纸箱打托、卷类不打托、卷类打托、其他不打托数据。",
        [
            "分别保存五类包装组合并上推/同步到CRM询价明细",
            "导出英文报价单",
            "核对Sheet名称与Packing",
        ],
        "纸箱+否→箱装不打托/Carton；纸箱+是→箱装打托/Pallet；卷类+否→卷装不打托/Roll；卷类+是→卷装打托/Pallet；其他+否→其他-不打托且Packing取手工英文名；其他+是阻断。",
        "表单“纸箱包装/卷类包装/其他”需与主PRD模板“箱装/卷装/其他”映射。",
        "是",
        "是",
    )
    add(
        "LNK",
        4,
        "P1",
        "数字格式两位小数与正整数校验",
        "G单编辑页包装数字字段。",
        [
            "对重量体积类字段输入超过2位小数",
            "对数量类字段输入小数或负数",
            "输入合法正整数/2位小数后保存",
        ],
        "正整数字段拒绝小数负数；2位小数字段按格式校验或按规则截断/提示；合法值可保存。",
        "表单字段格式：正整数 / 数字2位小数。",
        "否",
        "是",
    )
    add(
        "LNK",
        5,
        "P1",
        "S单产品信息必填项与下推G单校验",
        "创建英文商城S单。",
        [
            "按表单字段-2.2核对产品信息必填项",
            "缺必填保存应失败",
            "完整填写后下推/创建G单",
            "核对G单询价信息透传字段",
        ],
        "S单必填项生效；下推到G单的字段与“是否透传给供应商”标记一致。",
        "S表产品信息必填与透传列。",
        "否",
        "是",
    )

    add(
        "CRF",
        1,
        "P0",
        "CRM询价单详情展示商城包装基础字段",
        "商城S单/G单已维护包装方式、是否打托并同步到CRM询价单。",
        [
            "打开对应CRM询价单详情",
            "查看新增包装字段及字段顺序",
            "与商城端同一单据的包装方式、是否打托逐项核对",
        ],
        "CRM询价单详情展示包装方式和是否打托；字段名称、值及顺序以商城端“表单字段-2.2”为准，不发生枚举误映射。",
        "关联需求：CRM询价单及子单详情新增字段；以商城端为准。",
        "是",
        "是",
    )
    add(
        "CRF",
        2,
        "P0",
        "CRM询价子单详情展示商城包装基础字段",
        "同一询价主单下存在子单；商城端已维护包装方式、是否打托并同步。",
        [
            "打开CRM询价主单并进入目标子单详情",
            "查看新增包装字段",
            "与商城端及CRM主单对应数据核对",
        ],
        "询价子单详情展示包装方式和是否打托；取值与商城端对应单据一致，且主子单关联关系正确，不串单。",
        "关联需求明确覆盖“询价单及子单详情”。",
        "是",
        "是",
    )
    add(
        "CRF",
        3,
        "P0",
        "CRM详情展示纸箱包装联动字段",
        "准备纸箱包装且是否打托分别为否、是的商城询价数据。",
        [
            "分别打开两类数据对应的CRM询价单及子单详情",
            "核对Qty/Ctn、纸箱尺寸、单箱重量",
            "核对打托数据的打托总数、每托重量、每托体积及每托箱数",
        ],
        "纸箱包装字段与商城端保存值一致；打托=是时相关托盘字段完整展示，打托=否的数据不误取其他单据字段。",
        "字段集合来源：表单字段-2.2；待产品确认主单与子单各自承载字段范围。",
        "是",
        "是",
    )
    add(
        "CRF",
        4,
        "P0",
        "CRM详情展示卷类包装联动字段",
        "准备卷类包装且是否打托分别为否、是的商城询价数据。",
        [
            "分别打开对应CRM询价单及子单详情",
            "核对每卷重量、卷装直径、卷装高度、总卷数",
            "核对打托数据的托数、每托重量体积及每托盘装卷数",
        ],
        "卷类包装及托盘相关字段按商城端保存值展示；数值、单位和所属子单正确。",
        "字段集合来源：表单字段-2.2。",
        "是",
        "是",
    )
    add(
        "CRF",
        5,
        "P0",
        "CRM详情展示其他包装联动字段",
        "商城端包装方式=其他、是否打托=否，已维护其他包装英文名称及包装数量重量体积。",
        [
            "打开对应CRM询价单及子单详情",
            "核对其他包装名称、每包装数量、单个包装体积、单个包装重量及总包装数",
            "与商城端保存数据逐项核对",
        ],
        "CRM展示的其他包装字段和值与商城端一致；是否打托为否；英文包装名称可供报价单Packing取值。",
        "关联主需求“其他-不打托报价单”取数。",
        "是",
        "是",
    )
    add(
        "CRF",
        6,
        "P0",
        "商城到CRM再到英文报价单字段链路一致",
        "商城端准备纸箱、卷类、其他包装的完整字段组合；CRM具备导出报价单权限。",
        [
            "保存商城S单/G单包装字段",
            "打开CRM询价单及子单详情核对同步值",
            "从CRM导出英文报价单",
            "核对Sheet分组、Packing及包装数量重量体积字段",
        ],
        "商城端、CRM询价单/子单详情和导出Excel三处字段值一致；模板按包装方式+是否打托正确选择，公式使用正确源数据。",
        "三份需求的端到端关联验收。",
        "是",
        "是",
    )
    add(
        "CRF",
        7,
        "P1",
        "CRM历史询价单缺少新增字段时兼容展示",
        "存在字段上线前创建的CRM询价主单和子单，新增包装字段为空。",
        [
            "打开历史询价单及子单详情",
            "查看新增字段区域和页面其他信息",
            "尝试执行原有查看、流转操作",
        ],
        "页面可正常打开；空字段不展示错误值、不造成布局或接口异常；原有信息和操作不受影响。",
        "兼容历史数据；导出前是否必须补齐字段按主需求校验。",
        "否",
        "是",
    )
    add(
        "CRF",
        8,
        "P1",
        "CRM主子单字段权限与数据隔离正确",
        "准备有权限和无目标客户数据权限的账号；存在多个客户及多张询价单。",
        [
            "使用有权限账号查看主单及子单新增字段",
            "使用无数据权限账号尝试访问相同详情",
            "切换不同客户和子单核对字段归属",
        ],
        "有权限账号可见其数据范围内字段；无权限账号无法越权读取；不同客户、主单和子单之间不串值。",
        "新增字段沿用CRM询价单现有数据权限。",
        "否",
        "是",
    )

    add(
        "ACC",
        1,
        "P1",
        "导出接口防越权访问",
        "账号A无目标客户/询价单数据权限，账号B有权限。",
        ["使用账号A构造目标询价单导出请求", "观察接口响应", "检查文件与记录"],
        "接口拒绝越权导出；不泄露客户、联系人和询价明细数据；不生成文件或成功记录。",
        "权限与数据安全基线。",
        "是",
        "是",
    )
    add(
        "ACC",
        2,
        "P1",
        "并发重复导出不会生成重复异常记录",
        "同一询价单满足导出条件；可并发触发导出请求。",
        ["同时发起2次导出请求", "查看返回文件、Quotation No.与导出记录", "检查是否存在脏数据"],
        "系统按幂等/并发策略处理；无重复异常记录；文件和记录可追溯。",
        "PRD未明确并发策略，建议测前确认。",
    )
    add(
        "ACC",
        3,
        "P2",
        "大批量明细导出性能与可用性",
        "同一询价单包含接近业务上限数量的明细，覆盖多模板类型。",
        ["导出大批量报价单", "记录生成耗时和文件大小", "打开Excel检查公式与Sheet"],
        "导出在可接受时间内完成；浏览器未卡死；Excel可正常打开，公式无错误。",
        "待确认：单次导出明细数量上限与性能SLA。",
    )
    add(
        "ACC",
        4,
        "P2",
        "Excel文件在常用办公软件中可打开",
        "已成功导出包含多Sheet和图片的报价单。",
        ["使用Microsoft Excel打开文件", "使用WPS打开文件", "检查图片、合并单元格、公式和样式"],
        "文件可正常打开；图片、公式、样式和金额格式兼容；无修复文件提示。",
        "兼容性回归。",
    )


def tid() -> str:
    return str(uuid.uuid4()).replace("-", "")[:16]


def safe_sheet_name(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def row_values(case: dict[str, str]) -> list[str]:
    return [case[h] for h in HEADERS]


def style_header(ws, headers: list[str], fill_color: str = "4472C4") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def write_cases_sheet(ws, cases: list[dict[str, str]]) -> None:
    style_header(ws, HEADERS)
    block_fill = PatternFill("solid", fgColor="FFC7CE")
    p0_fill = PatternFill("solid", fgColor="DDEBF7")
    for row_idx, case in enumerate(cases, 2):
        values = row_values(case)
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if case["是否阻塞"] == "是":
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = block_fill
        elif case["优先级"] == "P0":
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = p0_fill

    widths = {
        1: 12,
        2: 22,
        3: 8,
        4: 32,
        5: 32,
        6: 46,
        7: 46,
        8: 14,
        9: 30,
        10: 12,
        11: 12,
        12: 12,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    status_col = HEADERS.index("用例状态") + 1
    dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}1000")


def build_excel() -> None:
    wb = Workbook()

    total = len(CASES)
    p0 = sum(1 for case in CASES if case["优先级"] == "P0")
    block = sum(1 for case in CASES if case["是否阻塞"] == "是")
    first = sum(1 for case in CASES if case["首轮必测"] == "是")

    ws_meta = wb.active
    ws_meta.title = "需求追溯"
    meta_rows = [
        ("需求文档", TITLE),
        (
            "飞书地址",
            f"{DOC_URL}\n关联资料：{SUPPLEMENT_DOC_URL}\nCRM字段需求：{CRM_FIELD_DOC_URL}\n字段表：{FIELD_SHEET_URL}",
        ),
        (
            "功能范围",
            "询价单TAB页、导出入口权限、联系人弹窗、模板选择、字段取数、公式空值、文件记录、英文商城字段调整、后台页面兼容、字段联动与只读展示、CRM询价主子单字段展示、回归",
        ),
        ("用例状态", "PASS / FAIL / BLOCK / N/A"),
        ("统计", f"合计{total}条；阻塞{block}条；首轮必测{first}条；P0={p0}条"),
    ]
    for row_idx, (key, value) in enumerate(meta_rows, 1):
        ws_meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_meta.cell(row=row_idx, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
    ws_meta.column_dimensions["A"].width = 16
    ws_meta.column_dimensions["B"].width = 100

    ws_all = wb.create_sheet("总测试用例")
    write_cases_sheet(ws_all, CASES)

    for _, module_name in MODULES:
        module_cases = [case for case in CASES if case["模块"] == module_name]
        if not module_cases:
            continue
        ws = wb.create_sheet(safe_sheet_name(module_name))
        write_cases_sheet(ws, module_cases)

    ws_smoke = wb.create_sheet("首轮冒烟")
    write_cases_sheet(ws_smoke, [case for case in CASES if case["首轮必测"] == "是"])

    ws_block = wb.create_sheet("阻塞场景清单")
    block_headers = ["用例ID", "模块", "场景", "优先级", "阻塞说明", "失败影响"]
    style_header(ws_block, block_headers, fill_color="C00000")
    for row_idx, case in enumerate([case for case in CASES if case["是否阻塞"] == "是"], 2):
        values = [
            case["用例ID"],
            case["模块"],
            case["场景"],
            case["优先级"],
            case["备注"] or case["场景"],
            "失败将影响英文报价单导出主流程、数据准确性或后续验收。",
        ]
        for col_idx, value in enumerate(values, 1):
            ws_block.cell(row=row_idx, column=col_idx, value=value).alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx in range(1, 7):
        ws_block.column_dimensions[get_column_letter(col_idx)].width = 24 if col_idx < 5 else 36
    ws_block.freeze_panes = "A2"

    ws_priority = wb.create_sheet("优先级说明")
    style_header(ws_priority, ["字段", "定义"], fill_color="7030A0")
    priority_rows = [
        ("P0", "核心功能/主流程/数据正确性；必须首轮执行"),
        ("P1", "重要分支、权限、筛选、补偿"),
        ("P2", "边界、UI、非关键异常；回归轮次"),
        ("是否阻塞=是", "失败则后续模块测试无意义"),
        ("首轮必测=是", "第一遍必须执行；见「首轮冒烟」"),
        ("用例状态", "PASS/FAIL/BLOCK/N/A；BLOCK=环境/依赖阻塞"),
    ]
    for row in priority_rows:
        ws_priority.append(row)
    ws_priority.column_dimensions["A"].width = 18
    ws_priority.column_dimensions["B"].width = 72

    ws_cov = wb.create_sheet("覆盖检查")
    style_header(ws_cov, ["需求点", "覆盖情况", "对应用例"], fill_color="548235")
    coverage = [
        ("询价单TAB页创建时间、报价单号、当前操作人", "已覆盖", "TAB-001~006"),
        ("导出按钮权限与展示条件", "已覆盖", "EXP-001~006"),
        ("联系人弹窗选择、搜索、邮箱为空、加载态", "已覆盖", "POP-001~007"),
        ("包装类型+是否打托模板选择与Sheet分组", "已覆盖", "TPL-001~009"),
        ("报价单表头字段取数与模板Notes", "已覆盖", "HDR-001~005"),
        ("商品通用字段与销售填写字段", "已覆盖", "COM-001~005"),
        ("箱装/卷装/其他包装字段与公式", "已覆盖", "BOX-001~002；ROL-001~003；OTH-001~002"),
        ("公式空值、错误值、TOTAL汇总", "已覆盖", "CAL-001~005"),
        ("文件名、样式、导出记录、JSON记录", "已覆盖", "FIL-001~005"),
        ("英文商城S单/G单包装字段、预估ICON、上下推取数", "已覆盖", "ECF-001~008"),
        ("商家后台S/G、供应商G、线下单导入模板兼容", "已覆盖", "ADM-001~005"),
        ("包装联动显隐、只读展示、模板映射、数字格式", "已覆盖", "LNK-001~005"),
        ("CRM询价单及子单详情新增商城包装字段", "已覆盖", "CRF-001~005"),
        ("商城字段到CRM详情再到英文报价单的端到端一致性", "已覆盖", "CRF-006"),
        ("CRM历史数据兼容、权限和主子单数据隔离", "已覆盖", "CRF-007~008"),
        ("权限、并发、性能和兼容性", "已覆盖", "ACC-001~004"),
        (
            "纸箱包装/卷类包装与主PRD箱装/卷装命名映射、CRM主单与子单各自承载字段范围、缺单位规格处理策略、性能SLA",
            "待确认",
            "LNK-003；CRF-003；ECF-008；ACC-002~003",
        ),
    ]
    for row in coverage:
        ws_cov.append(row)
    for col_idx, width in {1: 42, 2: 14, 3: 42}.items():
        ws_cov.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws_cov.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    global XLSX_PATH
    try:
        wb.save(XLSX_PATH)
    except PermissionError:
        XLSX_PATH = OUT_DIR / f"{TITLE}_Updated_TestCases.xlsx"
        wb.save(XLSX_PATH)


def add_topic(parent_attached: list[dict], title: str) -> dict:
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    return node


def build_xmind() -> None:
    root_attached: list[dict] = []
    for _, module_name in MODULES:
        module_cases = [case for case in CASES if case["模块"] == module_name]
        if not module_cases:
            continue
        module_node = add_topic(root_attached, module_name)
        for case in module_cases:
            title = f'{case["用例ID"]} {case["场景"]} [{case["优先级"]}]'
            if case["是否阻塞"] == "是":
                title += " 阻塞"
            add_topic(module_node["children"]["attached"], title)

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": TITLE,
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": f"{TITLE} 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "generate_export_quotation_testcases", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    global XMIND_PATH
    target_path = XMIND_PATH
    try:
        zf = zipfile.ZipFile(target_path, "w", zipfile.ZIP_DEFLATED)
    except PermissionError:
        XMIND_PATH = OUT_DIR / f"{TITLE}_Updated_TestMindMap.xmind"
        target_path = XMIND_PATH
        zf = zipfile.ZipFile(target_path, "w", zipfile.ZIP_DEFLATED)
    with zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    build_cases()
    build_excel()
    build_xmind()
    total = len(CASES)
    p0 = sum(1 for case in CASES if case["优先级"] == "P0")
    block = sum(1 for case in CASES if case["是否阻塞"] == "是")
    first = sum(1 for case in CASES if case["首轮必测"] == "是")
    print(f"generated: {XLSX_PATH}")
    print(f"generated: {XMIND_PATH}")
    print(f"total={total}, p0={p0}, block={block}, first={first}")
