# -*- coding: utf-8 -*-
"""Generate CRM UI redesign regression matrix (Excel).

Based on CRM left sidebar menus + existing Pyautotest assets.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "testcases"
OUT_PATH = OUT_DIR / "CRM_UI改版_回归矩阵.xlsx"

HEADERS = [
    "序号",
    "一级菜单",
    "二级/入口",
    "优先级",
    "是否阻塞上线",
    "首轮必测",
    "主流程（改版后行为应不变）",
    "建议关键接口/契约",
    "仓库可复用资产",
    "复用结论",
    "缺口与建议动作",
    "AI 加速方式",
    "执行方式",
    "备注",
]

# reuse: 可复用 / 部分复用 / 需新录 / 仅文档
ROWS: list[dict] = [
    {
        "menu": "首页",
        "entry": "首页",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "登录后进入 CRM；首页卡片/待办/快捷入口可打开且不报错",
        "api": "首页汇总类接口（若有）；权限菜单接口",
        "assets": "tests/test_login.py；pages/login_page.py；APP_HOME_URL",
        "reuse": "部分复用",
        "gap": "补「登录→首页可达」冒烟；断言菜单存在即可，不做视觉对比",
        "ai": "基于 LoginPage + CrmPage 生成首页可达用例骨架",
        "exec": "UI冒烟",
        "note": "改版后首页布局变化大时只验入口与关键跳转",
    },
    {
        "menu": "客户",
        "entry": "客户列表/详情",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "列表查询→打开详情→编辑基础信息→保存；跟进/活动 Tab 可切换",
        "api": "客户分页/详情/保存；活动列表（路径待抓包确认）",
        "assets": "testcases/CRM_Customer_BasicInfo_Optimization_*；CRM0403 线索历史；generate_crm_create_edit_customer_lead_*",
        "reuse": "需新录",
        "gap": "几乎无客户 API service；需录制列表/新建/编辑主路径并沉淀 CustomerService",
        "ai": "record_regression_session 录客户主路径 → 抽接口 → 生成 service+pytest",
        "exec": "录制+接口+手工P0",
        "note": "客户是改版高风险模块，优先于机会/联系人",
    },
    {
        "menu": "客户",
        "entry": "客户查重",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "输入关键信息查重→命中/未命中结果正确；可跳转已有客户",
        "api": "客户查重接口",
        "assets": "侧栏有入口；暂无自动化",
        "reuse": "需新录",
        "gap": "独立录制查重成功/失败/跳转；写入检查表",
        "ai": "录制 Network 后生成 3 条接口用例（命中/未命中/缺参）",
        "exec": "录制+接口",
        "note": "查重错误会导致重复客户，属阻塞",
    },
    {
        "menu": "销售机会",
        "entry": "机会列表/详情/新建",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "列表打开→新建机会→关联客户→保存→详情可查看",
        "api": "机会分页/保存/详情",
        "assets": "无专用自动化；CRM3.13 等文档可能有关联活动展示",
        "reuse": "需新录",
        "gap": "先菜单可达+列表加载；深测新建/阶段变更可第二轮",
        "ai": "生成菜单可达+列表冒烟；主路径录制后再扩",
        "exec": "UI冒烟+手工抽测",
        "note": "首轮保入口与列表，避免平均用力",
    },
    {
        "menu": "联系人",
        "entry": "联系人列表/详情/新建",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "列表打开→新建联系人→关联客户→保存；职务等筛选项可用",
        "api": "联系人分页/保存",
        "assets": "CRM3.13 联系人职务筛选相关用例文档",
        "reuse": "仅文档",
        "gap": "文档有规则可转检查表；自动化需新录",
        "ai": "从 CRM3.13 产物抽联系人相关点 → 检查表 + 可选录制",
        "exec": "手工P1+UI冒烟",
        "note": "与客户详情联系人子对象联动需抽测",
    },
    {
        "menu": "销售线索",
        "entry": "线索列表/新建/编辑",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "新建线索→列表可查→编辑保存；国家/来源/等级字典正常",
        "api": "/api/crm/lead/saveOrUpdate；/api/crm/lead/page；queryDicByType；country/list",
        "assets": "api/services/crm_lead_service.py；tests/test_api_create_lead*.py；recording 草稿能力",
        "reuse": "可复用",
        "gap": "接口已较全；补 1 条 UI 冒烟（菜单→列表→新建成功提示）即可",
        "ai": "在现有 pytest 上补断言/边界；UI 用录制润色定位",
        "exec": "接口回归为主+UI冒烟",
        "note": "改版期最稳底盘，优先进 Jenkins",
    },
    {
        "menu": "销售线索",
        "entry": "认领 / 分配 / 移公海",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "公海认领、分配给跟进人、移入公海后状态与跟进人正确",
        "api": "claimLead；assign；movePublicSea；effective/list",
        "assets": "tests/test_api_claim_lead.py；test_api_assign_lead.py；test_api_move_lead_public_sea.py",
        "reuse": "可复用",
        "gap": "确认改版未改接口契约；UI 按钮文案/入口变了则补定位",
        "ai": "对照 Network 校验字段；失败时 AI 读 Allure 附件初判",
        "exec": "接口回归",
        "note": "规则配置变更时需联动系统设置模块",
    },
    {
        "menu": "活动记录",
        "entry": "活动列表/新建",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "新建活动并关联线索/客户→列表可见；图片/类型筛选可用",
        "api": "/api/crm/common/activity/saveOrUpdate",
        "assets": "crm_lead_service.create_activity_record；test_api_create_lead_activity*.py；CRM3.13 活动相关用例",
        "reuse": "部分复用",
        "gap": "接口可复用；活动独立菜单页 UI 需冒烟；客户侧活动展示用文档检查表",
        "ai": "扩负向（缺关联对象/非法类型）；录制活动菜单页",
        "exec": "接口+UI冒烟+手工抽测",
        "note": "活动是跟进核心，阻塞标记保留",
    },
    {
        "menu": "拜访日程",
        "entry": "日程列表/新建/完成",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "打开日程→新建拜访→完成/取消；与客户/线索关联正确",
        "api": "拜访/日程相关接口（待抓包）",
        "assets": "CRM3.13 拜访相关用例文档",
        "reuse": "仅文档",
        "gap": "首轮：菜单可达+新建保存成功；状态流转第二轮",
        "ai": "从文档抽检查表；可选录制主路径",
        "exec": "UI冒烟+手工P1",
        "note": "侧栏高亮模块，入口可达必测",
    },
    {
        "menu": "系统设置",
        "entry": "线索分配规则",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "查看/编辑规则→保存→新线索按规则落到正确跟进人",
        "api": "规则 CRUD + 线索创建/分配结果校验",
        "assets": "线索 assign/claim API 可作结果校验；规则页无自动化",
        "reuse": "部分复用",
        "gap": "规则页需录制；结果用现有线索接口断言跟进人",
        "ai": "录制改规则操作；结果断言复用 CrmLeadService",
        "exec": "录制+接口验证",
        "note": "配置错误会污染全员线索流",
    },
    {
        "menu": "系统设置",
        "entry": "客户分配规则",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "规则保存后，客户分配/领取结果符合规则",
        "api": "客户分配规则 + 客户分页/跟进人字段",
        "assets": "无客户分配 API 封装",
        "reuse": "需新录",
        "gap": "与线索分配对称补测；优先手工+录制",
        "ai": "对照线索分配用例模板生成客户侧检查表",
        "exec": "录制+手工P0",
        "note": "与客户模块联动",
    },
    {
        "menu": "系统设置",
        "entry": "权限组管理",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "权限组增改→账号菜单/按钮显隐正确；无越权访问",
        "api": "权限/菜单配置接口",
        "assets": "无",
        "reuse": "需新录",
        "gap": "准备 2 个角色账号做菜单对比；AI 生成权限矩阵检查表",
        "ai": "按侧栏生成「角色×菜单」矩阵，人填可见性",
        "exec": "手工P0（双账号）",
        "note": "改版最易漏菜单权限，必须首轮",
    },
    {
        "menu": "系统设置",
        "entry": "线索回收规则",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "规则生效后到期线索进入公海/回收池；可与 movePublicSea 对照",
        "api": "回收规则 + lead/page 状态",
        "assets": "movePublicSea 接口；公海相关用例文档",
        "reuse": "部分复用",
        "gap": "定时回收难自动化，用手造到期数据+接口查状态",
        "ai": "生成「规则条件→预期状态」决策表",
        "exec": "手工+接口抽查",
        "note": "依赖时间条件，安排专项账号数据",
    },
    {
        "menu": "系统设置",
        "entry": "客户回收规则",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "客户回收规则保存与触发结果正确",
        "api": "客户回收规则 + 客户状态",
        "assets": "CRM3.13 客户回收相关文档点",
        "reuse": "仅文档",
        "gap": "同线索回收：决策表+抽样验证",
        "ai": "从文档抽规则边界",
        "exec": "手工P1",
        "note": "",
    },
    {
        "menu": "系统设置",
        "entry": "数据共享规则",
        "priority": "P1",
        "block": "否",
        "first": "否",
        "flow": "共享范围变更后，列表可见数据范围变化符合预期",
        "api": "共享规则 + 列表数据权限",
        "assets": "无",
        "reuse": "需新录",
        "gap": "双账号可见性对比；第二轮深入",
        "ai": "生成共享场景矩阵（本人/本部门/全部）",
        "exec": "手工抽测",
        "note": "首轮可抽样，非全量",
    },
    {
        "menu": "系统设置",
        "entry": "用户企微绑定",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "绑定/解绑入口可用；绑定态展示正确",
        "api": "企微绑定相关接口",
        "assets": "generate_crm_wework_test_artifacts.py；cases_ub_binding.py 等",
        "reuse": "仅文档",
        "gap": "有用例产物，转检查表执行；自动化可选",
        "ai": "从已有 Excel 抽「改版后入口仍在」子集",
        "exec": "手工按已有用例",
        "note": "依赖企微环境，失败常为环境问题",
    },
    {
        "menu": "系统设置",
        "entry": "企微好友池",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "好友池列表打开；来源/同步状态可查",
        "api": "好友池列表接口",
        "assets": "cases_fp_friend_pool.py；二维码/好友池相关文档",
        "reuse": "仅文档",
        "gap": "菜单可达+列表加载；深测跟同步任务",
        "ai": "从 friend pool 用例抽首轮子集",
        "exec": "手工P1+UI冒烟",
        "note": "",
    },
    {
        "menu": "线索看板",
        "entry": "线索看板",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "进入看板→默认指标展示→筛选/时间范围→下钻或导出（若有）",
        "api": "看板统计接口",
        "assets": "cases_lead_dashboard.py；generate_crm_lead_dashboard_*；CRM_Lead_Dashboard_TestCases*",
        "reuse": "仅文档",
        "gap": "用例很全，改版期执行「菜单/权限/主指标抽样」子集即可",
        "ai": "从已有用例筛 P0/首轮必测行生成执行清单",
        "exec": "手工按子集",
        "note": "数据对账成本高，首轮不做全量对账",
    },
    {
        "menu": "客户跟进看板",
        "entry": "客户跟进看板",
        "priority": "P1",
        "block": "否",
        "first": "是",
        "flow": "进入看板→指标与筛选可用→与客户跟进数据不严重矛盾",
        "api": "跟进看板统计接口",
        "assets": "generate_crm_customer_follow_dashboard_test_artifacts.py",
        "reuse": "仅文档",
        "gap": "同线索看板：入口+筛选+抽样指标",
        "ai": "抽首轮子集检查表",
        "exec": "手工按子集",
        "note": "",
    },
    {
        "menu": "横切",
        "entry": "全局：登录/鉴权/菜单",
        "priority": "P0",
        "block": "是",
        "first": "是",
        "flow": "登录成功；token 有效；侧栏菜单按权限完整展示；无 404",
        "api": "member/login；菜单/权限接口",
        "assets": "auth_service；test_login*.py；test_api_negative.py（无 token 等）",
        "reuse": "可复用",
        "gap": "补「侧栏全菜单点击可达」一条 UI 冒烟",
        "ai": "按本矩阵菜单列表生成 CrmPage 循环点击用例",
        "exec": "UI冒烟+接口负向",
        "note": "所有模块前置；改版后第一天先跑",
    },
    {
        "menu": "横切",
        "entry": "全局：录制回流能力",
        "priority": "P1",
        "block": "否",
        "first": "否",
        "flow": "对空白模块用录制生成 UI+接口草稿并人工润色入库",
        "api": "会话中捕获的业务 API",
        "assets": "recording/*；scripts/record_regression_session.py",
        "reuse": "可复用",
        "gap": "先跑通客户、机会各 1 次录制，沉淀模板",
        "ai": "draft_generator 出稿 → Cursor 按项目风格收口",
        "exec": "工具建设",
        "note": "支撑后续补覆盖，不直接挡上线",
    },
]


SUMMARY_HEADERS = [
    "维度",
    "结论",
    "说明",
]

SUMMARY_ROWS = [
    ("环境入口", "已具备", "BASE_URL/PLATFORM_BASE_URL = test-auth / test-platform.ysbpack.com"),
    ("UI 自动化", "覆盖低", "仅登录 + 通用 CrmPage.open_menu；无模块级 Page Object"),
    ("接口自动化", "线索强、其余弱", "CrmLeadService 覆盖创建/查询/活动/认领/分配/公海"),
    ("手工用例资产", "较丰富", "客户基础信息、看板、企微、展会、询价等 Excel/XMind 可转检查表"),
    ("录制能力", "已具备", "record_regression_session 可同时出 UI+主接口草稿"),
    ("CI", "线索 API 可进", "Jenkinsfile.crm-api / run_crm_api_jenkins.bat"),
    ("改版策略", "接口保业务 + 菜单冒烟 + 文档子集手工", "避免全量重写 UI 脚本"),
]

PLAN_HEADERS = ["周次", "目标", "交付物", "负责方式", "完成标准"]
PLAN_ROWS = [
    (
        "第 1 周",
        "菜单可达 + 风险签收",
        "本矩阵评审通过；侧栏全菜单 UI 冒烟用例",
        "AI 生成冒烟骨架，人补权限账号",
        "所有一级/二级入口可打开或明确无权限原因",
    ),
    (
        "第 1–2 周",
        "P0 接口与线索保底",
        "线索 API 回归进 CI；客户查重/客户主路径录制 1 次",
        "复用 CrmLeadService；客户模块新录",
        "P0 阻塞项有接口或检查表执行记录",
    ),
    (
        "第 2 周",
        "系统设置规则联动",
        "分配/权限组检查表；规则变更→线索结果用例",
        "规则页手工/录制，结果走接口断言",
        "分配与权限无越权、无错分",
    ),
    (
        "第 2–3 周",
        "P1 抽样与看板子集",
        "机会/联系人/拜访冒烟；两看板首轮子集",
        "从已有 Excel 筛首轮必测",
        "P1 入口可用，抽样主流程通过",
    ),
    (
        "上线前",
        "探索与放行",
        "失败清单+风险签收；可选视觉抽样",
        "人做放行决策，AI 辅助归因",
        "阻塞项清零或有明确豁免",
    ),
]

MENU_SMOKE_HEADERS = ["顺序", "菜单路径", "操作", "通过标准", "优先级"]
MENU_SMOKE_ROWS = [
    (1, "首页", "点击「首页」", "页面加载成功，无空白/报错弹窗", "P1"),
    (2, "客户", "点击「客户」", "客户列表出现或明确空态", "P0"),
    (3, "客户 / 客户查重", "点击「客户查重」", "查重页打开，可输入查询", "P0"),
    (4, "销售机会", "点击「销售机会」", "机会列表加载成功", "P1"),
    (5, "联系人", "点击「联系人」", "联系人列表加载成功", "P1"),
    (6, "销售线索", "点击「销售线索」", "线索列表加载成功", "P0"),
    (7, "活动记录", "点击「活动记录」", "活动列表加载成功", "P0"),
    (8, "拜访日程", "点击「拜访日程」", "日程页加载成功", "P1"),
    (9, "系统设置 / 线索分配规则", "展开设置并进入", "规则页可打开", "P0"),
    (10, "系统设置 / 客户分配规则", "进入", "规则页可打开", "P0"),
    (11, "系统设置 / 权限组管理", "进入", "权限组列表可打开", "P0"),
    (12, "系统设置 / 线索回收规则", "进入", "规则页可打开", "P1"),
    (13, "系统设置 / 客户回收规则", "进入", "规则页可打开", "P1"),
    (14, "系统设置 / 数据共享规则", "进入", "规则页可打开", "P1"),
    (15, "系统设置 / 用户企微绑定", "进入", "绑定页可打开", "P1"),
    (16, "系统设置 / 企微好友池", "进入", "好友池列表可打开", "P1"),
    (17, "线索看板", "点击「线索看板」", "看板加载或权限提示明确", "P1"),
    (18, "客户跟进看板", "点击「客户跟进看板」", "看板加载或权限提示明确", "P1"),
]


def _style_header(ws, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, size=11)
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = thin
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _fill_reuse(cell, value: str) -> None:
    colors = {
        "可复用": "C6EFCE",
        "部分复用": "FFE699",
        "仅文档": "BDD7EE",
        "需新录": "FFC7CE",
    }
    cell.fill = PatternFill("solid", fgColor=colors.get(value, "FFFFFF"))


def _fill_priority(cell, value: str) -> None:
    colors = {"P0": "FF6B6B", "P1": "FFA94D", "P2": "74C0FC"}
    if value in colors:
        cell.fill = PatternFill("solid", fgColor=colors[value])
        cell.font = Font(bold=True)


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1: matrix
    ws = wb.active
    ws.title = "回归矩阵"
    _style_header(ws, HEADERS)
    body_align = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for idx, row in enumerate(ROWS, 1):
        values = [
            idx,
            row["menu"],
            row["entry"],
            row["priority"],
            row["block"],
            row["first"],
            row["flow"],
            row["api"],
            row["assets"],
            row["reuse"],
            row["gap"],
            row["ai"],
            row["exec"],
            row["note"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(idx + 1, col, value)
            cell.alignment = body_align
            cell.border = thin
            if col == 4:
                _fill_priority(cell, str(value))
            if col == 10:
                _fill_reuse(cell, str(value))
        ws.row_dimensions[idx + 1].height = 48
    _autosize(
        ws,
        {
            1: 6,
            2: 12,
            3: 22,
            4: 8,
            5: 12,
            6: 10,
            7: 36,
            8: 34,
            9: 40,
            10: 12,
            11: 36,
            12: 32,
            13: 18,
            14: 28,
        },
    )

    # Sheet 2: summary
    ws2 = wb.create_sheet("覆盖现状")
    _style_header(ws2, SUMMARY_HEADERS)
    for r, row in enumerate(SUMMARY_ROWS, 2):
        for c, value in enumerate(row, 1):
            cell = ws2.cell(r, c, value)
            cell.alignment = body_align
            cell.border = thin
    _autosize(ws2, {1: 14, 2: 28, 3: 70})

    # Sheet 3: plan
    ws3 = wb.create_sheet("落地节奏")
    _style_header(ws3, PLAN_HEADERS)
    for r, row in enumerate(PLAN_ROWS, 2):
        for c, value in enumerate(row, 1):
            cell = ws3.cell(r, c, value)
            cell.alignment = body_align
            cell.border = thin
            ws3.row_dimensions[r].height = 40
    _autosize(ws3, {1: 12, 2: 22, 3: 42, 4: 32, 5: 36})

    # Sheet 4: menu smoke checklist
    ws4 = wb.create_sheet("菜单可达检查表")
    _style_header(ws4, MENU_SMOKE_HEADERS)
    for r, row in enumerate(MENU_SMOKE_ROWS, 2):
        for c, value in enumerate(row, 1):
            cell = ws4.cell(r, c, value)
            cell.alignment = body_align
            cell.border = thin
            if c == 5:
                _fill_priority(cell, str(value))
    # extra columns for execution
    ws4.cell(1, 6, "实际结果")
    ws4.cell(1, 7, "执行人")
    ws4.cell(1, 8, "备注")
    for col in range(6, 9):
        cell = ws4.cell(1, col)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws4.auto_filter.ref = "A1:H1"
    _autosize(ws4, {1: 8, 2: 28, 3: 22, 4: 36, 5: 10, 6: 12, 7: 10, 8: 24})

    # Sheet 5: legend
    ws5 = wb.create_sheet("图例与统计")
    legend = [
        ("字段", "含义"),
        ("可复用", "已有自动化，改版后优先跑通/微调定位或契约"),
        ("部分复用", "接口或文档可复用一部分，仍需补 UI/新接口"),
        ("仅文档", "有 Excel/XMind/cases，转检查表人工执行"),
        ("需新录", "建议用 record_regression_session 录主路径再沉淀"),
        ("P0", "阻塞或核心业务，首轮必测"),
        ("P1", "重要但可抽样，首轮保入口与主流程"),
    ]
    for r, row in enumerate(legend, 1):
        for c, value in enumerate(row, 1):
            cell = ws5.cell(r, c, value)
            cell.alignment = body_align
            if r == 1:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(color="FFFFFF", bold=True)
            elif c == 1 and r > 1:
                if value in {"可复用", "部分复用", "仅文档", "需新录"}:
                    _fill_reuse(cell, value)
                if value in {"P0", "P1"}:
                    _fill_priority(cell, value)

    # stats
    from collections import Counter

    reuse_counter = Counter(r["reuse"] for r in ROWS)
    prio_counter = Counter(r["priority"] for r in ROWS)
    block_n = sum(1 for r in ROWS if r["block"] == "是")
    first_n = sum(1 for r in ROWS if r["first"] == "是")
    ws5.cell(10, 1, "统计")
    ws5.cell(10, 2, "数量")
    ws5["A10"].fill = PatternFill("solid", fgColor="1F4E79")
    ws5["B10"].fill = PatternFill("solid", fgColor="1F4E79")
    ws5["A10"].font = Font(color="FFFFFF", bold=True)
    ws5["B10"].font = Font(color="FFFFFF", bold=True)
    stats = [
        ("矩阵条目总数", len(ROWS)),
        ("P0", prio_counter.get("P0", 0)),
        ("P1", prio_counter.get("P1", 0)),
        ("阻塞上线=是", block_n),
        ("首轮必测=是", first_n),
        ("可复用", reuse_counter.get("可复用", 0)),
        ("部分复用", reuse_counter.get("部分复用", 0)),
        ("仅文档", reuse_counter.get("仅文档", 0)),
        ("需新录", reuse_counter.get("需新录", 0)),
    ]
    for r, (k, v) in enumerate(stats, 11):
        ws5.cell(r, 1, k)
        ws5.cell(r, 2, v)
    _autosize(ws5, {1: 18, 2: 70})

    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Generated: {path}")
