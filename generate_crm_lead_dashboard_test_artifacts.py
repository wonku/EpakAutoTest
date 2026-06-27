# -*- coding: utf-8 -*-
"""CRM 线索看板：功能测试用例 Excel + XMind（以需求/图一/图二为准）。"""
import json
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cases_lead_dashboard import (
    CASES_CONFIG,
    CASES_DATA,
    CASES_E2E,
    CASES_EMP,
    CASES_FILTER,
    CASES_LD,
    CASES_MENU,
    CASES_SEA,
    CASES_SOURCE,
    CASES_UI,
)

OUT_DIR = Path(__file__).resolve().parent / "testcases"
OUT_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = OUT_DIR / "CRM_Lead_Dashboard_TestCases.xlsx"
XMIND_PATH = OUT_DIR / "CRM_Lead_Dashboard_TestMindMap.xmind"

ROOT_XLSX = Path(__file__).resolve().parent / "CRM_Lead_Dashboard_TestCases.xlsx"
ROOT_XMIND = Path(__file__).resolve().parent / "CRM_Lead_Dashboard_TestMindMap.xmind"

CASES = CASES_LD


def tid():
    return str(uuid.uuid4()).replace("-", "")[:16]


def add_topic(parent_attached, title, children_titles=None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    if children_titles:
        for ct in children_titles:
            add_topic(node["children"]["attached"], ct)
    return node


def build_xmind():
    root_attached = []

    m0 = add_topic(root_attached, f"0 菜单与权限（{len(CASES_MENU)}条）【阻塞】")
    add_topic(
        m0["children"]["attached"],
        "P0 首轮必测",
        [
            "MNU-001~003 菜单与进入",
            "MNU-004~006 特殊员工/商城号/权限过滤",
        ],
    )

    m1 = add_topic(root_attached, f"1 全局筛选区（{len(CASES_FILTER)}条）")
    for sec in [
        "FLT-001~007 分析时间范围（默认本月/近30/90/自定义）",
        "FLT-101~106 部门多选",
        "FLT-201~206 跟进人输入+多选",
        "FLT-301~304 线索所在国家0511",
        "FLT-401~404 查询/重置",
        "FLT-501~506 组件数/标签/无组件提示",
    ]:
        add_topic(m1["children"]["attached"], sec)

    m2 = add_topic(root_attached, f"2 组件配置（{len(CASES_CONFIG)}条）")
    add_topic(
        m2["children"]["attached"],
        "2.1 入口与抽屉",
        ["CFG-001 顶部按钮右侧抽屉", "CFG-002 三大类配置卡", "CFG-008 完成/X关闭"],
    )
    add_topic(
        m2["children"]["attached"],
        "2.1 开关与快捷操作",
        ["CFG-003 大类总开关", "CFG-004 指标卡/图表/明细", "CFG-006/007 全选/取消全选"],
    )
    add_topic(
        m2["children"]["attached"],
        "2.1 主页面联动",
        ["CFG-010~012 N/9与chips实时联动", "CFG-015/016 全关0/9与提示"],
    )
    add_topic(
        m2["children"]["attached"],
        "2.1 持久化与用户维度",
        ["CFG-013 用户互不影响", "CFG-014 刷新保留", "CFG-019 重置布局"],
    )
    add_topic(
        m2["children"]["attached"],
        "2.1.7 有组件无数据空态",
        ["CFG-023 指标卡0/--", "CFG-024 图表空态", "CFG-025 明细空行", "CFG-028 与全关区分"],
    )

    m3 = add_topic(root_attached, f"3 数据准确性（{len(CASES_DATA)}条）")
    add_topic(
        m3["children"]["attached"],
        "查询结果=权限∩时间∩部门∩跟进人∩国家",
        ["DAT-001 默认对账", "DAT-006 权限", "DAT-010 三类同步"],
    )
    add_topic(
        m3["children"]["attached"],
        "4.1~4.3 权限与处理顺序",
        [
            "DAT-011 复用CRM既有权限模型",
            "DAT-012 先权限过滤后统计",
            "DAT-013~014 口径一致与无权限不展示",
            "DAT-015~018 四步处理顺序校验",
        ],
    )

    m4 = add_topic(root_attached, f"4 来源与状态质量（{len(CASES_SOURCE)}条）")
    add_topic(
        m4["children"]["attached"],
        "3.1.1 指标卡（6张）",
        ["SRC-002~007 线索总数/跟进中占比/转换率/作废率/平均评分/公海存量"],
    )
    add_topic(
        m4["children"]["attached"],
        "3.1.2 图表区（6视图）",
        [
            "SRC-008~009 来源占比环图+右侧列表",
            "SRC-010~013 状态分布堆叠柱（0511倒序+滚动）",
            "SRC-014~015 活动类型环图+列表",
            "SRC-016~018 来源等级100%堆叠（tooltip含百分比与原始数）",
            "SRC-019~020 展会线索占比（2.69展会名称）",
            "SRC-021~023 展会等级分布（线下展会子集）",
        ],
    )
    add_topic(
        m4["children"]["attached"],
        "3.1.3 明细表",
        ["SRC-024 表名", "SRC-025~028 字段/计算/交叉校验"],
    )
    add_topic(
        m4["children"]["attached"],
        "联动与空态",
        ["SRC-029 全局筛选同步", "SRC-030 空数据展示规则"],
    )

    m5 = add_topic(root_attached, f"5 人员处理与工作量（{len(CASES_EMP)}条）")
    add_topic(
        m5["children"]["attached"],
        "3.2 口径：人员责任周期（不按lead_id去重）",
        ["EMP-002~003 责任周期口径与多次计入"],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.1 指标卡（4张）",
        [
            "EMP-004 累计被分配=责任周期数",
            "EMP-005 已处理=有有效人工跟进的责任周期",
            "EMP-006~008 首跟及时率(24h) 分子/分母/比值",
            "EMP-009 转化为客户数（按转换时责任人）",
            "EMP-010 处理率口径",
            "EMP-011~015 平均首跟时长（含三种情形+work_hours）",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.2 图表区（柱状+折线）",
        [
            "EMP-016~020 柱状：X轴全员+不截断+滚动+双序列+tooltip(承接/已处理/处理率)",
            "EMP-021~025 折线：X轴全员+点值含未跟进+tooltip(均值/承接/未首跟次数)+滚动",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 明细表",
        [
            "EMP-026~027 字段完整性+承接次数命名",
            "EMP-028~031 处理率/首跟及时率/平均首跟时长/分母为0显示--",
            "EMP-032 主动移入按操作人，系统自动单列",
            "EMP-033~034 转客户/活动总数(电/微)",
            "EMP-035 行内导出按钮",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 导出—入口与权限",
        [
            "EMP-036 顶部批量导出",
            "EMP-037 单独权限配置（无权限拦截）",
            "EMP-038 不影响查看",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 导出—对象与时间",
        [
            "EMP-039~041 导出对象：全部员工/指定人员/责任周期明细",
            "EMP-042~047 导出时间：默认跟随看板/自定义/最大90天/边界/超限/联动/异常",
            "EMP-048~049 弹窗取消/确认与文件下载",
            "EMP-050 行内单人导出仅截至当前日期&跟进人=该人员",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 导出—责任周期交集逻辑",
        [
            "EMP-051~053 与筛选时段有交集即计入/跨段/完全在外",
            "EMP-054 同一线索多周期分别多条",
            "EMP-055 不同员工各自一条",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 责任周期结束类型（6类）",
        [
            "EMP-056 当前持有=当前时间",
            "EMP-057 分配=转出时间",
            "EMP-058 主动移入公海=移入时间",
            "EMP-059 系统回收=系统回收时间",
            "EMP-060 转客户=转客户时间",
            "EMP-061 作废=作废时间",
            "EMP-062 枚举完备性",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "3.2.3 导出字段（28列）",
        [
            "EMP-063 列头与模板sheet2一致",
            "EMP-064~073 各字段口径",
        ],
    )
    add_topic(
        m5["children"]["attached"],
        "联动/权限/空态",
        [
            "EMP-074 全局筛选同步刷新",
            "EMP-075 导出与筛选一致",
            "EMP-076 权限不可越权",
            "EMP-077~079 空态/无数据导出/大数据性能",
        ],
    )

    m6 = add_topic(root_attached, f"6 公海运营与流转（{len(CASES_SEA)}条）")
    add_topic(
        m6["children"]["attached"],
        "3.3.1 指标卡（5张）",
        ["SEA-002~006 手动移入/自动回收/本期认领/平均停留/当前存量", "SEA-007 认领率口径"],
    )
    add_topic(
        m6["children"]["attached"],
        "3.3.2 图表区（2视图）",
        [
            "SEA-008~009 本期移入vs认领趋势（时间范围全展示）",
            "SEA-010~011 当前公海来源占比环图",
        ],
    )
    add_topic(
        m6["children"]["attached"],
        "3.3.3 明细表",
        ["SEA-012 表名", "SEA-013~018 字段/计算/边界/交叉校验"],
    )
    add_topic(
        m6["children"]["attached"],
        "联动与空态",
        ["SEA-019 全局筛选同步", "SEA-020 空数据展示规则"],
    )

    m7 = add_topic(root_attached, f"7 交互与兼容（{len(CASES_UI)}条）")
    m8 = add_topic(root_attached, f"8 端到端（{len(CASES_E2E)}条）")
    add_topic(m8["children"]["attached"], "建议 P0 冒烟通过后再执行")

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "CRM线索看板",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "CRM线索看板 测试脑图（需求/图一图二｜含阻塞与首轮）",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "CRMLeadDashboardGenerator", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def _style_header(ws, headers, fill_color="4472C4"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(wrap_text=True, vertical="center")


def _write_cases_sheet(ws, cases, headers):
    _style_header(ws, headers)
    block_fill = PatternFill("solid", fgColor="FFC7CE")
    p0_fill = PatternFill("solid", fgColor="FFEB9C")
    for i, row in enumerate(cases, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if len(row) > 4 and row[4] == "是":
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = block_fill
        elif len(row) > 3 and row[3] == "P0":
            for col in range(1, len(headers) + 1):
                cur = ws.cell(row=i, column=col)
                if not cur.fill or getattr(cur.fill, "fgColor", None) is None:
                    cur.fill = p0_fill
                elif getattr(cur.fill.fgColor, "rgb", "00000000") in ("00000000", None):
                    cur.fill = p0_fill
    for col in range(1, len(headers) + 1):
        if col in (6,):
            ws.column_dimensions[get_column_letter(col)].width = 12
        elif col in (7, 8, 9):
            ws.column_dimensions[get_column_letter(col)].width = 42
        elif col == len(headers) - 1:
            ws.column_dimensions[get_column_letter(col)].width = 28
        elif col == len(headers):
            ws.column_dimensions[get_column_letter(col)].width = 14
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _rows_from_cases(case_list):
    rows = []
    for c in case_list:
        module, cid, title, pri, blocking, first, pre, steps, expect, note, api_hint = c
        rows.append(
            (
                module,
                cid,
                title,
                pri,
                blocking,
                first,
                pre,
                steps,
                expect,
                note,
                api_hint or "",
                "",
                "",
            )
        )
    return rows


def _add_result_validation(ws, result_col_idx, start_row=2):
    if ws.max_row < start_row:
        return
    formula = '"PASS,FAIL,BLOCK"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    rng = f"{get_column_letter(result_col_idx)}{start_row}:{get_column_letter(result_col_idx)}{ws.max_row}"
    dv.add(rng)


def build_excel():
    wb = Workbook()
    headers = [
        "模块",
        "用例编号",
        "用例标题",
        "优先级",
        "是否阻塞",
        "首轮必测",
        "前置条件",
        "测试步骤",
        "预期结果",
        "备注",
        "关联接口/文档",
        "实际测试结果备注",
        "测试结果",
    ]

    ws0 = wb.active
    ws0.title = "需求追溯"
    ws0["A1"] = "需求背景"
    ws0["B1"] = (
        "CRM 新增「线索看板」：合并「来源与状态质量」「人员处理与工作量」「公海运营与流转」"
        "三类报表；全局筛选对指标卡/图表/明细统一生效；支持组件配置。"
    )
    ws0["A2"] = "菜单"
    ws0["B2"] = "CRM-线索看板（需菜单权限；公司特殊员工可配置全量数据权限）"
    ws0["A3"] = "全局筛选（图一/图二）"
    ws0["B3"] = (
        "分析时间范围（默认本月；近30/90/自定义；图二备注近7/60/90）；"
        "部门（多选，权限内）；跟进人（输入关键词+多选）；"
        "线索所在国家（中国大陆/海外，0511）；查询/重置；"
        "当前显示组件数/9；已选组件标签；无组件提示文案"
    )
    ws0["A4"] = "数据逻辑"
    ws0["B4"] = "查询结果 = 权限过滤后数据集 ∩ 时间 ∩ 部门 ∩ 跟进人 ∩ 国家"
    ws0["A5"] = "验收账号"
    ws0["B5"] = "商城账户 18143073992 需配置特殊员工全量看板权限"
    ws0["A6"] = "首轮测试建议"
    ws0["B6"] = (
        "按 Sheet「首轮冒烟」执行：0菜单权限→1筛选→2组件配置→3数据对账→4/5/6模块→8 E2E。"
        "粉色行=阻塞；失败则暂停并提缺陷。"
    )
    ws0["A7"] = "测试数据"
    ws0["B7"] = (
        "见 Sheet「测试数据准备」：全量/普通/无权限账号；跨部门跨时间线索；"
        "大陆/海外；公海流转与首跟事件；开发基准对账表。"
    )
    ws0["A8"] = "用例统计"
    total = len(CASES)
    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    p0 = sum(1 for c in CASES if c[3] == "P0")
    ws0["B8"] = (
        f"合计{total}条（菜单{len(CASES_MENU)}+筛选{len(CASES_FILTER)}+配置{len(CASES_CONFIG)}"
        f"+数据{len(CASES_DATA)}+来源{len(CASES_SOURCE)}+人员{len(CASES_EMP)}"
        f"+公海{len(CASES_SEA)}+交互{len(CASES_UI)}+E2E{len(CASES_E2E)}）；"
        f"阻塞{block}条；首轮必测{first}条；P0共{p0}条"
    )
    ws0["A9"] = "组件配置2.1"
    ws0["B9"] = (
        "顶部组件配置→右侧抽屉；三大类×(指标卡/图表/明细)；全选/取消全选；"
        "主页面 N/9 与 chips 联动；按用户保存；全关提示；"
        "有组件无数据时 KPI=0/--、图表空态、明细空行。"
    )
    ws0["A10"] = "说明"
    ws0["B10"] = "用例以需求说明与图一、图二为准；HTML 原型仅作动态效果参考，不作为验收标准。"
    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 78
    for r in range(1, 11):
        ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    rows_all = _rows_from_cases(CASES)
    ws_all = wb.create_sheet("测试用例")
    _write_cases_sheet(ws_all, rows_all, headers)
    _add_result_validation(ws_all, len(headers))

    def rows_for_module(prefix):
        return [r for r in rows_all if r[0].startswith(prefix)]

    for sheet_name, prefix in [
        ("0-菜单与权限", "0-"),
        ("1-全局筛选区", "1-"),
        ("2-组件配置", "2-"),
        ("3-数据准确性", "3-"),
        ("4-来源与状态质量", "4-"),
        ("5-人员处理与工作量", "5-"),
        ("6-公海运营与流转", "6-"),
        ("7-交互与兼容", "7-"),
        ("8-端到端", "8-"),
    ]:
        ws = wb.create_sheet(sheet_name)
        _write_cases_sheet(ws, rows_for_module(prefix), headers)
        _add_result_validation(ws, len(headers))

    ws_smoke = wb.create_sheet("首轮冒烟")
    smoke = [r for r in rows_all if r[5] == "是"]
    _write_cases_sheet(ws_smoke, smoke, headers)
    _add_result_validation(ws_smoke, len(headers))

    ws_block = wb.create_sheet("阻塞场景清单")
    block_headers = ["模块", "用例编号", "用例标题", "优先级", "阻塞说明", "失败影响"]
    _style_header(ws_block, block_headers, fill_color="C00000")
    block_impact = {
        "MNU": "无法进入看板或权限模型错误→数据不可信",
        "FLT": "筛选不可用或数据集公式错误→全页统计失效",
        "CFG": "无法打开配置/开关失效→自定义看板与空态无法验证",
        "DAT": "数据对账失败→看板不可验收",
        "SRC": "来源模块与全局筛选联动失败",
        "EMP": "人员模块与全局筛选联动失败",
        "SEA": "公海模块与全局筛选联动失败",
        "UI": "体验问题，一般不阻塞主流程",
        "E2E": "全链路验收失败",
    }
    i = 2
    for c in [x for x in CASES if x[4] == "是"]:
        prefix = c[1].split("-")[0]
        impact = block_impact.get(prefix, "阻塞关联模块测试")
        vals = [c[0], c[1], c[2], c[3], c[9] or "见用例步骤", impact]
        for col, v in enumerate(vals, 1):
            cell = ws_block.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        i += 1
    for col in range(1, 7):
        ws_block.column_dimensions[get_column_letter(col)].width = 22 if col == 5 else 18
    ws_block.freeze_panes = "A2"

    ws_data = wb.create_sheet("测试数据准备")
    data_headers = ["编号", "类型", "说明", "数量/要求", "用途"]
    _style_header(ws_data, data_headers, fill_color="548235")
    data_rows = [
        ("ACC-01", "账号", "特殊员工全量看板数据权限", "1", "基准对账、全量验收"),
        ("ACC-02", "账号", "商城账户 18143073992", "1", "需求指定验收账号"),
        ("ACC-03", "账号", "部门负责人（权限内多部门）", "1", "部门筛选、权限边界"),
        ("ACC-04", "账号", "普通销售（仅本人/本部门）", "1", "权限过滤、跟进人筛选"),
        ("ACC-05", "账号", "无菜单权限", "1", "菜单拦截"),
        ("ACC-06", "账号", "有菜单无数据", "1", "空态"),
        ("ORG-01", "组织", "部门树≥3级", "销售一部/二部等", "部门多选"),
        ("ORG-02", "人员", "重名员工", "2人同名", "跟进人匹配"),
        ("LEAD-01", "线索", "跨时间分布", "≥200", "本月/近30/90/自定义边界"),
        ("LEAD-02", "线索", "跨部门", "每部门≥50", "部门筛选"),
        ("LEAD-03", "线索", "跨跟进人", "每人≥30", "跟进人筛选"),
        ("LEAD-04", "线索", "国家", "大陆/海外各≥40", "0511国家筛选"),
        ("LEAD-05", "线索", "多来源多状态", "多来源≥10/状态", "来源模块图表"),
        ("LEAD-06", "线索", "等级 A/B/C/D/空", "均有", "等级分布图"),
        ("LEAD-07", "事件", "公海流转", "≥80", "公海模块"),
        ("LEAD-08", "事件", "分配/首跟/转化", "≥100", "人员模块 SLA"),
        ("LEAD-09", "活动", "电/微等类型", "≥50", "活动类型统计"),
        ("BASE-01", "基准", "开发/BI 对账导出", "≥6组筛选条件", "DAT-001~010"),
    ]
    for ri, row in enumerate(data_rows, 2):
        for ci, v in enumerate(row, 1):
            ws_data.cell(row=ri, column=ci, value=v).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    for col in range(1, 6):
        ws_data.column_dimensions[get_column_letter(col)].width = 18 if col < 3 else 28
    ws_data.freeze_panes = "A2"

    ws_pri = wb.create_sheet("优先级说明")
    ws_pri.append(["优先级", "定义", "首轮策略"])
    ws_pri.append(["P0", "核心功能/主流程/数据正确性", "必须首轮全部执行"])
    ws_pri.append(["P1", "重要分支、筛选组合、图表明细", "首轮时间允许则执行"])
    ws_pri.append(["P2", "边界、UI细节", "回归轮次执行"])
    ws_pri.append(["阻塞=是", "失败则后续模块测试无意义或结果不可信", "失败即停，修复后从该模块重测"])
    ws_pri.append(["准入", "阻塞全过+首轮冒烟全过+DAT-001/006/007与基准零差异", "提测门禁"])

    wb.save(XLSX_PATH)


def main():
    build_excel()
    build_xmind()
    import shutil

    print(f"Generated {len(CASES)} cases ->")
    print(" ", XLSX_PATH)
    print(" ", XMIND_PATH)
    for src, dst in ((XLSX_PATH, ROOT_XLSX), (XMIND_PATH, ROOT_XMIND)):
        try:
            shutil.copy2(src, dst)
            print(" ", dst)
        except OSError as e:
            print(f"  (skip copy to {dst.name}: {e})")


if __name__ == "__main__":
    main()
