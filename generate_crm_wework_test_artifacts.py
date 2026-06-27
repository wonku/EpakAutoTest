# -*- coding: utf-8 -*-
"""CRM-企业微信对接：功能测试用例 Excel + XMind（含阻塞场景与首轮冒烟）。"""
import json
import uuid
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cases_ub_binding import CASES_UB
from cases_qr_management import CASES_QR
from cases_auto_sync import CASES_AUTO
from cases_fp_friend_pool import CASES_FP, CASES_MAP


OUT_DIR = Path(__file__).resolve().parent / "testcases"
OUT_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = OUT_DIR / "CRM_WeWork_Integration_TestCases.xlsx"
XMIND_PATH = OUT_DIR / "CRM_WeWork_Integration_TestMindMap.xmind"

# 用例元组: module, case_id, title, priority, blocking, first_round, pre, steps, expect, note
# blocking: 是/否 — 阻塞则后续模块无法有效验证
# first_round: 是/否 — 建议第一遍测试执行
CASES_PRE = [
    # ========== 0. 前置：企微应用与 CRM 对接（全局阻塞） ==========
    (
        "0-企微应用与CRM对接",
        "PRE-001",
        "CRM录入CorpId/AgentId/Secret并连接测试成功",
        "P0",
        "是",
        "是",
        "具备企微管理员；已创建自建应用",
        "1.企微后台创建自建应用，记录CorpId、AgentId、Secret\n2.CRM企微配置页录入并保存\n3.点击「测试连接」或触发gettoken",
        "连接成功；失败时有明确错误码/文案（非空白）",
        "全链路前置；失败则所有同步不可用",
        "gettoken",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-002",
        "通讯录只读权限开启—员工同步90196可用",
        "P0",
        "是",
        "是",
        "应用Secret有效",
        "1.关闭通讯录只读权限\n2.CRM执行员工同步\n3.重新开启后同步",
        "步骤2失败且提示权限相关；步骤3可拉取员工列表",
        "阻塞「用户企微绑定」同步",
        "90196",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-003",
        "外部联系人权限—externalcontact/get可用",
        "P0",
        "是",
        "是",
        "客户联系功能已开通",
        "1.关闭外部联系人相关API权限\n2.触发好友详情拉取\n3.恢复权限重试",
        "步骤2失败；步骤3成功返回联系人详情",
        "阻塞好友池字段补全与自动建线索",
        "92114 externalcontact/get",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-004",
        "企业可信IP白名单—外网API不被拒",
        "P0",
        "是",
        "是",
        "已知CRM出口公网IP",
        "1.白名单清空或填错IP\n2.调用需外网的企微API\n3.填入正确IP",
        "步骤2返回IP限制类错误；步骤3调用成功",
        "常见上线阻塞项",
        "",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-005",
        "接收事件服务器GET校验通过",
        "P0",
        "是",
        "是",
        "已配置URL、Token、EncodingAESKey且与CRM一致",
        "1.企微后台保存/验证回调URL\n2.查看CRM日志",
        "校验通过；CRM按协议回显；无500",
        "阻塞扫码自动入库",
        "92130 回调",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-006",
        "回调POST解密—add_external_contact事件可解析",
        "P0",
        "是",
        "是",
        "已订阅add_external_contact、change_external_contact",
        "1.真实或模拟加好友触发POST\n2.核对解密后UserID、ExternalUserID、CreateTime、State\n3.构造错误签名",
        "步骤2字段完整进入处理队列；步骤3拒绝并记日志",
        "阻塞自动同步线索",
        "92130",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-007",
        "Secret重置后CRM未更新时的失败提示",
        "P1",
        "否",
        "否",
        "已完成对接",
        "1.企微重置Secret\n2.不更新CRM触发同步\n3.更新Secret后重试",
        "步骤2鉴权失败提示明确；步骤3恢复",
        "",
        "gettoken",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-008",
        "回调Token/AESKey与CRM不一致时校验失败",
        "P0",
        "是",
        "是",
        "可改配置",
        "1.故意错Token保存\n2.仅错EncodingAESKey\n3.全部一致后保存",
        "步骤1/2企微校验失败；步骤3通过",
        "",
        "",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-009",
        "未订阅加好友事件时业务漏单可发现",
        "P0",
        "是",
        "是",
        "可改事件订阅",
        "1.取消add_external_contact订阅\n2.加好友\n3.恢复订阅再加好友",
        "步骤2不产生好友池/线索任务；步骤3恢复",
        "",
        "92130",
    ),
    (
        "0-企微应用与CRM对接",
        "PRE-010",
        "回调重复投递幂等—同external_userid不重复建线索",
        "P0",
        "是",
        "是",
        "可重复推送相同事件",
        "1.短时间重复相同MsgId或相同业务键\n2.查线索数与活动数",
        "不产生重复线索；日志可追踪去重",
        "",
        "",
    ),
]

CASES_E2E = [
    # ========== 6. 端到端 ==========
    (
        "6-端到端",
        "E2E-001",
        "全链路：配置→同步员工→绑定→扫码→好友池→建线索",
        "P0",
        "是",
        "是",
        "测试环境完整；测试号可加好友",
        "1.完成PRE与UB绑定\n2.客户扫码加好友\n3.好友池确认数据\n4.设来源与负责人\n5.批量新建线索",
        "各环节数据一致；线索可跟进",
        "首轮最后执行",
        "",
    ),
    (
        "6-端到端",
        "E2E-002",
        "解绑后新加好友—负责人不再自动带出原商城用户",
        "P1",
        "否",
        "是",
        "销售曾绑定后解绑",
        "1.解绑\n2.同销售加新客户\n3.查负责人",
        "负责人不错误映射旧绑定",
        "",
        "",
    ),
    (
        "6-端到端",
        "E2E-003",
        "多销售分别添加同一客户—好友池记录策略",
        "P1",
        "否",
        "否",
        "两个已绑定销售",
        "分别添加同一external_userid",
        "符合PRD：一条或多条及跟进人规则明确",
        "",
        "",
    ),
]

CASES = CASES_PRE + CASES_UB + CASES_QR + CASES_AUTO + CASES_FP + CASES_MAP + CASES_E2E

ROOT_XLSX = Path(__file__).resolve().parent / "CRM_WeWork_Integration_TestCases.xlsx"
ROOT_XMIND = Path(__file__).resolve().parent / "CRM_WeWork_Integration_TestMindMap.xmind"


def tid():
    return str(uuid.uuid4()).replace("-", "")[:16]


def add_topic(parent_attached, title, children_titles=None):
    node = {
        "id": tid(),
        "class": "topic",
        "title": title,
        "children": {"attached": []},
    }
    parent_attached.append(node)
    if children_titles:
        for ct in children_titles:
            add_topic(node["children"]["attached"], ct)
    return node


def build_xmind():
    root_attached = []

    m0 = add_topic(root_attached, "0 前置｜企微应用与CRM对接 【阻塞】")
    add_topic(
        m0["children"]["attached"],
        "P0 首轮必测",
        [
            "PRE-001 配置与连接测试",
            "PRE-002 通讯录权限90196",
            "PRE-003 外部联系人权限92114",
            "PRE-004 企业可信IP",
            "PRE-005/006 回调GET/POST",
            "PRE-008/009/010 回调配置与幂等",
        ],
    )

    m1 = add_topic(root_attached, "1 用户企微绑定（65条详细）")
    for sec in [
        "A菜单权限 UB-A01~A05",
        "B同步企微 B01~B12",
        "C汇总统计 C01~C07",
        "D筛选搜索 D01~D11",
        "E列表展示 E01~E08",
        "F绑定 F01~F14",
        "G解绑 G01~G06",
    ]:
        add_topic(m1["children"]["attached"], sec)

    m2 = add_topic(root_attached, "2 二维码管理（46条详细）92228")
    for sec in [
        "A菜单权限 QR-A01~A04",
        "B同步list_contact_way B01~B11",
        "C列表筛选 C01~C11",
        "D预览下载 D01~D06",
        "E state/remark映射 E01~E05",
        "F联动好友池 F01~F05",
        "G异常安全 G01~G04",
    ]:
        add_topic(m2["children"]["attached"], sec)

    m3 = add_topic(root_attached, "3 扫码/回调→自动进好友池")
    add_topic(
        m3["children"]["attached"],
        "主流程",
        ["AUTO-001 进池", "AUTO-002/003 负责人", "AUTO-013/014 去重"],
    )
    add_topic(
        m3["children"]["attached"],
        "字段映射（列表）",
        [
            "AUTO-004 客户名兜底",
            "AUTO-005 来源add_way+state",
            "AUTO-006~010 手机/企业/属性/标签组",
        ],
    )

    m4 = add_topic(root_attached, f"4 企微好友池/待入库池（{len(CASES_FP)}条详细）")
    for sec in [
        "A菜单权限 FP-A01~A04",
        "B筛选 FP-B01~B14",
        "C批量操作 FP-C01~C18",
        "D行内操作 FP-D01~D06",
        "E列表映射 FP-E01~E18",
        "F手动同步 FP-F01~F06",
    ]:
        add_topic(m4["children"]["attached"], sec)

    m5 = add_topic(root_attached, f"5 转线索映射（{len(CASES_MAP)}条）")
    add_topic(
        m5["children"]["attached"],
        "MAP-A01~A16",
        ["职务/企业兜底备注", "多手机号", "转介绍/展会", "批量一致"],
    )

    m6 = add_topic(root_attached, "6 端到端 E2E-001")
    add_topic(m6["children"]["attached"], "建议前置全部P0通过后执行")

    sheet = {
        "id": tid(),
        "revisionId": tid(),
        "class": "sheet",
        "title": "CRM企微对接",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "CRM-企业微信对接 测试脑图（功能为主｜含阻塞与首轮）",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": root_attached},
        },
    }
    content = [sheet]
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    metadata = {
        "creator": {"name": "CRMWeWorkGenerator", "version": "2.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet["id"],
    }
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        zf.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        zf.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))


def _style_header(ws, headers, fill_color="4472C4"):
    hf = PatternFill("solid", fgColor=fill_color)
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(wrap_text=True, vertical="center")


def _write_cases_sheet(ws, cases, headers):
    _style_header(ws, headers)
    block_fill = PatternFill("solid", fgColor="FFC7CE")
    p0_fill = PatternFill("solid", fgColor="FFEB9C")
    for i, row in enumerate(cases, 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if len(row) > 4 and row[4] == "是":
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = block_fill
        elif len(row) > 3 and row[3] == "P0":
            for col in range(1, len(headers) + 1):
                if not ws.cell(row=i, column=col).fill or ws.cell(row=i, column=col).fill.fgColor.rgb == "00000000":
                    ws.cell(row=i, column=col).fill = p0_fill
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col not in (6, 7, 8) else 42
    ws.freeze_panes = "A2"


def api_link(note, steps, title, api_hint):
    blob = f"{note}\n{steps}\n{title}\n{api_hint}"
    links = []

    def add(u):
        if u not in links:
            links.append(u)

    if "90196" in blob:
        add("https://developer.work.weixin.qq.com/document/path/90196")
    if "92130" in blob or "回调" in title:
        add("https://developer.work.weixin.qq.com/document/path/92130")
    if "92114" in blob or "externalcontact" in blob.lower():
        add("https://developer.work.weixin.qq.com/document/path/92114")
    if "92228" in blob or "联系我" in title or "contact_way" in blob.lower():
        add("https://developer.work.weixin.qq.com/document/path/92228")
    return "\n".join(links)


def build_excel():
    wb = Workbook()
    headers = [
        "模块",
        "用例编号",
        "用例标题",
        "优先级",
        "是否阻塞",
        "首轮必测",
        "前置条件",
        "测试步骤",
        "预期结果",
        "备注",
        "关联接口/文档",
    ]

    # --- 需求追溯 ---
    ws0 = wb.active
    ws0.title = "需求追溯"
    ws0["A1"] = "需求背景"
    ws0["B1"] = (
        "CRM对接企业微信：①销售人员账号同步+绑定CRM；②管理销售人员二维码；"
        "③扫码线索自动同步进CRM（好友池+建线索）。"
    )
    ws0["A2"] = "菜单"
    ws0["B2"] = "设置-用户企微绑定；设置-二维码管理；设置-企微好友池（均需独立菜单权限）"
    ws0["A3"] = "关键接口"
    ws0["B3"] = (
        "员工同步90196；联系我列表92228；外部联系人92114；回调92130"
    )
    ws0["A4"] = "首轮测试建议"
    ws0["B4"] = (
        "按Sheet「首轮冒烟」顺序执行：先0前置→1绑定→2二维码→3自动进池→4好友池批量→5映射→6 E2E。"
        "阻塞项失败则暂停后续模块并提缺陷。粉色行=阻塞场景。"
    )
    ws0["A5"] = "用例统计"
    total = len(CASES)
    block = sum(1 for c in CASES if c[4] == "是")
    first = sum(1 for c in CASES if c[5] == "是")
    p0 = sum(1 for c in CASES if c[3] == "P0")
    ws0["B5"] = (
        f"合计{total}条（绑定{len(CASES_UB)}+二维码{len(CASES_QR)}+好友池{len(CASES_FP)}"
        f"+映射{len(CASES_MAP)}+自动{len(CASES_AUTO)}+前置{len(CASES_PRE)}+E2E{len(CASES_E2E)}）；"
        f"阻塞{block}条；首轮必测{first}条；P0共{p0}条"
    )
    ws0["A6"] = "分模块Sheet"
    ws0["B6"] = (
        "测试用例=全量；首轮冒烟；1-用户企微绑定；2-二维码管理；"
        "4-企微好友池；5-转线索映射"
    )
    ws0.column_dimensions["A"].width = 14
    ws0.column_dimensions["B"].width = 78
    for r in range(1, 6):
        ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- 全量用例 ---
    ws_all = wb.create_sheet("测试用例")
    rows_all = []
    for c in CASES:
        module, cid, title, pri, blocking, first, pre, steps, expect, note, api_hint = c
        doc = api_link(note, steps, title, api_hint)
        rows_all.append(
            (module, cid, title, pri, blocking, first, pre, steps, expect, note, doc or api_hint)
        )
    _write_cases_sheet(ws_all, rows_all, headers)

    def rows_for_module(prefix):
        out = []
        for c in CASES:
            if c[0].startswith(prefix):
                module, cid, title, pri, blocking, first, pre, steps, expect, note, api_hint = c
                doc = api_link(note, steps, title, api_hint)
                out.append(
                    (module, cid, title, pri, blocking, first, pre, steps, expect, note, doc or api_hint)
                )
        return out

    ws_ub = wb.create_sheet("1-用户企微绑定")
    _write_cases_sheet(ws_ub, rows_for_module("1-用户企微绑定"), headers)

    ws_qr = wb.create_sheet("2-二维码管理")
    _write_cases_sheet(ws_qr, rows_for_module("2-二维码管理"), headers)

    ws_fp = wb.create_sheet("4-企微好友池")
    _write_cases_sheet(ws_fp, rows_for_module("4-企微好友池"), headers)

    ws_map = wb.create_sheet("5-转线索映射")
    _write_cases_sheet(ws_map, rows_for_module("5-好友池转线索映射"), headers)

    # --- 首轮冒烟 ---
    ws_smoke = wb.create_sheet("首轮冒烟")
    smoke = [r for r in rows_all if r[5] == "是"]
    _write_cases_sheet(ws_smoke, smoke, headers)

    # --- 阻塞清单 ---
    ws_block = wb.create_sheet("阻塞场景清单")
    block_headers = ["模块", "用例编号", "用例标题", "优先级", "阻塞说明", "失败影响"]
    _style_header(ws_block, block_headers, fill_color="C00000")
    block_impact = {
        "PRE": "后续同步、回调、好友池、建线索均无法进行",
        "UB": "员工无法绑定→负责人映射错误→线索分配异常",
        "QR": "二维码无法同步/识别→扫码来源与展会名错误",
        "AUTO": "扫码数据无法入库→主业务链路中断",
        "FP": "好友无法转线索→销售无法跟进",
        "MAP": "线索字段错误→影响CRM后续流程",
        "E2E": "全链路验收失败",
    }
    for i, c in enumerate([x for x in CASES if x[4] == "是"], 2):
        prefix = c[1].split("-")[0]
        if c[0].startswith("1-"):
            prefix = "UB"
        elif c[0].startswith("2-"):
            prefix = "QR"
        elif c[0].startswith("4-"):
            prefix = "FP"
        elif c[0].startswith("5-"):
            prefix = "MAP"
        impact = block_impact.get(prefix, "阻塞关联模块测试")
        vals = [c[0], c[1], c[2], c[3], c[9] or "见用例步骤", impact]
        for col, v in enumerate(vals, 1):
            cell = ws_block.cell(row=i, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 7):
        ws_block.column_dimensions[get_column_letter(col)].width = 22 if col == 5 else 18
    ws_block.freeze_panes = "A2"

    # --- 优先级说明 ---
    ws_pri = wb.create_sheet("优先级说明")
    ws_pri.append(["优先级", "定义", "首轮策略"])
    ws_pri.append(["P0", "核心功能/主流程/数据正确性", "必须首轮全部执行"])
    ws_pri.append(["P1", "重要分支、权限、筛选、补偿逻辑", "首轮时间允许则执行"])
    ws_pri.append(["P2", "边界、UI细节、非关键异常", "回归轮次执行"])
    ws_pri.append(["阻塞=是", "该项失败会导致后续模块测试无意义或结果不可信", "失败即停，修复后从该模块重测"])

    wb.save(XLSX_PATH)


def main():
    build_excel()
    build_xmind()
    import shutil

    print(
        f"Generated {len(CASES)} cases "
        f"(UB={len(CASES_UB)}, QR={len(CASES_QR)}, FP={len(CASES_FP)}, MAP={len(CASES_MAP)}) ->"
    )
    print(" ", XLSX_PATH)
    print(" ", XMIND_PATH)
    for src, dst in ((XLSX_PATH, ROOT_XLSX), (XMIND_PATH, ROOT_XMIND)):
        try:
            shutil.copy2(src, dst)
            print(" ", dst)
        except OSError as e:
            print(f"  (skip copy to {dst.name}: {e})")


if __name__ == "__main__":
    main()