# -*- coding: utf-8 -*-
"""CRM 客户/线索分配后企微提醒：生成 Excel 与 XMind 测试用例。"""
import json
import uuid
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "testcases"
OUT_DIR.mkdir(parents=True, exist_ok=True)
SHORT_NAME = "CRM客户线索分配企微提醒"
DOC_TITLE = "CRM-客户/线索分配后，新增企微提醒"
DOC_URL = "https://tvd6quau8vr.feishu.cn/wiki/PYadwUf6Iig6KsktT0sc4KvunTc"
COMMENT_IMAGE = (
    r"C:\Users\caitiantian\.cursor\projects\e-cursor-Pyautotest\assets"
    r"\c__Users_caitiantian_AppData_Roaming_Cursor_User_workspaceStorage_empty-window_images"
    r"_image-309dd8fe-c4cb-402d-8107-7efdecbee056.png"
)
XLSX_PATH = OUT_DIR / f"{SHORT_NAME}_TestCases.xlsx"
XMIND_PATH = OUT_DIR / f"{SHORT_NAME}_TestMindMap.xmind"

HEADERS = [
    "用例ID", "模块", "优先级", "场景", "前置条件", "测试步骤", "预期结果",
    "实际结果", "备注", "用例状态", "是否阻塞", "首轮必测",
]


def case(cid, module, priority, scene, precondition, steps, expected,
         note="", blocking="否", smoke="否"):
    return {
        "用例ID": cid, "模块": module, "优先级": priority, "场景": scene,
        "前置条件": precondition, "测试步骤": steps, "预期结果": expected,
        "实际结果": "", "备注": note, "用例状态": "",
        "是否阻塞": blocking, "首轮必测": smoke,
    }


CASES = []


CUSTOMER_PATHS = [
    ("中文商城注册", "通过中文商城注册产生新客户并分配给用户A"),
    ("英文商城注册", "通过英文商城注册产生新客户并分配给用户A"),
    ("页面新建国内客户", "在CRM页面新建国内客户并直接指定用户A"),
    ("页面新建国外客户", "在CRM页面新建国外客户并直接指定用户A"),
    ("导入国内客户", "导入国内客户并在导入时指定用户A"),
    ("导入国外客户", "导入国外客户并在导入时指定用户A"),
    ("个人客户转换企业客户", "将个人客户转换为企业客户并分配给用户A"),
    ("线索转换国内客户", "将线索转换为国内客户并指定用户A"),
    ("线索转换国外客户", "将线索转换为国外客户并指定用户A"),
    ("一般分配", "对单条客户执行一般分配到用户A"),
    ("公海分配", "从客户公海将单条客户分配给用户A"),
    ("批量分配", "选择多条客户批量分配给用户A"),
    ("单条认领", "用户A认领一条客户"),
    ("批量认领", "用户A批量认领多条客户"),
]
for index, (name, action) in enumerate(CUSTOMER_PATHS, 1):
    batch = "批量" in name
    CASES.append(case(
        f"CA-{index:03d}", "1-客户新增提醒", "P0",
        f"{name}后向被分配人发送新增客户企微提醒",
        "商城用户A已绑定企微员工；A具备客户查看权限；准备含客户名称、首个联系人和产品需求的数据",
        f"1. {action}\n2. 检查用户A企微消息\n3. 点击消息链接并核对客户归属",
        (
            "用户A实时收到新增客户提醒；标题为“您名下有一条新增客户！”；"
            "详情中的客户名称、首个联系人和产品需求与本次客户一致；链接进入对应客户详情；"
            + ("批量数据均有对应提醒且无遗漏、重复或错配" if batch else "仅本次目标客户产生一条提醒")
        ),
        "全文评论补充场景" + ("；待产品确认批量场景按单条发送还是聚合发送" if batch else ""),
        "是", "是",
    ))

CASES.extend([
    case("CA-015", "1-客户新增提醒", "P0", "客户改分配人时仅新被分配人收到提醒",
         "客户当前归属用户A；用户B已绑定企微", "1. 将客户从A手动分配给B\n2. 检查A、B企微消息",
         "B实时收到该客户新增提醒；A不收到本次新增提醒；CRM客户跟进人更新为B",
         "分配对象应以操作后的实际归属为准", "是", "是"),
    case("CA-016", "1-客户新增提醒", "P1", "客户分配接口重复回调保持提醒幂等",
         "可模拟同一客户同一分配事件重复提交或重试", "1. 对相同事件重复调用分配接口\n2. 检查归属、消息和发送记录",
         "客户归属正确；同一业务事件不产生重复企微轰炸；发送记录可关联同一事件",
         "分配、一般分配、公海分配、批量分配共用一个接口的补充口径"),
])


LEAD_PATHS = [
    ("页面新建", "在CRM页面新建线索并直接指定用户A"),
    ("导入", "导入线索并在导入时指定用户A"),
    ("单条分配", "将一条线索分配给用户A"),
    ("批量分配", "将多条线索批量分配给用户A"),
    ("单条认领", "用户A认领一条线索"),
    ("批量认领", "用户A批量认领多条线索"),
    ("企微好友池新建", "从企微好友池新建线索并归属用户A"),
    ("中文商城展会登记同步", "中文商城展会登记后同步创建线索并归属用户A"),
    ("英文商城展会登记同步", "英文商城展会登记后同步创建线索并归属用户A"),
    ("百度渠道留资同步", "百度渠道留资同步创建线索并归属用户A"),
    ("腾讯渠道留资同步", "腾讯渠道留资同步创建线索并归属用户A"),
    ("抖音渠道留资同步", "抖音渠道留资同步创建线索并归属用户A"),
    ("其他渠道留资同步", "通过已接入的其他渠道留资同步创建线索并归属用户A"),
]
for index, (name, action) in enumerate(LEAD_PATHS, 1):
    batch = "批量" in name
    CASES.append(case(
        f"LA-{index:03d}", "2-线索新增提醒", "P0",
        f"{name}后向被分配人发送新增线索企微提醒",
        "商城用户A已绑定企微员工；A具备线索查看权限；准备含线索名称、来源和产品需求的数据",
        f"1. {action}\n2. 检查用户A企微消息\n3. 点击消息链接并核对线索归属",
        (
            "用户A实时收到新增线索提醒；标题为“您名下有一条新增线索！”；"
            "详情中的线索名称、线索来源和产品需求与本次线索一致；链接进入对应线索详情；"
            + ("批量数据均有对应提醒且无遗漏、重复或错配" if batch else "仅本次目标线索产生一条提醒")
        ),
        "全文评论补充场景" + ("；待产品确认批量场景按单条发送还是聚合发送" if batch else ""),
        "是", "是",
    ))

CASES.extend([
    case("LA-014", "2-线索新增提醒", "P0", "线索系统自动分配后实时发送提醒",
         "已配置可触发的线索自动流转规则；用户A已绑定企微", "1. 创建满足自动分配规则的线索\n2. 等待规则执行\n3. 检查归属和企微消息",
         "线索自动归属A后，A实时收到一条新增线索提醒；消息字段和链接对应目标线索",
         "正文明确包含系统自动分配", "是", "是"),
    case("LA-015", "2-线索新增提醒", "P1", "线索分配接口重复回调保持提醒幂等",
         "可模拟同一线索分配事件重复提交", "1. 重复提交同一分配事件\n2. 检查消息和发送日志",
         "同一业务事件不产生重复提醒；线索归属和发送记录一致",
         "分配/批量分配共用一个接口的补充口径"),
])


CASES.extend([
    case("CR-001", "3-客户回收提醒", "P0", "客户触发回收规则后在当天9点发送提醒",
         "用户A已绑定企微；客户满足即将回收规则；系统时间任务可执行",
         "1. 使客户触发回收规则\n2. 在触发当天9:00检查A企微\n3. 检查发送记录",
         "用户A在当天9:00收到一条客户回收提醒；发送时间和接收人正确",
         "正文明确：触发回收消息当天早上9:00", "是", "是"),
    case("CR-002", "3-客户回收提醒", "P0", "客户实际未回收仍发送已触发的回收提醒",
         "客户已触发回收规则；发送前通过跟进等操作避免实际回收", "1. 触发回收规则\n2. 使客户最终不被回收\n3. 当天9:00检查消息",
         "即使客户最终未被回收，仍按已触发事件向A发送提醒",
         "正文明确“实际不回收也发消息”", "是", "是"),
    case("CR-003", "3-客户回收提醒", "P0", "客户回收提醒标题与详情字段正确",
         "客户含名称、首个联系人和产品需求；已触发回收", "1. 等待9:00收到消息\n2. 核对标题和详情",
         "标题为“您名下有一条客户即将被回收公海！”；详情包含正确客户名称、联系人名称、产品需求及“请及时跟进！”",
         "联系人取值口径正文仅在新增客户处明确为首个联系人，回收场景待确认是否同样取首个联系人", "是", "是"),
    case("CR-004", "3-客户回收提醒", "P1", "同一客户重复命中回收任务不重复提醒",
         "可模拟同一天多次扫描同一客户", "1. 多次执行回收规则扫描\n2. 9:00后检查消息和日志",
         "同一客户同一回收触发周期不重复发送提醒；日志可追踪扫描和发送结果"),
    case("CR-005", "3-客户回收提醒", "P1", "多个客户同时触发回收提醒无遗漏错配",
         "A名下多个客户同日触发回收规则", "1. 执行回收规则\n2. 9:00检查全部消息\n3. 逐条核对字段和链接",
         "每个应提醒客户均有对应消息；客户字段和详情链接不串数据；无遗漏和重复"),
    case("CR-006", "3-客户回收提醒", "P1", "9点任务延迟或失败后结果可追踪",
         "可模拟定时任务延迟、企微接口暂时失败", "1. 触发回收\n2. 在9:00制造发送故障\n3. 恢复服务并检查日志",
         "失败被后台记录；不误记成功；恢复后的处理符合重试策略且不重复轰炸",
         "待技术确认：重试次数、间隔和补发时限"),
])

CASES.extend([
    case("LR-001", "4-线索回收提醒", "P0", "线索触发回收规则后在当天9点发送提醒",
         "用户A已绑定企微；线索满足即将回收规则", "1. 使线索触发回收规则\n2. 当天9:00检查A企微",
         "A在当天9:00收到一条线索回收提醒；发送时间和接收人正确",
         "正文及全文评论均明确触发回收", "是", "是"),
    case("LR-002", "4-线索回收提醒", "P0", "线索实际未回收仍发送已触发的回收提醒",
         "线索已触发回收规则；发送前使其最终不回收", "1. 触发回收规则\n2. 使线索不实际回收\n3. 9:00检查消息",
         "即使最终未回收，仍按已触发事件向A发送提醒",
         "正文明确“实际不回收也发消息”", "是", "是"),
    case("LR-003", "4-线索回收提醒", "P0", "线索回收提醒标题与详情字段正确",
         "线索含名称、来源和产品需求；已触发回收", "1. 收到消息\n2. 核对标题和详情",
         "标题为“您名下有一条线索即将被回收公海！”；详情包含正确线索名称、线索来源、产品需求及“请及时跟进！”",
         "正文消息格式", "是", "是"),
    case("LR-004", "4-线索回收提醒", "P1", "同一线索重复命中回收任务不重复提醒",
         "可模拟同一天多次扫描同一线索", "1. 多次执行回收扫描\n2. 检查消息和日志",
         "同一线索同一回收触发周期不重复发送提醒；日志记录一致"),
    case("LR-005", "4-线索回收提醒", "P1", "多条线索同时触发回收无遗漏错配",
         "A名下多条线索同日触发回收", "1. 执行规则\n2. 9:00检查消息\n3. 核对链接",
         "所有应提醒线索均有对应消息；字段和链接不串数据；无遗漏和重复"),
    case("LR-006", "4-线索回收提醒", "P1", "回收提醒按正确时区在9点执行",
         "测试环境可配置或观察服务器、用户时区", "1. 触发回收\n2. 对照系统时区观察发送时间",
         "在产品定义时区的当天9:00发送，不因服务器时区造成提前或延后一天",
         "待产品确认：早上9:00采用北京时间、组织时区还是用户时区"),
])


CASES.extend([
    case("WB-001", "5-企微绑定与发送", "P0", "已绑定商城用户可定位唯一企微员工",
         "商城用户A在CRM建立有效企微绑定", "1. 触发客户或线索分配给A\n2. 检查发送目标和绑定记录",
         "系统根据商城用户A准确找到其绑定的企微员工；消息只发送给该员工",
         "正文“根据商城用户A在CRM中建立的企微绑定关系”", "是", "是"),
    case("WB-002", "5-企微绑定与发送", "P0", "未绑定企微员工时记录失败且不发送",
         "商城用户A未绑定企微员工", "1. 将客户/线索分配给A\n2. 检查企微消息和后台日志",
         "不向任何企微员工发送消息；后台记录发送失败及未绑定原因；CRM分配业务本身成功",
         "正文明确要求", "是", "是"),
    case("WB-003", "5-企微绑定与发送", "P1", "无效或已离职企微绑定发送失败可追踪",
         "A绑定的企微员工已停用、离职或企微接口返回无效用户", "1. 触发提醒\n2. 检查接口结果和日志",
         "消息不错误发送给其他员工；后台记录真实失败原因；客户/线索归属不回滚"),
    case("WB-004", "5-企微绑定与发送", "P1", "更换企微绑定后消息发送给新员工",
         "商城用户A的绑定由企微员工X变更为Y", "1. 完成绑定变更\n2. 触发新分配\n3. 检查X和Y消息",
         "新提醒仅发送给当前有效绑定Y；X不再收到；历史发送记录保持原接收人"),
    case("WB-005", "5-企微绑定与发送", "P1", "不同商城用户的企微消息严格隔离",
         "用户A、B分别绑定不同企微员工", "1. 分别将不同客户和线索分配给A、B\n2. 核对两人消息",
         "A、B仅收到各自名下数据提醒；消息字段和链接不串用户或跨权限泄露"),
    case("WB-006", "5-企微绑定与发送", "P1", "企微接口超时不阻塞CRM分配",
         "可模拟企微发送接口超时", "1. 执行客户/线索分配\n2. 等待接口超时\n3. 检查CRM结果与日志",
         "CRM分配结果正常落库；提醒失败或处理中状态可追踪；页面不因企微超时长时间卡死",
         "实时发送超时阈值和异步机制待技术确认"),
    case("WB-007", "5-企微绑定与发送", "P1", "本期不发送短信或邮件提醒",
         "用户A同时配置手机号、邮箱和企微", "1. 触发四类提醒\n2. 检查短信、邮件及企微",
         "仅按需求发送企微消息；不新增短信或邮件提醒",
         "正文已确认本期不需要短信/邮件"),
])


CASES.extend([
    case("LK-001", "6-链接权限与跳转", "P0", "有客户查看权限时跳转指定客户详情",
         "A有客户菜单及目标客户查看权限；已收到客户提醒", "1. 在企微点击消息链接\n2. 完成必要登录\n3. 检查落地页",
         "进入消息对应的客户详情页；客户ID和展示数据正确",
         "企微内置浏览器登录后直达能力需开发评估", "是", "是"),
    case("LK-002", "6-链接权限与跳转", "P0", "无目标客户查看权限时降级客户菜单页",
         "A有客户菜单权限但无目标客户查看权限", "1. 点击客户提醒链接\n2. 检查落地页及数据",
         "不展示目标客户详情或敏感数据；跳转客户菜单页",
         "正文明确权限降级", "是", "是"),
    case("LK-003", "6-链接权限与跳转", "P0", "无客户菜单权限时返回首页并提示",
         "A无客户菜单权限", "1. 点击客户提醒链接\n2. 检查落地页和提示",
         "跳转CRM首页并提示“您没有客户菜单权限，请联系系统管理员开启”或最终确认文案；不泄露客户信息",
         "原文末尾标点为“、”，建议按正常中文标点确认", "是", "是"),
    case("LK-004", "6-链接权限与跳转", "P0", "有线索查看权限时跳转指定线索详情",
         "A有线索菜单及目标线索查看权限；已收到线索提醒", "1. 在企微点击链接\n2. 完成必要登录\n3. 检查落地页",
         "进入消息对应的线索详情页；线索ID和数据正确",
         "企微内置浏览器登录后直达能力需开发评估", "是", "是"),
    case("LK-005", "6-链接权限与跳转", "P0", "无目标线索查看权限时降级线索菜单页",
         "A有线索菜单权限但无目标线索查看权限", "1. 点击线索提醒链接\n2. 检查落地页",
         "不展示目标线索详情或敏感数据；跳转线索菜单页",
         "正文明确权限降级", "是", "是"),
    case("LK-006", "6-链接权限与跳转", "P0", "无线索菜单权限时返回首页并提示",
         "A无线索菜单权限", "1. 点击线索提醒链接\n2. 检查落地页和提示",
         "跳转CRM首页并提示“您没有线索菜单权限，请联系系统管理员开启。”或最终确认文案；不泄露线索信息",
         "正文明确权限降级", "是", "是"),
    case("LK-007", "6-链接权限与跳转", "P1", "企微内未登录时登录后恢复目标跳转",
         "企微内置浏览器无CRM登录态；A有目标详情权限", "1. 点击提醒链接\n2. 按提示登录\n3. 检查登录后页面",
         "未登录时不直接暴露详情；登录成功后进入原目标详情或按最终评估方案落地",
         "待开发评估：登录后直接跳转目标详情的工作量与最终方案"),
    case("LK-008", "6-链接权限与跳转", "P1", "消息链接ID被篡改时执行服务端权限校验",
         "A仅有原目标数据权限", "1. 修改链接中的客户或线索标识\n2. 在企微打开篡改链接",
         "服务端重新校验登录和数据权限；不能查看无权数据；按菜单/首页规则降级或提示"),
    case("LK-009", "6-链接权限与跳转", "P1", "数据已删除或进入公海时链接安全降级",
         "提醒发送后目标数据被删除、作废或回收公海", "1. 点击历史消息链接\n2. 检查页面响应",
         "不出现白屏或泄露历史数据；展示不存在/无权限提示，或按当前权限降级到对应菜单页",
         "具体提示文案待产品确认"),
    case("LK-010", "6-链接权限与跳转", "P1", "PC浏览器打开企微消息链接可正确处理",
         "已复制一条有效消息链接；A有权限", "1. 在PC浏览器打开链接\n2. 登录CRM\n3. 检查跳转",
         "链接可识别目标环境并完成权限校验；登录后进入目标详情或按最终方案降级；无错误域名"),
])


CASES.extend([
    case("MC-001", "7-消息内容与数据", "P0", "新增客户默认取首个联系人",
         "客户有多个联系人且顺序明确", "1. 分配客户给A\n2. 核对消息联系人\n3. 与客户联系人顺序比较",
         "消息联系人名称取客户首个联系人；不随机取其他联系人",
         "正文明确默认取客户首个联系人", "是", "是"),
    case("MC-002", "7-消息内容与数据", "P1", "客户无联系人时消息可发送",
         "客户没有联系人但可合法分配", "1. 分配客户给A\n2. 查看消息详情",
         "提醒仍可发送且客户名称、产品需求正确；联系人空值按最终占位规则展示，不出现null、undefined或模板报错",
         "待产品确认：无联系人时隐藏字段、留空或显示“--”"),
    case("MC-003", "7-消息内容与数据", "P1", "客户首个联系人变更后取触发时最新数据",
         "客户存在多个联系人；可调整首个联系人", "1. 变更首个联系人\n2. 再分配客户\n3. 核对消息",
         "消息取提醒触发时的首个联系人；不使用已失效缓存"),
    case("MC-004", "7-消息内容与数据", "P1", "产品需求为空或多值时格式可读",
         "分别准备产品需求为空、单值、多值的客户和线索", "1. 逐一触发提醒\n2. 核对产品需求字段",
         "单值准确展示；多值完整且分隔清晰；空值不展示null/undefined或破坏消息结构",
         "待产品确认：空值占位及多值分隔格式"),
    case("MC-005", "7-消息内容与数据", "P1", "客户线索名称含特殊字符时安全展示",
         "名称包含中英文、emoji、HTML字符、逗号和分号", "1. 触发客户及线索提醒\n2. 查看消息并点击链接",
         "合法字符正确转义和展示；脚本不执行；字段分隔仍可辨识；链接指向正确记录"),
    case("MC-006", "7-消息内容与数据", "P1", "超长名称和产品需求不导致发送失败",
         "准备达到字段长度上限的数据", "1. 触发提醒\n2. 检查企微消息和发送日志",
         "消息按企微限制完整展示或安全截断；核心标识和链接可用；后台不因超长内容异常",
         "待技术确认：企微消息长度限制和截断规则"),
    case("MC-007", "7-消息内容与数据", "P1", "线索来源与实际同步渠道一致",
         "准备页面、导入、企微好友池、展会及百度/腾讯/抖音等线索", "1. 各触发一次提醒\n2. 对比CRM线索来源和消息",
         "消息中的线索来源取CRM最终落库值；不同来源不混淆、不使用渠道内部错误编码"),
    case("MC-008", "7-消息内容与数据", "P1", "四类消息标题不会混用",
         "分别触发新增客户、新增线索、客户回收、线索回收", "1. 收集四类消息\n2. 核对标题和详情字段",
         "四类标题分别准确；客户消息不出现线索字段，线索消息不出现联系人字段；新增与回收标题不混用"),
])


CASES.extend([
    case("EX-001", "8-异常幂等与性能", "P0", "CRM分配成功但企微发送失败时业务数据一致",
         "可模拟企微接口返回失败", "1. 分配客户或线索\n2. 检查CRM归属和日志",
         "CRM分配成功且归属正确；企微发送记录为失败并含原因；不因通知失败回滚分配",
         "通知为分配后的附加动作", "是", "是"),
    case("EX-002", "8-异常幂等与性能", "P1", "CRM分配事务失败时不发送成功提醒",
         "可模拟分配接口事务失败", "1. 发起分配并制造落库失败\n2. 检查归属、消息和日志",
         "客户/线索未归属A时，不向A发送“已到名下”提醒；避免消息与业务事实不一致"),
    case("EX-003", "8-异常幂等与性能", "P1", "批量分配部分成功时仅成功项发送提醒",
         "批量数据中包含可分配和不可分配记录", "1. 执行批量分配\n2. 核对逐条结果和消息",
         "仅实际分配成功且到A名下的数据产生提醒；失败项不发送成功提醒；数量与接口结果一致"),
    case("EX-004", "8-异常幂等与性能", "P1", "批量认领部分成功时仅成功项发送提醒",
         "批量认领中包含已被他人认领或无权限记录", "1. A执行批量认领\n2. 核对结果和企微消息",
         "仅A实际认领成功的数据产生提醒；失败项不发送；无跨用户错发"),
    case("EX-005", "8-异常幂等与性能", "P1", "高并发分配提醒无丢失和串数据",
         "可并发创建、导入或批量分配多条客户线索", "1. 并发触发大量分配\n2. 对账业务事件、发送日志和企微消息",
         "已成功事件与发送记录可一一对应；无串数据、重复接收或错误接收人；CRM核心操作性能无明显劣化",
         "实时发送性能指标待技术确认"),
    case("EX-006", "8-异常幂等与性能", "P1", "消息发送日志支持定位失败原因",
         "准备成功、未绑定、接口失败、无效员工等事件", "1. 分别触发\n2. 查询后台发送结果",
         "每次事件可查看业务对象、接收商城用户、企微员工、触发类型、时间、结果及失败原因；敏感信息不明文泄露",
         "正文只明确未绑定需记录失败，其他日志字段待技术确认"),
    case("EX-007", "8-异常幂等与性能", "P1", "跨环境消息链接不会跳错环境",
         "分别在测试/UAT等环境触发提醒", "1. 点击各环境消息链接\n2. 检查域名和目标数据",
         "链接使用当前环境CRM地址；不会从测试消息跳生产或访问其他环境同ID数据"),
    case("EX-008", "8-异常幂等与性能", "P1", "分配给未启用商城用户不错误发送",
         "目标用户已禁用、离职或商城状态无效", "1. 尝试分配或由规则分配给该用户\n2. 检查业务结果和提醒",
         "系统按现有分配规则阻止或处理无效用户；不会向错误企微员工发送；结果和原因可追踪",
         "待产品确认：已禁用商城用户是否允许成为分配人"),
])


CASES.extend([
    case("E2E-001", "9-端到端", "P0", "新建客户到企微详情页完整闭环",
         "A已绑定企微并有客户权限", "1. 页面新建客户并指定A\n2. A接收消息\n3. 点击进入详情\n4. 核对字段和归属",
         "创建、分配、实时提醒、权限校验和详情跳转全链路成功；数据一致且仅一条消息",
         "核心验收", "是", "是"),
    case("E2E-002", "9-端到端", "P0", "渠道线索同步到企微详情页完整闭环",
         "A已绑定企微并有线索权限；渠道同步可用", "1. 从任一渠道提交留资\n2. 同步创建并分配线索\n3. A接收消息并进入详情",
         "渠道留资、CRM线索、跟进人、企微消息字段和详情数据一致；仅发送给A",
         "核心验收", "是", "是"),
    case("E2E-003", "9-端到端", "P0", "客户线索同日回收提醒完整闭环",
         "A名下客户和线索均触发回收规则", "1. 触发两类回收\n2. 当天9:00收取消息\n3. 点击链接",
         "两类消息按时发送且标题字段不混用；链接按各自权限进入详情或正确降级",
         "核心验收", "是", "是"),
    case("E2E-004", "9-端到端", "P0", "有权限无权限未绑定三角色矩阵闭环",
         "准备有详情权限A、仅菜单权限B、无菜单且未绑定企微C", "1. 分别向三人分配数据\n2. 检查消息和跳转\n3. 检查失败日志",
         "A收到消息并进入详情；B收到消息但降级菜单；C未发送且后台记录未绑定失败；无越权泄露",
         "权限与绑定核心验收", "是", "是"),
])


MODULES = [
    "1-客户新增提醒", "2-线索新增提醒", "3-客户回收提醒", "4-线索回收提醒",
    "5-企微绑定与发送", "6-链接权限与跳转", "7-消息内容与数据",
    "8-异常幂等与性能", "9-端到端",
]

COVERAGE = [
    ("客户创建/导入/转换/商城注册", "已覆盖", "CA-001~009"),
    ("客户一般分配/公海分配/批量分配", "已覆盖", "CA-010~012, CA-015~016"),
    ("客户认领/批量认领", "已覆盖", "CA-013~014"),
    ("线索页面新建/导入/分配/认领", "已覆盖", "LA-001~006, LA-014~015"),
    ("企微好友池、中文/英文展会登记", "已覆盖", "LA-007~009"),
    ("百度、腾讯、抖音及其他渠道留资同步", "已覆盖", "LA-010~013"),
    ("客户/线索触发回收且实际不回收仍提醒", "已覆盖", "CR-001~006, LR-001~006"),
    ("企微绑定查找、未绑定失败记录", "已覆盖", "WB-001~007"),
    ("详情、菜单、首页三级权限降级", "部分覆盖", "LK-001~010；企微内登录后直达方案待开发确认"),
    ("四类消息标题和字段", "部分覆盖", "MC-001~008；空值、多值、超长格式待确认"),
    ("批量消息形式、9点时区和重试策略", "待确认", "CA-012, CA-014, LA-004, LA-006, LR-006, CR-006"),
    ("短信/邮件提醒", "不在本期", "WB-007"),
]

BLOCK_IMPACT = {
    "CA": "客户新增提醒主链路不可验收",
    "LA": "线索新增提醒主链路不可验收",
    "CR": "客户回收提醒不可验收",
    "LR": "线索回收提醒不可验收",
    "WB": "企微接收人定位或发送不可用",
    "LK": "链接权限控制或跳转不可用",
    "MC": "消息核心字段错误",
    "EX": "通知失败影响CRM核心数据一致性",
    "E2E": "跨模块发布验收失败",
}


def style_header(ws, headers, color="4472C4"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_case_sheet(ws, items):
    style_header(ws, HEADERS)
    red = PatternFill("solid", fgColor="FFC7CE")
    blue = PatternFill("solid", fgColor="DDEBF7")
    for row_no, item in enumerate(items, 2):
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row_no, col, item[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        fill = red if item["是否阻塞"] == "是" else blue if item["优先级"] == "P0" else None
        if fill:
            for col in range(1, 13):
                ws.cell(row_no, col).fill = fill
    for col, width in enumerate([14, 23, 9, 38, 34, 48, 48, 16, 38, 12, 10, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(2, len(items) + 1)}"
    status_col = get_column_letter(HEADERS.index("用例状态") + 1)
    validation = DataValidation(type="list", formula1='"PASS,FAIL,BLOCK,N/A"', allow_blank=True)
    validation.error = "请选择 PASS、FAIL、BLOCK 或 N/A"
    ws.add_data_validation(validation)
    validation.add(f"{status_col}2:{status_col}{max(500, len(items) + 1)}")


def build_excel():
    wb = Workbook()
    meta = wb.active
    meta.title = "需求追溯"
    stats = Counter(item["优先级"] for item in CASES)
    blocking = sum(item["是否阻塞"] == "是" for item in CASES)
    smoke = sum(item["首轮必测"] == "是" for item in CASES)
    rows = [
        ("需求文档", DOC_TITLE),
        ("飞书地址", DOC_URL),
        ("评论来源", COMMENT_IMAGE),
        ("评论补充", "客户：中英文商城注册、页面新建/导入国内外客户、个人转企业、线索转客户、一般/公海/批量分配、认领/批量认领、回收；线索：页面新建、导入、分配/认领及批量、企微好友池、中英文展会、百度/腾讯/抖音等渠道、回收"),
        ("功能范围", "客户/线索到达名下实时企微提醒；客户/线索触发回收当天9点提醒；企微绑定查找；四类消息格式；详情/菜单/首页权限降级；发送失败记录"),
        ("用例状态", "PASS / FAIL / BLOCK / N/A"),
        ("统计", f"合计{len(CASES)}条；阻塞{blocking}条；首轮必测{smoke}条；P0={stats['P0']}条"),
    ]
    for row_no, (key, value) in enumerate(rows, 1):
        meta.cell(row_no, 1, key).font = Font(bold=True)
        meta.cell(row_no, 2, value).alignment = Alignment(vertical="top", wrap_text=True)
    meta.column_dimensions["A"].width = 16
    meta.column_dimensions["B"].width = 120

    write_case_sheet(wb.create_sheet("总测试用例"), CASES)
    for module in MODULES:
        write_case_sheet(wb.create_sheet(module), [c for c in CASES if c["模块"] == module])
    write_case_sheet(wb.create_sheet("首轮冒烟"), [c for c in CASES if c["首轮必测"] == "是"])

    block_ws = wb.create_sheet("阻塞场景清单")
    block_headers = ["用例ID", "模块", "场景", "优先级", "阻塞说明", "失败影响"]
    style_header(block_ws, block_headers, "C00000")
    for row_no, item in enumerate((c for c in CASES if c["是否阻塞"] == "是"), 2):
        prefix = item["用例ID"].split("-", 1)[0]
        values = [
            item["用例ID"], item["模块"], item["场景"], item["优先级"],
            item["备注"] or "核心提醒或权限主链路", BLOCK_IMPACT[prefix],
        ]
        for col, value in enumerate(values, 1):
            block_ws.cell(row_no, col, value).alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in enumerate([14, 24, 40, 10, 40, 36], 1):
        block_ws.column_dimensions[get_column_letter(col)].width = width
    block_ws.freeze_panes = "A2"

    pri = wb.create_sheet("优先级说明")
    pri_rows = [
        ("字段", "定义"), ("P0", "核心触发、接收人、消息数据与权限主流程；首轮执行"),
        ("P1", "重要分支、异常、幂等、性能与数据隔离"), ("P2", "非关键体验与补充边界"),
        ("是否阻塞=是", "失败导致关联提醒或权限测试无意义"),
        ("首轮必测=是", "第一轮必须执行；见“首轮冒烟”"),
        ("用例状态", "PASS/FAIL/BLOCK/N/A；BLOCK=环境或依赖阻塞"),
    ]
    for row in pri_rows:
        pri.append(row)
    style_header(pri, pri_rows[0], "7030A0")
    pri.column_dimensions["A"].width = 20
    pri.column_dimensions["B"].width = 76

    cov = wb.create_sheet("覆盖检查")
    style_header(cov, ["需求点", "覆盖情况", "对应用例ID"])
    for row_no, values in enumerate(COVERAGE, 2):
        for col, value in enumerate(values, 1):
            cov.cell(row_no, col, value).alignment = Alignment(vertical="top", wrap_text=True)
    cov.column_dimensions["A"].width = 56
    cov.column_dimensions["B"].width = 16
    cov.column_dimensions["C"].width = 74
    cov.freeze_panes = "A2"
    wb.save(XLSX_PATH)


def topic(title, children=None):
    node = {"id": uuid.uuid4().hex[:16], "class": "topic", "title": title}
    if children:
        node["children"] = {"attached": children}
    return node


def build_xmind():
    branches = []
    for module in MODULES:
        module_cases = [c for c in CASES if c["模块"] == module]
        children = [
            topic(
                f"{item['用例ID']} {item['场景']}",
                [topic(f"优先级：{item['优先级']}"), topic(f"阻塞：{item['是否阻塞']}｜首轮：{item['首轮必测']}")],
            )
            for item in module_cases
        ]
        branches.append(topic(module, children))
    sheet_id = uuid.uuid4().hex[:16]
    content = [{
        "id": sheet_id, "revisionId": uuid.uuid4().hex[:16], "class": "sheet",
        "title": SHORT_NAME,
        "rootTopic": {
            "id": "root-topic", "class": "topic", "title": f"{SHORT_NAME} 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": branches},
        },
    }]
    metadata = {
        "creator": {"name": "CRMWeComNoticeGenerator", "version": "1.0"},
        "dataStructureVersion": "2", "activeSheetId": sheet_id,
    }
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))


def verify():
    wb = load_workbook(XLSX_PATH, read_only=False)
    required = {
        "需求追溯", "总测试用例", "首轮冒烟", "阻塞场景清单", "优先级说明", "覆盖检查", *MODULES,
    }
    assert required.issubset(wb.sheetnames)
    assert [cell.value for cell in wb["总测试用例"][1]] == HEADERS
    assert wb["总测试用例"].max_row == len(CASES) + 1
    ids = [item["用例ID"] for item in CASES]
    assert len(ids) == len(set(ids))
    with zipfile.ZipFile(XMIND_PATH) as archive:
        assert {"content.json", "manifest.json", "metadata.json"}.issubset(archive.namelist())
        json.loads(archive.read("content.json").decode("utf-8"))


def main():
    build_excel()
    build_xmind()
    verify()
    stats = Counter(item["优先级"] for item in CASES)
    print(f"Generated {len(CASES)} cases: {dict(stats)}")
    print(f"Blocking={sum(c['是否阻塞'] == '是' for c in CASES)}")
    print(f"Smoke={sum(c['首轮必测'] == '是' for c in CASES)}")
    print(XLSX_PATH.resolve())
    print(XMIND_PATH.resolve())


if __name__ == "__main__":
    main()
