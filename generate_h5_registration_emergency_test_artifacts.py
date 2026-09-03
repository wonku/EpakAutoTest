# -*- coding: utf-8 -*-
"""增量更新“中文商城注册流程优化”测试产物：采购商 H5 紧急迭代。"""
from __future__ import annotations

import json
import uuid
import zipfile
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "testcases"
XLSX_PATH = OUT_DIR / "中文商城注册流程优化_TestCases.xlsx"
XMIND_PATH = OUT_DIR / "中文商城注册流程优化_TestMindMap.xmind"
PRD_URL = "https://tvd6quau8vr.feishu.cn/wiki/MEDpwB2uxiq9bukZtFUcZhM3n2d"
HEADERS = [
    "用例ID", "模块", "优先级", "场景", "前置条件", "测试步骤", "预期结果",
    "实际结果", "备注", "用例状态", "是否阻塞", "首轮必测",
]


def case(cid, module, priority, scene, precondition, steps, expected,
         note="紧急迭代：采购商 H5 注册/认证入口补充", blocker="否", smoke="否"):
    return [
        cid, module, priority, scene, precondition, steps, expected,
        "", note, "", blocker, smoke,
    ]


CASES = [
    case("TC-CNREG-059", "14-H5注册页改版", "P0", "采购商 H5 注册入口可打开新版注册页",
         "H5 测试环境可访问；未登录；准备主流移动浏览器",
         "1. 从采购商 H5 登录页点击注册\n2. 观察页面加载及标题\n3. 刷新页面",
         "进入采购商新版注册页；页面无白屏、报错或跳入 PC 布局；刷新后仍停留在有效注册页面",
         blocker="是", smoke="是"),
    case("TC-CNREG-060", "14-H5注册页改版", "P0", "H5 注册页字段和操作与 PC 新版注册页一致",
         "分别打开同环境 PC 与 H5 注册页",
         "1. 对比两端注册字段\n2. 对比步骤条、按钮和辅助入口\n3. 对比字段必填标识",
         "H5 与 PC 均以手机号和验证码完成第一步；仅保留两步且第二步为企业认证；业务字段、必填规则和主要操作一致",
         blocker="是", smoke="是"),
    case("TC-CNREG-061", "14-H5注册页改版", "P0", "H5 注册页不展示已移除字段和旧按钮",
         "打开采购商 H5 新版注册页",
         "1. 检查表单字段\n2. 检查底部按钮\n3. 检查页面顶部入口",
         "不展示登录密码、确认密码、邮箱；不展示“注册并完善资料”“注册并返回首页”和“供应商入驻”；展示“提交注册”",
         smoke="是"),
    case("TC-CNREG-062", "14-H5注册页改版", "P1", "H5 注册页适配常用屏幕尺寸",
         "准备常见小屏、全面屏手机或对应模拟设备",
         "1. 分别打开注册页\n2. 上下滚动并聚焦各输入框\n3. 检查横向溢出和控件遮挡",
         "内容按屏宽自适应；无横向滚动、文字截断、控件重叠；底部操作可见且可点击"),
    case("TC-CNREG-063", "14-H5注册页改版", "P1", "软键盘弹起时可完成注册表单操作",
         "真机打开 H5 注册页",
         "1. 聚焦手机号和验证码输入框\n2. 输入内容并切换字段\n3. 点击提交注册",
         "软键盘不永久遮挡当前输入框及提交按钮；页面可滚动；输入内容不丢失；按钮可触发"),
    case("TC-CNREG-064", "14-H5注册页改版", "P1", "微信注册入口在 H5 的展示与交互符合产品口径",
         "打开采购商 H5 新版注册页",
         "1. 检查微信注册入口\n2. 点击入口\n3. 观察后续授权或扫码交互",
         "微信注册入口及后续交互与最终 H5 设计一致；无无法关闭的 PC 扫码弹窗或失效二维码",
         note="待产品确认：H5 内微信注册采用微信授权还是展示扫码弹窗"),

    case("TC-CNREG-065", "15-H5验证码注册", "P0", "有效手机号可获取验证码并显示倒计时",
         "未注册或允许注册的手机号；短信服务可用",
         "1. 输入有效手机号\n2. 点击获取验证码\n3. 观察按钮和短信",
         "请求成功且手机号收到验证码；按钮进入倒计时，倒计时内不可重复发送",
         blocker="是", smoke="是"),
    case("TC-CNREG-066", "15-H5验证码注册", "P1", "手机号为空或格式错误时不可获取验证码",
         "打开 H5 注册页",
         "1. 手机号留空点击获取验证码\n2. 输入位数或格式错误手机号再次点击",
         "均不发送短信；在手机号字段附近给出明确且可见的校验提示"),
    case("TC-CNREG-067", "15-H5验证码注册", "P0", "正确验证码提交后进入企业认证",
         "已获取当前手机号的有效验证码",
         "1. 输入正确验证码\n2. 点击提交注册\n3. 等待页面跳转",
         "验证码校验通过；账号仅创建一次；进入第二步企业认证页面；登录会话有效",
         blocker="是", smoke="是"),
    case("TC-CNREG-068", "15-H5验证码注册", "P0", "错误验证码提交时提示验证码有误",
         "已获取验证码",
         "1. 输入错误验证码\n2. 点击提交注册",
         "停留在注册第一步；明确提示“验证码有误”或等价文案；不创建账号、不进入企业认证",
         smoke="是"),
    case("TC-CNREG-069", "15-H5验证码注册", "P1", "过期验证码不可完成注册",
         "准备已过期验证码",
         "1. 输入手机号和过期验证码\n2. 点击提交注册",
         "提示验证码过期或无效；不创建账号；允许重新获取验证码"),
    case("TC-CNREG-070", "15-H5验证码注册", "P1", "验证码为空时阻止提交",
         "输入有效手机号但不输入验证码",
         "1. 点击提交注册\n2. 观察字段提示和网络请求",
         "验证码字段显示必填提示；不发起有效注册请求；不跳转"),
    case("TC-CNREG-071", "15-H5验证码注册", "P1", "连续点击提交注册保持幂等",
         "输入有效手机号和正确验证码；网络响应较慢",
         "1. 快速连续点击提交注册\n2. 检查请求、账号和页面结果",
         "仅接受一次有效提交；只创建一个账号；只进入一次企业认证页面"),

    case("TC-CNREG-072", "16-H5企业认证", "P0", "注册成功后企业认证页完整展示认证资料项",
         "通过 H5 验证码完成第一步注册",
         "1. 进入第二步\n2. 检查步骤条、资料字段、上传控件和提交按钮",
         "步骤条当前为企业认证；认证资料项及必填标识与 PC 端一致；控件可在 H5 操作",
         blocker="是", smoke="是"),
    case("TC-CNREG-073", "16-H5企业认证", "P0", "必填认证资料完整时可提交",
         "处于 H5 企业认证页；准备合法企业资料和文件",
         "1. 填写所有必填资料\n2. 完成所需文件上传\n3. 点击提交认证资料",
         "前端和服务端校验通过；资料仅提交一次；提示提交成功",
         blocker="是", smoke="是"),
    case("TC-CNREG-074", "16-H5企业认证", "P0", "缺少必填认证资料时阻止提交",
         "处于 H5 企业认证页",
         "1. 分别缺失一个必填字段或必传文件\n2. 点击提交认证资料",
         "不提交认证申请；对应缺失项给出明确提示并可定位；已填写资料不被清空",
         smoke="是"),
    case("TC-CNREG-075", "16-H5企业认证", "P0", "认证提交成功后 5 秒返回商城",
         "认证资料填写完整且接口可用",
         "1. 提交认证资料\n2. 观察成功提示和倒计时\n3. 等待 5 秒",
         "显示提交成功及返回提示；约 5 秒后自动返回采购商商城页面；用户保持登录态",
         smoke="是"),
    case("TC-CNREG-076", "16-H5企业认证", "P1", "认证资料提交失败时可安全重试",
         "模拟认证提交接口失败或网络中断",
         "1. 填写完整资料并提交\n2. 观察失败反馈\n3. 网络恢复后重试",
         "失败提示明确；不会误显示成功或生成重复申请；已填写资料尽可能保留；恢复后可成功重试"),
    case("TC-CNREG-077", "16-H5企业认证", "P1", "企业认证页返回或刷新后账号与流程状态正确",
         "已完成验证码注册并进入企业认证，尚未提交",
         "1. 刷新页面\n2. 使用浏览器返回后重新进入\n3. 检查账号和步骤状态",
         "不会重复创建账号或退回无效状态；重新进入时能继续认证或按明确规则恢复",
         note="待产品确认：未提交认证资料的草稿是否保存"),

    case("TC-CNREG-078", "17-个人中心认证入口", "P0", "未认证用户在个人中心可见企业认证入口",
         "采购商账号已登录且企业认证状态为未认证",
         "1. 进入 H5 个人中心\n2. 打开账户资料或对应区域\n3. 检查认证入口",
         "展示清晰可点击的“发起企业认证”或等价入口；状态显示为未认证",
         blocker="是", smoke="是"),
    case("TC-CNREG-079", "17-个人中心认证入口", "P0", "个人中心认证入口跳转到企业认证页面",
         "未认证采购商已登录并位于个人中心",
         "1. 点击企业认证入口\n2. 检查目标页和登录状态",
         "进入当前账号的企业认证资料页；登录态和账号上下文正确；可填写并提交认证资料",
         blocker="是", smoke="是"),
    case("TC-CNREG-080", "17-个人中心认证入口", "P1", "认证状态变化后个人中心入口和文案同步",
         "分别准备未认证、审核中、已认证、已驳回账号",
         "1. 各账号登录个人中心\n2. 查看认证状态及可用操作\n3. 刷新后再次检查",
         "各状态展示与后台一致；未认证可发起，审核中不可重复提交，已认证展示认证信息，已驳回可查看原因并按规则重新提交"),
    case("TC-CNREG-081", "17-个人中心认证入口", "P1", "未登录访问个人中心认证地址先完成登录",
         "清除登录态；已知企业认证页面地址",
         "1. 直接访问认证地址\n2. 完成登录\n3. 观察后续跳转",
         "未登录时不能查看或提交他人认证资料；引导登录；登录后进入当前账号可访问的认证页面或按规则返回"),

    case("TC-CNREG-082", "18-下单认证拦截", "P0", "未认证用户下单时出现企业认证提示",
         "未认证采购商已登录；商品可下单",
         "1. 选择商品并触发需求定义的下单动作\n2. 观察拦截提示",
         "下单被拦截；提示需先完成企业认证；提供前往认证和取消/关闭操作；不生成有效订单",
         blocker="是", smoke="是"),
    case("TC-CNREG-083", "18-下单认证拦截", "P0", "下单认证提示确认后跳转企业认证页面",
         "未认证用户已触发下单认证提示",
         "1. 点击前往认证\n2. 检查目标页、账号和商品订单数据",
         "进入当前账号企业认证页面；账号上下文正确；未认证订单未生成；不存在重复订单或脏数据",
         blocker="是", smoke="是"),
    case("TC-CNREG-084", "18-下单认证拦截", "P1", "取消下单认证提示后停留在原业务页面",
         "未认证用户已触发下单认证提示",
         "1. 点击取消或关闭\n2. 检查页面和订单数据",
         "提示关闭并停留原业务页面；商品选择不异常；不生成订单；可再次触发提示"),
    case("TC-CNREG-085", "18-下单认证拦截", "P0", "已认证用户下单不被认证提示拦截",
         "已认证采购商已登录；商品可下单",
         "1. 选择商品\n2. 提交下单\n3. 检查结果",
         "不展示未认证提示；按原下单流程成功提交；订单归属当前账号",
         smoke="是"),
    case("TC-CNREG-086", "18-下单认证拦截", "P1", "审核中和已驳回用户下单按状态口径处理",
         "分别准备审核中、已驳回采购商账号",
         "1. 分别登录并触发下单\n2. 检查提示文案和跳转目标",
         "两种状态均按最终产品规则处理；提示准确区分状态；已驳回用户可进入可修改/重提页面；不误判为已认证",
         note="待产品确认：审核中账号是否允许下单，以及对应提示文案"),
    case("TC-CNREG-087", "18-下单认证拦截", "P1", "多个下单入口执行一致的认证拦截",
         "未认证采购商已登录；商城存在立即购买、购物车结算等入口",
         "1. 分别从各可用下单入口提交\n2. 比较拦截、文案和跳转",
         "所有实际生成订单的入口均执行一致认证校验；无入口可绕过认证直接生成订单",
         note="待产品确认并核对：需纳入拦截的全部下单入口"),

    case("TC-CNREG-088", "19-端间一致性", "P0", "H5 注册账号可在 PC 登录且注册来源一致",
         "通过 H5 新注册采购商账号",
         "1. 在 PC 端使用该手机号登录\n2. 检查客户及注册来源数据",
         "PC 可正常登录；账号、客户归属和注册来源一致；后台标记为自助注册",
         smoke="是"),
    case("TC-CNREG-089", "19-端间一致性", "P0", "企业认证状态在 H5、PC 和后台同步",
         "同一账号已提交或完成企业认证",
         "1. 查看 H5 个人中心\n2. 查看 PC 账户资料\n3. 查看后台客户认证状态",
         "三端对同一账号显示一致的最新认证状态；刷新后无长期旧状态",
         smoke="是"),
    case("TC-CNREG-090", "19-端间一致性", "P1", "认证完成后再次下单不再触发未认证拦截",
         "账号之前因未认证被拦截，现已完成认证",
         "1. 重新进入商品或结算页\n2. 再次提交下单",
         "系统读取最新认证状态；不再出现未认证提示；可继续原下单流程"),
]


COVERAGE = [
    ("采购商 H5 注册页改版并与 PC 业务规则一致", "已覆盖", "TC-CNREG-059~064"),
    ("H5 手机号验证码注册", "已覆盖", "TC-CNREG-065~071"),
    ("H5 注册后企业认证", "已覆盖", "TC-CNREG-072~077"),
    ("登录后个人中心企业认证入口", "已覆盖", "TC-CNREG-078~081"),
    ("未认证状态下单提示并跳转认证", "部分覆盖", "TC-CNREG-082~087"),
    ("H5、PC 与后台数据状态一致", "已覆盖", "TC-CNREG-088~090"),
    ("H5 微信注册交互", "待确认", "TC-CNREG-064"),
    ("审核中账号下单规则及全部下单拦截入口", "待确认", "TC-CNREG-086~087"),
]


def style_header(ws, headers):
    for index, value in enumerate(headers, 1):
        cell = ws.cell(1, index, value)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_case_sheet(ws):
    style_header(ws, HEADERS)
    widths = [17, 22, 8, 38, 34, 46, 46, 16, 38, 12, 10, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(ws.max_row, 2)}"
    red = PatternFill("solid", fgColor="FCE4D6")
    blue = PatternFill("solid", fgColor="DDEBF7")
    for row in ws.iter_rows(min_row=2, max_col=12):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        fill = red if row[10].value == "是" else blue if row[2].value == "P0" else None
        if fill:
            for cell in row:
                cell.fill = fill
    status_col = get_column_letter(HEADERS.index("用例状态") + 1)
    dv = DataValidation(type="list", formula1='"PASS,FAIL,BLOCK,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{status_col}2:{status_col}{max(ws.max_row, 500)}")


def replace_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def build_excel():
    wb = load_workbook(XLSX_PATH)
    total = wb["总测试用例"]
    existing_ids = {total.cell(row, 1).value for row in range(2, total.max_row + 1)}
    for row in CASES:
        if row[0] not in existing_ids:
            total.append(row)
    style_case_sheet(total)

    grouped = OrderedDict()
    for row in CASES:
        grouped.setdefault(row[1], []).append(row)
    for module, rows in grouped.items():
        ws = replace_sheet(wb, module[:31])
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
        style_case_sheet(ws)

    all_rows = list(total.iter_rows(min_row=2, max_col=12, values_only=True))
    trace = wb["需求追溯"] if "需求追溯" in wb.sheetnames else wb.worksheets[0]
    trace.title = "需求追溯"
    trace.delete_rows(1, trace.max_row)
    trace_rows = [
        ("需求文档", "中文商城注册流程优化"),
        ("飞书地址", PRD_URL),
        ("功能范围", "原 PC 注册优化；紧急增量：采购商 H5 注册页、注册后企业认证、个人中心认证入口、未认证下单跳转认证、端间一致性"),
        ("用例状态", "PASS / FAIL / BLOCK / N/A"),
        ("统计", (
            f"合计{len(all_rows)}条；阻塞{sum(r[10] == '是' for r in all_rows)}条；"
            f"首轮必测{sum(r[11] == '是' for r in all_rows)}条；P0={sum(r[2] == 'P0' for r in all_rows)}条"
        )),
        ("本次增量", f"新增{len(CASES)}条：TC-CNREG-059~090"),
        ("需求补充", "采购商 H5 页面改版为与 PC 端一致的注册页面；认证资料入口包含注册页、登录后个人中心、未认证下单提示跳转"),
    ]
    for row in trace_rows:
        trace.append(row)
    trace.column_dimensions["A"].width = 16
    trace.column_dimensions["B"].width = 110
    for row in trace.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in trace["A"]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    smoke = replace_sheet(wb, "首轮冒烟")
    smoke.append(HEADERS)
    for row in all_rows:
        if row[11] == "是":
            smoke.append(row)
    style_case_sheet(smoke)

    blockers = replace_sheet(wb, "阻塞场景清单")
    blocker_headers = ["用例ID", "模块", "场景", "优先级", "阻塞说明", "失败影响"]
    blockers.append(blocker_headers)
    for row in all_rows:
        if row[10] == "是":
            blockers.append([row[0], row[1], row[3], row[2], "核心入口或主链路不可用", "后续关联场景无法有效执行"])
    style_header(blockers, blocker_headers)
    for col, width in enumerate([18, 24, 42, 10, 34, 38], 1):
        blockers.column_dimensions[get_column_letter(col)].width = width
    for row in blockers.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    priority = replace_sheet(wb, "优先级说明")
    priority.append(["字段", "定义"])
    for row in [
        ("P0", "核心功能/主流程/数据正确性；必须首轮执行"),
        ("P1", "重要分支、权限、校验与补偿"),
        ("P2", "边界、UI、非关键异常；回归轮次"),
        ("是否阻塞=是", "失败则后续模块测试无意义"),
        ("首轮必测=是", "第一遍必须执行；见“首轮冒烟”"),
        ("用例状态", "PASS/FAIL/BLOCK/N/A；BLOCK=环境/依赖阻塞"),
    ]:
        priority.append(row)
    style_header(priority, ["字段", "定义"])
    priority.column_dimensions["A"].width = 20
    priority.column_dimensions["B"].width = 70

    coverage = replace_sheet(wb, "覆盖检查")
    coverage.append(["需求点", "覆盖情况", "对应用例"])
    for row in COVERAGE:
        coverage.append(row)
    style_header(coverage, ["需求点", "覆盖情况", "对应用例"])
    coverage.column_dimensions["A"].width = 55
    coverage.column_dimensions["B"].width = 16
    coverage.column_dimensions["C"].width = 32
    for row in coverage.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(XLSX_PATH)


def topic(title, children=None):
    node = {"id": uuid.uuid4().hex[:16], "class": "topic", "title": title}
    if children:
        node["children"] = {"attached": children}
    return node


def build_xmind():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    rows = list(wb["总测试用例"].iter_rows(min_row=2, max_col=12, values_only=True))
    modules = OrderedDict()
    for row in rows:
        modules.setdefault(str(row[1]), []).append(row)
    branches = []
    for module, module_rows in modules.items():
        branches.append(topic(module, [
            topic(f"{row[0]} {row[3]}", [
                topic(f"优先级：{row[2]}"),
                topic(f"阻塞：{row[10]}｜首轮：{row[11]}"),
            ])
            for row in module_rows
        ]))
    sheet_id = uuid.uuid4().hex[:16]
    content = [{
        "id": sheet_id,
        "revisionId": uuid.uuid4().hex[:16],
        "class": "sheet",
        "title": "中文商城注册流程优化",
        "rootTopic": {
            "id": "root-topic",
            "class": "topic",
            "title": "中文商城注册流程优化 测试脑图",
            "structureClass": "org.xmind.ui.map.unbalanced",
            "children": {"attached": branches},
        },
    }]
    metadata = {
        "creator": {"name": "Cursor", "version": "1.0"},
        "dataStructureVersion": "2",
        "activeSheetId": sheet_id,
    }
    manifest = {"file-entries": {"content.json": {}, "metadata.json": {}}}
    with zipfile.ZipFile(XMIND_PATH, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))


def verify():
    wb = load_workbook(XLSX_PATH, read_only=True)
    required = {"需求追溯", "总测试用例", "首轮冒烟", "阻塞场景清单", "优先级说明", "覆盖检查"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Excel 缺少 Sheet: {sorted(missing)}")
    rows = list(wb["总测试用例"].iter_rows(min_row=2, max_col=12, values_only=True))
    ids = [row[0] for row in rows]
    if len(ids) != len(set(ids)):
        raise RuntimeError("存在重复用例 ID")
    if any(cid not in ids for cid in (row[0] for row in CASES)):
        raise RuntimeError("紧急迭代用例未完整写入")
    with zipfile.ZipFile(XMIND_PATH) as archive:
        if set(archive.namelist()) != {"content.json", "metadata.json", "manifest.json"}:
            raise RuntimeError("XMind 文件结构不完整")
        json.loads(archive.read("content.json"))
    print(
        f"Generated {len(rows)} cases; P0={sum(r[2] == 'P0' for r in rows)}; "
        f"blockers={sum(r[10] == '是' for r in rows)}; smoke={sum(r[11] == '是' for r in rows)}"
    )
    print(XLSX_PATH)
    print(XMIND_PATH)


if __name__ == "__main__":
    build_excel()
    build_xmind()
    verify()
